"""
PS Tools — Shared Utilities
Excel helpers, credit logic, and report builder.
"""
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
from shared.config import (
    NAVY, TEAL, WHITE, LTGRAY, MID_GRAY,
    TAG_COLORS, TAG_BADGE,
    PTO_KEYWORDS, UTIL_EXEMPT_EMPLOYEES,
    EMPLOYEE_LOCATION, PS_REGION_OVERRIDE, PS_REGION_MAP,
    AVAIL_HOURS, DEFAULT_SCOPE,
)


def match_ff_task(task_val):
    t = str(task_val).strip()
    return t if t and t.lower() not in ("", "nan", "none") else None

def thin_border():
    s = Side(style="thin", color=MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_fill(hex_color):  return PatternFill("solid", fgColor=hex_color)
def row_fill(hex_color):  return PatternFill("solid", fgColor=hex_color)

GROUP_COLORS = ["EEF2FB", "FFFFFF"]

def group_bg(value, prev_value, group_idx):
    if value != prev_value: group_idx = 1 - group_idx
    return GROUP_COLORS[group_idx], group_idx

def style_header(ws, row, headers, fill_color=NAVY):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = Font(name="Manrope", bold=True, color=WHITE, size=10)
        c.fill      = hdr_fill(fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[row].height = 22

def style_cell(cell, bg, fmt=None, bold=False, align="left"):
    cell.fill      = row_fill(bg)
    cell.font      = Font(name="Manrope", size=10, bold=bold)
    cell.border    = thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: cell.number_format = fmt

def write_title(ws, title, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Manrope", bold=True, size=14, color=WHITE)
    c.fill      = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

# ── Column detection ──────────────────────────────────────────────────────────
def auto_detect_columns(df):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    checks = {
        "employee":      ["employee", "employee name", "name", "resource"],
        "project":       ["project", "project name", "job", "customer:project", "customer: project", "project: name"],
        "project_type":  ["project type", "type", "project_type"],
        "time_item_sku": ["time item sku", "item sku", "sku", "time item"],
        "date":          ["date", "time entry date", "entry date", "work date"],
        "hours":         ["hours", "duration", "hours logged", "time", "qty"],
        "approval":      ["approval status", "approval", "status"],
        "task":          ["case/task/event", "cask/task/event", "task", "case", "event", "memo"],
        "non_billable":  ["non-billable", "non billable", "nonbillable", "non_billable", "is non billable"],
        "billing_type":  ["billing type", "billing_type", "bill type", "billtype"],
        "hours_to_date": ["hours to date", "hours_to_date", "htd", "prior hours", "cumulative hours", "hours booked to date"],
        "project_id":    ["project id", "project_id", "projectid", "project internal id"],
        "region":        ["employee location", "location", "region", "country", "office"],
        "customer_region":  ["customer region", "customer_region", "cust region", "client region"],
        "project_manager":  ["project manager", "project_manager", "pm", "manager"],
        "project_phase":    ["project phase", "phase", "project_phase", "stage"],
        "start_date":       ["start date", "project start date", "start_date", "project start", "commenced"],
    }
    mapping = {}
    unmatched = []
    for standard, candidates in checks.items():
        for candidate in candidates:
            if candidate in cols_lower:
                mapping[cols_lower[candidate]] = standard
                break
        else:
            if standard in ["employee", "project", "project_type", "hours", "non_billable"]:
                unmatched.append(standard)
    return mapping, unmatched

# ── Core credit logic ─────────────────────────────────────────────────────────
def assign_credits(df, scope_map):
    col_map, unmatched = auto_detect_columns(df)
    if unmatched:
        st.warning(f"Could not auto-detect columns: {unmatched}. Check your file headers.")

    df = df.rename(columns=col_map)

    df = df.copy()
    for _c in list(df.columns):
        _dt = str(df[_c].dtype).lower()
        if _dt in ('string', 'str') or 'string' in _dt or 'arrow' in _dt:
            try:
                df[_c] = df[_c].astype(object)
            except Exception:
                df[_c] = df[_c].to_numpy(dtype=object, na_value=None)

    df["non_billable"] = df["non_billable"].fillna("").astype(str).str.strip().str.upper()
    if "project" in df.columns:
        df["project"] = df["project"].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
    # Normalize project_id — strip float suffix e.g. 709470.0 -> 709470
    if "project_id" in df.columns:
        df["project_id"] = df["project_id"].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    df["hours"]  = pd.to_numeric(df["hours"], errors="coerce").fillna(0)
    df["date"]   = pd.to_datetime(df["date"], errors="coerce")
    df["period"] = df["date"].dt.strftime("%Y-%m").fillna("Unknown")
    if "start_date" in df.columns:
        df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
    if "region" not in df.columns:
        df["region"] = pd.array([""] * len(df), dtype=object)

    if "employee" in df.columns:
        _loc_map_norm = {}
        for _k, _v in EMPLOYEE_LOCATION.items():
            _loc_map_norm[_k.lower().strip()] = _v
            _loc_map_norm[" ".join(_k.split()).lower()] = _v

        def _fast_loc(emp_raw):
            emp_n = " ".join(str(emp_raw or "").split()).strip()
            emp_l = emp_n.lower()
            if not emp_l: return ""
            def _extract(v): return v[0] if isinstance(v, tuple) else (v or "")
            if emp_l in _loc_map_norm: return _extract(_loc_map_norm[emp_l])
            for _kl, _vl in _loc_map_norm.items():
                if emp_l.startswith(_kl) or _kl.startswith(emp_l): return _extract(_vl)
                if emp_l.startswith(_kl.split(",")[0].strip()): return _extract(_vl)
            return ""

        _needs_loc = df["region"].fillna("").str.strip() == ""
        if _needs_loc.any():
            df["region"] = df["region"].astype(object)
            df.loc[_needs_loc, "region"] = df.loc[_needs_loc, "employee"].map(_fast_loc)

    _ps_ovr_l = {_k.lower(): _v for _k, _v in PS_REGION_OVERRIDE.items()}
    if "employee" in df.columns:
        def _fast_ps(emp_raw):
            e = " ".join(str(emp_raw or "").split()).lower()
            for _kl, _vl in _ps_ovr_l.items():
                if e == _kl or e.startswith(_kl): return _vl
            return ""
        _emp_ps = df["employee"].map(_fast_ps)
        df["ps_region"] = _emp_ps.astype(object)
        _no_ovr = _emp_ps.fillna("") == ""
        if _no_ovr.any():
            df["ps_region"] = df["ps_region"].astype(object)
            df.loc[_no_ovr, "ps_region"] = df.loc[_no_ovr, "region"].map(
                lambda r: PS_REGION_MAP.get(str(r).strip(), "Other"))
    else:
        df["ps_region"] = df["region"].map(lambda r: PS_REGION_MAP.get(str(r).strip(), "Other"))

    if "date" in df.columns:
        df = df.sort_values(["project", "date"], na_position="first").reset_index(drop=True)

    # ── Pre-compute prior hours per project ───────────────────────────────────
    # Key prior_htd by BOTH project_id and project_name so downstream lookup
    # works regardless of which key _con_key resolves to.
    _htd_col = "hours_to_date"
    prior_htd = {}  # keyed by project_id AND project_name — both point to same value
    if _htd_col in df.columns:
        _pid_col = "project_id" if "project_id" in df.columns else "project"
        for _proj_key, _grp in df.groupby(_pid_col):
            _proj_n = " ".join(str(_grp["project"].iloc[0]).strip().split()) \
                if _pid_col == "project_id" else " ".join(str(_proj_key).strip().split())
            _pid_str = str(_proj_key).strip().replace(".0", "") \
                if _pid_col == "project_id" else None
            try:
                _max_htd    = float(_grp[_htd_col].dropna().astype(float).max() or 0)
                _period_hrs = float(_grp["hours"].dropna().astype(float).sum() or 0)
                _val = max(0.0, _max_htd - _period_hrs)
                # Store under both project name AND project_id so lookup always hits
                prior_htd[_proj_n] = _val
                if _pid_str and _pid_str not in ("nan", "None", ""):
                    prior_htd[_pid_str] = _val
            except Exception:
                prior_htd[_proj_n] = 0.0
                if _pid_str and _pid_str not in ("nan", "None", ""):
                    prior_htd[_pid_str] = 0.0

    consumed = {}
    credit_hrs_list   = []
    variance_hrs_list = []
    credit_tag_list   = []
    notes_list        = []
    htd_start_list     = []  # track starting HTD per row for output
    consumed_after_list = []  # track running consumed position after each row

    _proj_idx  = df.columns.get_loc("project")       if "project"           in df.columns else None
    _ptype_idx = df.columns.get_loc("project_type")  if "project_type"      in df.columns else None
    _hrs_idx   = df.columns.get_loc("hours")         if "hours"             in df.columns else None
    _nb_idx    = df.columns.get_loc("non_billable")  if "non_billable"      in df.columns else None
    _bt_idx    = df.columns.get_loc("billing_type")  if "billing_type"      in df.columns else None
    _sku_idx   = df.columns.get_loc("time_item_sku") if "time_item_sku"     in df.columns else None
    _drs_idx   = df.columns.get_loc("_drs_project_name") if "_drs_project_name" in df.columns else None
    _pid_idx   = df.columns.get_loc("project_id")    if "project_id"        in df.columns else None

    for _tup in df.itertuples(index=False, name=None):
        proj      = " ".join(str(_tup[_proj_idx]  if _proj_idx  is not None else "").split())
        ptype     = str(_tup[_ptype_idx] if _ptype_idx is not None else "").strip()
        hrs       = float(_tup[_hrs_idx]  if _hrs_idx  is not None else 0)
        nb        = str(_tup[_nb_idx]    if _nb_idx   is not None else "NO").strip().upper()
        bill_type = str(_tup[_bt_idx]    if _bt_idx   is not None else "").strip().lower()
        is_tm     = bill_type == "t&m"

        if hrs <= 0:
            credit_hrs_list.append(0); variance_hrs_list.append(0)
            credit_tag_list.append("SKIPPED"); notes_list.append("Zero or missing hours")
            htd_start_list.append(0)
            consumed_after_list.append(0.0)
            continue

        if bill_type == "internal":
            credit_hrs_list.append(0); variance_hrs_list.append(0)
            credit_tag_list.append("NON-BILLABLE"); notes_list.append("Internal: excluded from utilization")
            htd_start_list.append(0)
            consumed_after_list.append(0.0)
            continue

        if is_tm:
            credit_hrs_list.append(hrs); variance_hrs_list.append(0)
            credit_tag_list.append("CREDITED"); notes_list.append("T&M: full credit")
            htd_start_list.append(0)
            consumed_after_list.append(0.0)
            continue

        _ptype_lower = ptype.strip().lower()
        if "premium" in _ptype_lower:
            _sku = str(_tup[_sku_idx] if _sku_idx is not None else "") or ""
            _sku_nums = _re_constants.findall(r"IMPL(\d+)", _sku.upper())
            if _sku_nums:
                scope_hrs = float(_sku_nums[0])
            else:
                _proj_name_scope = str(_tup[_drs_idx] if _drs_idx is not None else "") or str(_tup[_proj_idx] if _proj_idx is not None else "") or ""
                _name_nums = _re_constants.findall(r"(?<![\d])(10|20)(?![\d])", _proj_name_scope)
                scope_hrs = float(_name_nums[0]) if _name_nums else None
        else:
            _matches = [(k, float(v)) for k, v in scope_map.items()
                        if k.strip().lower() in _ptype_lower]
            scope_hrs = max(_matches, key=lambda x: len(x[0]))[1] if _matches else None

        if scope_hrs is None:
            credit_hrs_list.append(0); variance_hrs_list.append(hrs)
            credit_tag_list.append("UNCONFIGURED"); notes_list.append(f"Fixed Fee but no scope defined for: {ptype}")
            htd_start_list.append(0)
            consumed_after_list.append(0.0)
            continue

        # Key on project_id if available
        _raw_pid = str(_tup[_pid_idx] if _pid_idx is not None else "").strip().replace(".0", "")
        _con_key = _raw_pid if _raw_pid and _raw_pid not in ("nan", "None", "") else proj

        if _con_key not in consumed:
            consumed[_con_key] = prior_htd.get(_con_key, prior_htd.get(proj, 0.0))

        already   = consumed[_con_key]
        remaining = scope_hrs - already
        htd_start_list.append(already)

        if remaining <= 0:
            credit_hrs_list.append(0); variance_hrs_list.append(hrs)
            credit_tag_list.append("OVERRUN"); notes_list.append(f"Scope exhausted (cap: {scope_hrs:.0f}h)")
            consumed[_con_key] = already + hrs
            consumed_after_list.append(float(already))
        elif hrs <= remaining:
            consumed[_con_key] = already + hrs
            credit_hrs_list.append(hrs); variance_hrs_list.append(0)
            credit_tag_list.append("CREDITED"); notes_list.append(f"NB within scope ({already:.1f}/{scope_hrs:.0f}h used)")
            consumed_after_list.append(float(already))
        else:
            consumed[_con_key] = already + remaining
            credit_hrs_list.append(remaining); variance_hrs_list.append(hrs - remaining)
            credit_tag_list.append("PARTIAL")
            notes_list.append(f"Split: {remaining:.2f}h credited / {hrs - remaining:.2f}h overrun")
            consumed_after_list.append(float(already))

    for _col in df.select_dtypes(include="string").columns:
        df[_col] = df[_col].astype(object)

    df["credit_hrs"]   = credit_hrs_list
    df["variance_hrs"] = variance_hrs_list
    df["credit_tag"]   = credit_tag_list
    df["notes"]        = notes_list
    df["htd_start"]    = htd_start_list
    df["previous_htd"] = [float(v) for v in consumed_after_list]

    if "task" in df.columns:
        df["ff_task"] = df["task"].fillna("").apply(match_ff_task)
    else:
        df["ff_task"] = ""

    skipped_df = df[df["credit_tag"] == "SKIPPED"][
        [c for c in ["employee","project","project_type","billing_type","date","hours","notes"] if c in df.columns]
    ].copy()

    return df, consumed, skipped_df


# ── Excel builder ─────────────────────────────────────────────────────────────
def _xl_val(val):
    if val is None: return ""
    if isinstance(val, tuple): return val[0] if val else ""
    try:
        if pd.isna(val): return ""
    except (TypeError, ValueError): pass
    return val

def build_excel(df, scope_map, consumed):
    wb  = Workbook()
    wb.remove(wb.active)
    bgs = [WHITE, LTGRAY]

    ws = wb.create_sheet("PROCESSED_DATA")
    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = "A3"

    headers = ["Employee","Location","Customer Region","Project Manager","Project",
               "Project Type","Billing Type","Hrs to Date","Date","Hours Logged",
               "Approval","Task/Case","Non-Billable","Credit Hrs","Variance Hrs",
               "Previous Hrs to Date","Credit Tag","Period","Notes",
               "Project Phase","Start Date","Days Active","Scoped Hrs","Variance Flag"]
    widths  = [20,16,18,20,35,20,14,13,14,13,14,25,13,12,12,18,16,12,45,16,14,12,12,16]
    cols    = ["employee","region","customer_region","project_manager","project",
               "project_type","billing_type","hours_to_date","date","hours",
               "approval","task","non_billable","credit_hrs","variance_hrs",
               "previous_htd","credit_tag","period","notes",
               "project_phase","start_date_display","days_active","_scoped_hrs","_variance_flag"]

    write_title(ws, "PROCESSED DATA — Utilization Credit Detail", len(headers))
    style_header(ws, 2, headers, TEAL)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def _get_scoped(ptype):
        _m = [(k, float(v)) for k, v in scope_map.items()
              if k.strip().lower() in str(ptype).strip().lower()]
        return max(_m, key=lambda x: len(x[0]))[1] if _m else None

    def _variance_flag(tag, scoped):
        t = str(tag).strip().upper()
        if t in ("SKIPPED", "NON-BILLABLE", "CREDITED"): return "N/A"
        if t == "UNCONFIGURED": return "No Scope Set"
        if t in ("OVERRUN", "PARTIAL"): return "True Overrun" if scoped and scoped > 0 else "No Scope Set"
        return "Within Budget"

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        tag     = str(row.get("credit_tag","")).strip()
        bg      = TAG_COLORS.get(tag, "F2F2F2")
        _scoped = _get_scoped(row.get("project_type",""))
        _vflag  = _variance_flag(tag, _scoped)
        _row_ext = dict(row)
        _row_ext["_scoped_hrs"]    = round(_scoped, 2) if _scoped is not None else ""
        _row_ext["_variance_flag"] = _vflag
        for c_idx, col in enumerate(cols, 1):
            val  = _xl_val(_row_ext.get(col, ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            fmt, bold, align = None, False, "left"
            if col == "date" and pd.notna(val): fmt = "YYYY-MM-DD"; align = "center"
            elif col in ("hours","credit_hrs","variance_hrs","hours_to_date","previous_htd"):
                fmt = "#,##0.00"; align = "right"
            elif col == "credit_tag": bold = True; align = "center"
            elif col in ("period","billing_type","region"): align = "center"
            elif col == "_scoped_hrs": fmt = "#,##0.00"; align = "right"
            elif col == "_variance_flag":
                align = "center"; bold = True
                val_str = str(val)
                if val_str == "True Overrun": bg = "FFC7CE"
                elif val_str == "No Scope Set": bg = "FFEB9C"
                elif val_str == "Within Budget": bg = "C6EFCE"
                else: bg = TAG_COLORS.get(tag, "F2F2F2")
            style_cell(cell, bg, fmt=fmt, bold=bold, align=align)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    ws2 = wb.create_sheet("SUMMARY - By Employee")
    ws2.sheet_properties.tabColor = NAVY
    ws2.freeze_panes = "A3"

    eh = ["Employee","Location","PS Region","Period",
          "Avail Hrs (Capacity)","Hours This Period","Utilization Credits",
          "FF Project Overrun Hrs","Admin Hrs",
          "Util % (vs Hours Logged)","Util % (vs Capacity)",
          "Project Util %","Gap (pts)","Projected Full Month Util %"]
    ew = [22,16,14,12,16,15,18,18,14,20,20,16,12,22]
    write_title(ws2, "SUMMARY — Utilization by Employee", len(eh))
    style_header(ws2, 2, eh, TEAL)
    ws2.auto_filter.ref = "A2:L2"
    for i, w in enumerate(ew, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    emp_region = {}
    emp_cust_region = {}
    emp_pm = {}
    if "region" in df.columns:
        emp_region = df.dropna(subset=["region"]).groupby("employee")["region"].first().to_dict()
    if "customer_region" in df.columns:
        emp_cust_region = df.dropna(subset=["customer_region"]).groupby("employee")["customer_region"].first().to_dict()
    if "project_manager" in df.columns:
        emp_pm = df.dropna(subset=["project_manager"]).groupby("employee")["project_manager"].first().to_dict()

    admin_hrs = df[df["billing_type"].str.lower() == "internal"].groupby(
        ["employee","period"])["hours"].sum().reset_index().rename(columns={"hours":"admin_hrs"})

    emp_sum = df[df["credit_tag"] != "SKIPPED"].groupby(
        ["employee","period"], as_index=False
    ).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        ff_overrun_hrs=("variance_hrs","sum"),
    ).sort_values(["employee","period"])
    emp_sum = emp_sum.merge(admin_hrs, on=["employee","period"], how="left")
    emp_sum["admin_hrs"] = emp_sum["admin_hrs"].fillna(0)

    _pto_lookup = {}
    _pto_task_col = None
    for _tc in ["task", "case_task_event"]:
        if _tc in df.columns and not df[_tc].fillna("").str.strip().eq("").all():
            _pto_task_col = _tc; break
    if _pto_task_col:
        _pto_mask = df[_pto_task_col].fillna("").str.lower().apply(
            lambda t: any(k in t for k in PTO_KEYWORDS))
        _pto_df = df[_pto_mask].groupby(["employee","period"])["hours"].sum()
        _pto_lookup = {(emp, per): hrs for (emp, per), hrs in _pto_df.items()}

    _period_days = None
    _total_period_days = None
    if "date" in df.columns:
        _dates = df["date"].dropna()
        if len(_dates) > 0:
            import calendar
            _min_d = _dates.min(); _max_d = _dates.max()
            _period_days = len(pd.bdate_range(_min_d, _max_d))
            _yr, _mo = _min_d.year, _min_d.month
            _month_start = pd.Timestamp(_yr, _mo, 1)
            _month_end   = pd.Timestamp(_yr, _mo, calendar.monthrange(_yr, _mo)[1])
            _total_period_days = len(pd.bdate_range(_month_start, _month_end))

    _prev_emp = None; _grp_idx = 0
    for r_idx, (_, row) in enumerate(emp_sum.iterrows(), 3):
        emp    = row["employee"]
        period = row["period"]
        region = emp_region.get(emp, "")
        avail  = get_avail_hours(region, period) if region else None
        util   = row["credit_hrs"] / row["hours_this_period"] if row["hours_this_period"] > 0 else 0
        bg, _grp_idx = group_bg(emp, _prev_emp, _grp_idx)
        _prev_emp = emp
        util_bg = ("EAF9F1" if util >= 0.8 else "FEF9E7" if util >= 0.6 else "FDECED")
        _ps_reg = df[df["employee"]==emp]["ps_region"].iloc[0] if len(df[df["employee"]==emp]) > 0 else ""
        cap_util = row["credit_hrs"] / avail if avail and avail > 0 else None
        proj_util_pct = (row["credit_hrs"] + row["ff_overrun_hrs"]) / avail if avail and avail > 0 else None
        gap_pts = (proj_util_pct - cap_util) if proj_util_pct is not None and cap_util is not None else None
        proj_util = None
        if _period_days is not None and _total_period_days is not None and _period_days < _total_period_days:
            daily_rate = row["credit_hrs"] / _period_days if _period_days > 0 else 0
            proj_util = (daily_rate * _total_period_days) / avail if avail and avail > 0 else None

        vals = [emp, region, _ps_reg, period, avail or "—",
                row["hours_this_period"], row["credit_hrs"], row["ff_overrun_hrs"], row.get("admin_hrs", 0),
                util if row["hours_this_period"] > 0 else "—",
                cap_util if cap_util is not None else "—",
                proj_util_pct if proj_util_pct is not None else "—",
                gap_pts if gap_pts is not None else "—",
                proj_util if proj_util is not None else "—"]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%","0.0%","0.0%","+0.0%;-0.0%;-","0.0%"]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=_xl_val(val))
            is_util_col = c_idx in (10, 11, 12)
            if c_idx == 13 and isinstance(val, float):
                gap_bg = ("FCE4D6" if val >= 0.20 else "FFF2CC" if val >= 0.10 else bg)
                style_cell(cell, gap_bg, fmt=fmt, bold=(val >= 0.10), align="right")
            else:
                style_cell(cell, util_bg if is_util_col else bg, fmt=fmt,
                           align="right" if c_idx > 4 else "center" if c_idx == 4 else "left")

    ws3 = wb.create_sheet("FF By Project Utilization")
    ws3.sheet_properties.tabColor = "E67E22"
    ws3.freeze_panes = "A3"

    ph = ["Project","Project Type","Project Manager",
          "Scoped Hrs","Previous Hrs to Date","Hours This Period","Credit Hrs",
          "FF Project Overrun Hrs","Hours to Date","Hours Balance","Burn %","Status"]
    pw = [35,22,22,12,15,15,12,18,18,16,10,12]
    write_title(ws3, "SUMMARY — Utilization by Project (Fixed Fee)", len(ph))
    style_header(ws3, 2, ph, TEAL)
    ws3.auto_filter.ref = "A2:L2"
    for i, w in enumerate(pw, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ff_proj_df = df[
        (df["credit_tag"] != "SKIPPED") &
        (df["billing_type"].str.lower() == "fixed fee")
    ] if "billing_type" in df.columns else df[df["credit_tag"] != "SKIPPED"]

    proj_sum = ff_proj_df.groupby(
        ["project","project_type"], as_index=False
    ).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        variance_hrs=("variance_hrs","sum"),
        htd_start=("htd_start","first"),
    ).sort_values("project")

    htd_seeds = dict(zip(proj_sum["project"], proj_sum["htd_start"]))
    proj_cust_region = {}
    proj_pm = {}
    if "customer_region" in df.columns:
        proj_cust_region = df.dropna(subset=["customer_region"]).groupby("project")["customer_region"].first().to_dict()
    if "project_manager" in df.columns:
        proj_pm = df.dropna(subset=["project_manager"]).groupby("project")["project_manager"].first().to_dict()
    proj_ps_region = df.groupby("project")["ps_region"].first().to_dict() if "ps_region" in df.columns else {}
    proj_phase = {}
    if "project_phase" in df.columns:
        proj_phase = df.dropna(subset=["project_phase"]).groupby("project")["project_phase"].first().to_dict()
    proj_start = {}
    if "start_date" in df.columns:
        proj_start = df.dropna(subset=["start_date"]).groupby("project")["start_date"].min().to_dict()
    _as_of = pd.to_datetime(df["date"], errors="coerce").max() if "date" in df.columns else pd.Timestamp.now()
    df["project_phase"] = df["project"].map(proj_phase) if proj_phase else ""
    if proj_start:
        df["start_date_mapped"] = df["project"].map(proj_start)
        df["days_active"] = df["start_date_mapped"].apply(
            lambda s: int((_as_of - s).days) if pd.notna(s) and pd.notna(_as_of) else None)
        df["start_date_display"] = df["start_date_mapped"]
    else:
        df["start_date_display"] = df.get("start_date", None)
        df["days_active"] = None

    _prev_ptype = None; _grp_idx_p = 0
    for r_idx, (_, row) in enumerate(proj_sum.iterrows(), 3):
        ptype   = str(row["project_type"]).strip()
        _pm     = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in ptype.lower()]
        scope_h = max(_pm, key=lambda x: len(x[0]))[1] if _pm else 0
        seed      = float(row["htd_start"]) if row["htd_start"] else 0
        previous_h = max(0.0, seed - row["hours_this_period"])
        burn = seed / scope_h if scope_h > 0 else 0
        vari_h   = row["variance_hrs"]
        htd_total = previous_h + row["hours_this_period"]
        if scope_h > 0 and htd_total > scope_h:   status = "OVERRUN"
        elif scope_h > 0 and htd_total == scope_h: status = "AT LIMIT"
        elif scope_h == 0 and htd_total > 0:       status = "REVIEW"
        elif burn >= 0.9:                          status = "REVIEW"
        elif burn > 0:                             status = "ON TRACK"
        else:                                      status = "—"

        status_bg = {"OVERRUN":"FDECED","AT LIMIT":"FDECED","REVIEW":"FEF9E7","ON TRACK":"EAF9F1"}.get(status, LTGRAY)
        bg, _grp_idx_p = group_bg(ptype, _prev_ptype, _grp_idx_p)
        _prev_ptype = ptype
        pm_name = proj_pm.get(row["project"], "")
        start_dt = proj_start.get(row["project"])
        vals = [row["project"], ptype, pm_name, scope_h or "—", previous_h,
                row["hours_this_period"], row["credit_hrs"], vari_h,
                previous_h + row["hours_this_period"],
                (previous_h + row["hours_this_period"]) - scope_h if scope_h > 0 else "—",
                burn if scope_h > 0 else "—", status]
        fmts = [None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%",None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=_xl_val(val))
            style_cell(cell, status_bg if c_idx == 12 else bg, fmt=fmt, bold=(c_idx == 12),
                       align="right" if c_idx in (5,6,7,8,9,10,11) else "center" if c_idx == 12 else "left")

    ws_ot = wb.create_sheet("FF Overrun by Project Type")
    ws_ot.sheet_properties.tabColor = "8E44AD"
    ws_ot.freeze_panes = "A3"
    oth = ["Project Type","# Projects","# Over Budget","% Over Budget",
           "Total Overrun Hrs","Avg Scoped Hrs","Avg Actual Hrs","Avg +/− vs Scope"]
    otw = [28,11,14,14,17,14,14,17]
    write_title(ws_ot, "FIXED FEE OVERRUN BY PROJECT TYPE", len(oth))
    style_header(ws_ot, 2, oth, TEAL)
    ws_ot.auto_filter.ref = "A2:H2"
    for i, w in enumerate(otw, 1):
        ws_ot.column_dimensions[get_column_letter(i)].width = w

    _ot_ff = df[
        (df["credit_tag"] != "SKIPPED") &
        (df.get("billing_type", pd.Series(dtype="object")).str.lower() == "fixed fee")
    ].copy() if "billing_type" in df.columns else df[df["credit_tag"] != "SKIPPED"].copy()

    _ot_proj = _ot_ff.groupby(["project","project_type"], as_index=False).agg(
        hours_total=("hours","sum"), overrun_hrs=("variance_hrs","sum"))
    def _ot_scope(ptype):
        _m = [(k, float(v)) for k, v in scope_map.items()
              if k.strip().lower() in str(ptype).strip().lower()]
        return max(_m, key=lambda x: len(x[0]))[1] if _m else None
    _ot_proj["scoped_hrs"] = _ot_proj["project_type"].apply(_ot_scope)
    _ot_proj = _ot_proj[_ot_proj["scoped_hrs"].notna()].copy()
    _ot_proj["over_under"] = _ot_proj["hours_total"].astype(float) - _ot_proj["scoped_hrs"].astype(float)
    _ot_proj["is_over"] = (_ot_proj["over_under"] > 0).astype(int)
    _ot_type = _ot_proj.groupby("project_type", as_index=False).agg(
        total_projects=("project","count"), over_count=("is_over","sum"),
        total_overrun_hrs=("overrun_hrs","sum"), avg_scoped=("scoped_hrs","mean"),
        avg_actual=("hours_total","mean"), avg_over_under=("over_under","mean"),
    ).sort_values("total_overrun_hrs", ascending=False)
    _ot_type["over_pct"] = _ot_type["over_count"] / _ot_type["total_projects"]
    _ot_totals_proj  = int(_ot_type["total_projects"].sum())
    _ot_totals_over  = int(_ot_type["over_count"].sum())
    _ot_totals_ovhrs = float(_ot_type["total_overrun_hrs"].sum())

    for r_idx_ot, (_, row_ot) in enumerate(_ot_type.iterrows(), 3):
        bg_ot = GROUP_COLORS[r_idx_ot % 2]
        avg_ou = float(row_ot["avg_over_under"])
        over_pct = float(row_ot["over_pct"])
        vals_ot = [row_ot["project_type"], int(row_ot["total_projects"]), int(row_ot["over_count"]),
                   over_pct, float(row_ot["total_overrun_hrs"]),
                   float(row_ot["avg_scoped"]), float(row_ot["avg_actual"]), avg_ou]
        fmts_ot = [None,"#,##0","#,##0","0.0%","#,##0.0","0.0","0.0","+0.0;-0.0;\"-\""]
        for c_idx_ot, (val_ot, fmt_ot) in enumerate(zip(vals_ot, fmts_ot), 1):
            cell_ot = ws_ot.cell(row=r_idx_ot, column=c_idx_ot, value=val_ot)
            if c_idx_ot == 4:
                if over_pct >= 0.40:
                    style_cell(cell_ot, "FCE4D6", fmt=fmt_ot, bold=True, align="right")
                    cell_ot.font = Font(name="Manrope", size=10, bold=True, color="9C2A00")
                elif over_pct >= 0.25:
                    style_cell(cell_ot, "FFF2CC", fmt=fmt_ot, bold=True, align="right")
                    cell_ot.font = Font(name="Manrope", size=10, bold=True, color="7F4F00")
                else:
                    style_cell(cell_ot, bg_ot, fmt=fmt_ot, align="right")
            elif c_idx_ot == 8:
                style_cell(cell_ot, bg_ot, fmt=fmt_ot, bold=(avg_ou > 0), align="right")
                cell_ot.font = Font(name="Manrope", size=10, bold=(avg_ou>0),
                                    color="9C2A00" if avg_ou > 0 else "595959")
            elif c_idx_ot == 5 and float(row_ot["total_overrun_hrs"]) >= 500:
                style_cell(cell_ot, bg_ot, fmt=fmt_ot, bold=True, align="right")
                cell_ot.font = Font(name="Manrope", size=10, bold=True, color="9C2A00")
            else:
                style_cell(cell_ot, bg_ot, fmt=fmt_ot, align="left" if c_idx_ot == 1 else "right")
        ws_ot.row_dimensions[r_idx_ot].height = 15

    _ot_tr = 3 + len(_ot_type)
    for c_idx_ot, (val_ot, fmt_ot) in enumerate(zip(
        ["Total", _ot_totals_proj, _ot_totals_over, None, _ot_totals_ovhrs, None, None, None],
        [None, "#,##0", "#,##0", None, "#,##0.0", None, None, None]), 1):
        cell_ot = ws_ot.cell(row=_ot_tr, column=c_idx_ot, value=val_ot)
        cell_ot.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
        cell_ot.fill  = PatternFill("solid", fgColor=NAVY)
        cell_ot.border = thin_border()
        cell_ot.alignment = Alignment(horizontal="left" if c_idx_ot == 1 else "right", vertical="center")
        if fmt_ot and val_ot is not None: cell_ot.number_format = fmt_ot
    ws_ot.row_dimensions[_ot_tr].height = 18

    ws5 = wb.create_sheet("Non-Billable")
    ws5.sheet_properties.tabColor = "95A5A6"
    ws5.freeze_panes = "A3"
    znh = ["Task / Activity","Employee","Period","Hours"]
    znw = [35,22,12,12]
    write_title(ws5, "NON-BILLABLE — Hours by Employee by Activity", len(znh))
    style_header(ws5, 2, znh, TEAL)
    ws5.auto_filter.ref = "A2:E2"
    for i, w in enumerate(znw, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    if "billing_type" in df.columns:
        zco_df = df[df["billing_type"].fillna("").str.lower() == "internal"].copy()
    else:
        zco_df = df[df["credit_tag"] == "NON-BILLABLE"].copy()

    if len(zco_df) > 0:
        if "task" not in zco_df.columns:
            zco_df["task"] = "Internal (no task)"
        else:
            zco_df["task"] = zco_df["task"].fillna("").replace("", "Internal (no task)")

    if "task" in zco_df.columns and len(zco_df) > 0:
        zco_sum = zco_df.groupby(["task","employee","period"], as_index=False
        ).agg(hours=("hours","sum")).sort_values(["task","employee","period"])
        emp_period_totals = df[df["credit_tag"] != "SKIPPED"].groupby(
            ["employee","period"])["hours"].sum().to_dict()
        znh[3] = "Hours"
        ws5.cell(row=2, column=5, value="% of Total Hrs").font = Font(name="Manrope", bold=True, color=WHITE, size=10)
        ws5.cell(row=2, column=5).fill = hdr_fill(TEAL)
        ws5.cell(row=2, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws5.cell(row=2, column=5).border = thin_border()
        ws5.column_dimensions["E"].width = 16
        _prev_task_z = None; _grp_idx_z = 0; r_idx = 3
        for _, row in zco_sum.iterrows():
            task = row.get("task","")
            if task != _prev_task_z:
                _task_hrs = zco_sum[zco_sum["task"]==task]["hours"].sum()
                for ci, (hval, hfmt) in enumerate([(task, None), ("", None), ("", None), (_task_hrs, "#,##0.00"), ("", None)], 1):
                    hcell = ws5.cell(row=r_idx, column=ci, value=hval)
                    hcell.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
                    hcell.fill  = PatternFill("solid", fgColor=NAVY)
                    hcell.border = thin_border()
                    hcell.alignment = Alignment(horizontal="right" if ci==4 else "left", vertical="center")
                    if hfmt: hcell.number_format = hfmt
                r_idx += 1; _prev_task_z = task; _grp_idx_z = 0
            bg, _grp_idx_z = group_bg(task, task, _grp_idx_z)
            total_hrs = emp_period_totals.get((row["employee"], row["period"]), 0)
            pct = row["hours"] / total_hrs if total_hrs > 0 else 0
            vals = ["", row["employee"], row["period"], row["hours"], pct]
            fmts = [None, None, None, "#,##0.00", "0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=val)
                style_cell(cell, bg, fmt=fmt, align="right" if c_idx in (4,5) else "center" if c_idx == 3 else "left")
            r_idx += 1
    else:
        ws5.cell(row=3, column=1, value="No Non-Billable (Internal) entries in this period.")

    ws6 = wb.create_sheet("Task Analysis")
    ws6.sheet_properties.tabColor = "27AE60"
    ws6.freeze_panes = "A3"
    tah = ["Task Category","Project Type","Hours This Period","Avg Hrs / Project","% of Type Hrs"]
    taw = [28,25,16,18,16]
    write_title(ws6, "TASK ANALYSIS — Hours by Task › Project Type", len(tah))
    style_header(ws6, 2, tah, TEAL)
    ws6.auto_filter.ref = "A2:E2"
    for i, w in enumerate(taw, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    ff_df = df[(df["billing_type"].str.lower() == "fixed fee") & (df["ff_task"].notna())].copy() \
        if "billing_type" in df.columns else df[df["ff_task"].notna()].copy()

    if len(ff_df) > 0:
        task_sum = ff_df.groupby(["ff_task","project_type"], as_index=False).agg(hours=("hours","sum")).sort_values(["ff_task","project_type"])
        all_ff = df[df["billing_type"].str.lower() == "fixed fee"] if "billing_type" in df.columns else df
        proj_count_by_type = all_ff.groupby("project_type")["project"].nunique().to_dict()
        type_totals = ff_df.groupby("project_type")["hours"].sum().to_dict()
        _prev_task_t = None; _grp_idx_t = 0; r_idx_t = 3
        for _, row in task_sum.iterrows():
            ff_task = row["ff_task"]
            type_total = type_totals.get(row["project_type"], 0)
            pct = row["hours"] / type_total if type_total > 0 else 0
            proj_cnt = proj_count_by_type.get(row["project_type"], 1)
            avg_hrs = round((row["hours"] / proj_cnt if proj_cnt > 0 else 0) * 4) / 4
            if ff_task != _prev_task_t:
                _task_total_hrs = task_sum[task_sum["ff_task"]==ff_task]["hours"].sum()
                for ci, (hval, hfmt) in enumerate([(ff_task, None), ("— ALL TYPES —", None), (_task_total_hrs, "#,##0.00"), ("", None), ("", None)], 1):
                    hcell = ws6.cell(row=r_idx_t, column=ci, value=hval)
                    hcell.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
                    hcell.fill  = PatternFill("solid", fgColor=NAVY)
                    hcell.border = thin_border()
                    hcell.alignment = Alignment(horizontal="right" if ci==3 else "left", vertical="center")
                    if hfmt: hcell.number_format = hfmt
                r_idx_t += 1; _prev_task_t = ff_task; _grp_idx_t = 0
            task_colors = {"Configuration":"EBF5FB","Post Go-Live Consulting":"FEF9E7",
                           "Project Management":"F4ECF7","Training & UAT":"EAF9F1","Customer Communication":"FDF2F8"}
            bg, _grp_idx_t = group_bg(ff_task, ff_task, _grp_idx_t)
            task_bg = task_colors.get(ff_task, bg)
            vals = ["", row["project_type"], row["hours"], avg_hrs, pct]
            fmts = [None, None, "#,##0.00", "#,##0.00", "0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws6.cell(row=r_idx_t, column=c_idx, value=val)
                style_cell(cell, task_bg if c_idx > 1 else bg, fmt=fmt, align="right" if c_idx > 2 else "left")
            r_idx_t += 1
    else:
        ws6.cell(row=3, column=1, value="No Fixed Fee task data found. Check Billing Type and Task/Case columns.")

    ws_pc = wb.create_sheet("Project Count")
    ws_pc.sheet_properties.tabColor = "2980B9"
    ws_pc.freeze_panes = "A3"
    pch = ["Project Type","Billing Type","Project Count"]
    pcw = [35, 14, 14]
    write_title(ws_pc, "PROJECT COUNT — Distinct Projects by Type (excl. Internal)", len(pch))
    style_header(ws_pc, 2, pch, TEAL)
    ws_pc.auto_filter.ref = "A2:D2"
    for i, w in enumerate(pcw, 1):
        ws_pc.column_dimensions[get_column_letter(i)].width = w
    pc_df = df[df["billing_type"].str.lower() != "internal"].copy() if "billing_type" in df.columns else df.copy()
    pc_sum = pc_df.groupby(["project_type","billing_type"], as_index=False).agg(
        project_count=("project","nunique")).sort_values(["project_type","billing_type"])
    grand_total = pc_sum["project_count"].sum()
    for r_idx, (_, row) in enumerate(pc_sum.iterrows(), 3):
        bg = LTGRAY if r_idx % 2 == 0 else WHITE
        vals = [row["project_type"], row["billing_type"], row["project_count"]]
        fmts = [None, None, "#,##0"]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_pc.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, bg, fmt=fmt, align="center" if c_idx == 2 else "right" if c_idx == 3 else "left")
    total_row = r_idx + 1 if len(pc_sum) > 0 else 3
    for c_idx, (val, fmt, bold) in enumerate([("Grand Total", None, True), ("", None, False), (grand_total, "#,##0", True)], 1):
        cell = ws_pc.cell(row=total_row, column=c_idx, value=val)
        cell.font = Font(name="Manrope", bold=True, size=10, color=WHITE)
        cell.fill = hdr_fill(NAVY); cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right" if c_idx == 2 else "left", vertical="center")
        if fmt: cell.number_format = fmt

    ws_pta = wb.create_sheet("FF Project Type Analysis")
    ws_pta.sheet_properties.tabColor = "8E44AD"
    ws_pta.freeze_panes = "A4"
    ptah = ["Project Type","Task Category","Hours This Period","Avg Hrs / Project","% of Type Hrs"]
    ptaw = [28,25,16,18,16]
    write_title(ws_pta, "FF PROJECT TYPE ANALYSIS — Hours by Project Type › Task", len(ptah))
    style_header(ws_pta, 2, ptah, TEAL)
    ws_pta.auto_filter.ref = "A2:E2"
    ws_pta.cell(row=3, column=1, value="Grouped by Project Type → Task Category").font = Font(name="Manrope", size=9, italic=True, color="808080")
    for i, w in enumerate(ptaw, 1):
        ws_pta.column_dimensions[get_column_letter(i)].width = w

    if len(ff_df) > 0:
        pta_sum = ff_df.groupby(["project_type","ff_task"], as_index=False).agg(hours=("hours","sum"))
        pta_sum = pta_sum[pta_sum["ff_task"].notna() & (pta_sum["ff_task"] != "")].sort_values(["project_type","ff_task"])
        _all_ff_pta = df[df["billing_type"].fillna("").str.lower() == "fixed fee"] if "billing_type" in df.columns else df.copy()
        _proj_count_pta = _all_ff_pta.groupby("project_type")["project"].nunique().to_dict()
        _type_totals_pta = ff_df.groupby("project_type")["hours"].sum().to_dict()
        _prev_ptype_pta = None; _grp_idx_pta = 0; r_idx_pta = 4
        for _, row in pta_sum.iterrows():
            ptype_pta = row["project_type"]; ff_task_pta = row["ff_task"]
            type_total = _type_totals_pta.get(ptype_pta, 0)
            pct = row["hours"] / type_total if type_total > 0 else 0
            proj_cnt = _proj_count_pta.get(ptype_pta, 1)
            avg_hrs = round((row["hours"] / proj_cnt if proj_cnt > 0 else 0) * 4) / 4
            if ptype_pta != _prev_ptype_pta:
                _ptype_total = pta_sum[pta_sum["project_type"]==ptype_pta]["hours"].sum()
                for ci, (hval, hfmt) in enumerate([(ptype_pta, None), ("— ALL TASKS —", None), (_ptype_total, "#,##0.00"), ("", None), ("", None)], 1):
                    hcell = ws_pta.cell(row=r_idx_pta, column=ci, value=hval)
                    hcell.font = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
                    hcell.fill = PatternFill("solid", fgColor=NAVY); hcell.border = thin_border()
                    hcell.alignment = Alignment(horizontal="right" if ci==3 else "left", vertical="center")
                    if hfmt: hcell.number_format = hfmt
                r_idx_pta += 1; _prev_ptype_pta = ptype_pta; _grp_idx_pta = 0
            task_colors = {"Configuration":"EBF5FB","Post Go-Live Consulting":"FEF9E7",
                           "Project Management":"F4ECF7","Training & UAT":"EAF9F1","Customer Communication":"FDF2F8"}
            bg_pta, _grp_idx_pta = group_bg(ptype_pta, ptype_pta, _grp_idx_pta)
            task_bg_pta = task_colors.get(ff_task_pta, bg_pta)
            vals = ["", ff_task_pta, row["hours"], avg_hrs, pct]
            fmts = [None, None, "#,##0.00", "#,##0.00", "0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws_pta.cell(row=r_idx_pta, column=c_idx, value=val)
                style_cell(cell, task_bg_pta if c_idx > 1 else bg_pta, fmt=fmt, align="right" if c_idx > 2 else "left")
            r_idx_pta += 1
    else:
        ws_pta.cell(row=4, column=1, value="No Fixed Fee task data available.")

    ws_cr = wb.create_sheet("By Customer Region (WIP)")
    ws_cr.sheet_properties.tabColor = "1e2c63"
    ws_cr.freeze_panes = "A3"
    crh = ["Customer Region","Hours This Period","Utilization Credits","FF Project Overrun Hrs","Util %"]
    crw = [22,16,18,20,10]
    write_title(ws_cr, "SUMMARY — Utilization by Customer Region", len(crh))
    style_header(ws_cr, 2, crh, TEAL)
    ws_cr.auto_filter.ref = "A2:E2"
    for i, w in enumerate(crw, 1):
        ws_cr.column_dimensions[get_column_letter(i)].width = w
    if "customer_region" in df.columns:
        cr_base = df[df["credit_tag"] != "SKIPPED"].copy()
        cr_base["customer_region"] = cr_base["customer_region"].fillna("Unassigned")
        cr_sum = cr_base.groupby("customer_region", as_index=False).agg(
            hours_this_period=("hours","sum"), credit_hrs=("credit_hrs","sum"),
            ff_overrun_hrs=("variance_hrs","sum")).sort_values("customer_region")
        for r_idx, (_, row) in enumerate(cr_sum.iterrows(), 3):
            cr = row["customer_region"]; total_h = row["hours_this_period"]
            util = row["credit_hrs"] / total_h if total_h > 0 else 0
            util_bg = ("EAF9F1" if util >= 0.8 else "FEF9E7" if util >= 0.6 else "FDECED")
            bg = bgs[r_idx % 2]
            vals = [cr, total_h, row["credit_hrs"], row["ff_overrun_hrs"], util if total_h > 0 else "—"]
            fmts = [None,"#,##0.00","#,##0.00","#,##0.00","0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws_cr.cell(row=r_idx, column=c_idx, value=val)
                style_cell(cell, util_bg if c_idx == 5 else bg, fmt=fmt, align="right" if c_idx > 1 else "left")
    else:
        ws_cr.cell(row=3, column=1, value="No 'Customer Region' column found in import.")

    ws_ps = wb.create_sheet("By PS Region")
    ws_ps.sheet_properties.tabColor = "4472C4"
    ws_ps.freeze_panes = "A4"
    psh = ["PS Region","Project Type","Billing Type","Avail Hrs","Hours This Period",
           "Utilization Credits","FF Project Overrun Hrs","Admin Hrs","Util %"]
    psw = [14,28,14,12,16,18,20,14,10]
    write_title(ws_ps, "SUMMARY — Utilization by PS Region (APAC / EMEA / NOAM)", len(psh))
    style_header(ws_ps, 2, psh, TEAL)
    ws_ps.auto_filter.ref = "A3:I3"
    ws_ps.cell(row=3, column=1, value="Grouped by PS Region → Project Type → Billing Type").font = Font(name="Manrope", size=9, italic=True, color="808080")
    for i, w in enumerate(psw, 1):
        ws_ps.column_dimensions[get_column_letter(i)].width = w

    ps_avail = {}
    _seen_ep = set()
    for _emp, _grp in df.groupby("employee"):
        _loc = emp_region.get(_emp, ""); _ps = PS_REGION_MAP.get(_loc, "Other")
        for _p in _grp["period"].unique():
            if (_emp, _p) not in _seen_ep:
                _seen_ep.add((_emp, _p))
                ps_avail[_ps] = ps_avail.get(_ps, 0) + (get_avail_hours(_loc, _p) or 0)

    ps_admin = {}
    if "billing_type" in df.columns:
        for _, _ar in df[df["billing_type"].str.lower()=="internal"].iterrows():
            _ps = PS_REGION_MAP.get(_ar.get("region",""), "Other")
            ps_admin[_ps] = ps_admin.get(_ps, 0) + _ar.get("hours", 0)

    _ps_base = df[df["credit_tag"] != "SKIPPED"].copy()
    if "billing_type" not in _ps_base.columns: _ps_base["billing_type"] = "Unknown"
    _ps_detail = _ps_base.groupby(["ps_region","project_type","billing_type"], as_index=False).agg(
        hours_this_period=("hours","sum"), credit_hrs=("credit_hrs","sum"), ff_overrun_hrs=("variance_hrs","sum"))
    _ps_reg_total = _ps_base.groupby("ps_region", as_index=False).agg(
        hours_this_period=("hours","sum"), credit_hrs=("credit_hrs","sum"), ff_overrun_hrs=("variance_hrs","sum"))
    region_order = ["APAC","EMEA","NOAM","Other"]
    _ps_detail["_rord"] = _ps_detail["ps_region"].map({r:i for i,r in enumerate(region_order)}).fillna(99)
    _ps_detail = _ps_detail.sort_values(["_rord","ps_region","project_type","billing_type"]).drop(columns=["_rord"])

    r_idx = 4; _last_region = None
    for _, row in _ps_detail.iterrows():
        ps_reg = row["ps_region"]
        if ps_reg != _last_region:
            _last_region = ps_reg
            _rt = _ps_reg_total[_ps_reg_total["ps_region"]==ps_reg]
            _rh = _rt.iloc[0]["hours_this_period"] if len(_rt) else 0
            _rc = _rt.iloc[0]["credit_hrs"] if len(_rt) else 0
            _ro = _rt.iloc[0]["ff_overrun_hrs"] if len(_rt) else 0
            _ra = ps_admin.get(ps_reg, 0); _rv = ps_avail.get(ps_reg, 0)
            _ru = _rc / _rh if _rh > 0 else 0
            _ru_bg = "EAF9F1" if _ru>=0.7 else "FEF9E7" if _ru>=0.6 else "FDECED"
            reg_vals = [ps_reg, "— ALL TYPES —", "", _rv or "—", _rh, _rc, _ro, _ra, _ru if _rh > 0 else "—"]
            reg_fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(reg_vals, reg_fmts), 1):
                cell = ws_ps.cell(row=r_idx, column=c_idx, value=val)
                cell.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF" if c_idx <= 2 else "000000")
                cell.fill  = PatternFill("solid", fgColor=NAVY if c_idx <= 2 else (_ru_bg if c_idx == 9 else "D6DCF0"))
                cell.border = thin_border()
                if fmt: cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right" if c_idx > 3 else "left", vertical="center")
            r_idx += 1
        hrs  = row["hours_this_period"]
        util = row["credit_hrs"] / hrs if hrs > 0 else 0
        util_bg = "EAF9F1" if util >= 0.7 else "FEF9E7" if util >= 0.6 else "FDECED"
        bg = bgs[r_idx % 2]
        vals = ["", row["project_type"], row["billing_type"], "", hrs, row["credit_hrs"], row["ff_overrun_hrs"], "", util if hrs > 0 else "—"]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00",None,"0.0%"]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_ps.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, util_bg if c_idx == 9 else bg, fmt=fmt, align="right" if c_idx > 3 else "left")
        r_idx += 1

    ws_wl = wb.create_sheet("Watch List")
    ws_wl.sheet_properties.tabColor = "E74C3C"
    ws_wl.freeze_panes = "A3"
    wlh = ["Project","Project Type","PS Region","Project Manager",
           "Scoped Hrs","Previous Hrs to Date","Hours to Date","Hours Balance","Burn %","FF Overrun Hrs","Status"]
    wlw = [35,20,14,22,12,18,14,16,10,14,12]
    write_title(ws_wl, "PROJECT WATCH LIST — Overrun & At-Risk Projects", len(wlh))
    style_header(ws_wl, 2, wlh, "E74C3C")
    ws_wl.auto_filter.ref = f"A2:{get_column_letter(len(wlh))}2"
    for i, w in enumerate(wlw, 1):
        ws_wl.column_dimensions[get_column_letter(i)].width = w

    wl_df = ff_proj_df.groupby(["project","project_type"], as_index=False).agg(
        hours_this_period=("hours","sum"), credit_hrs=("credit_hrs","sum"),
        variance_hrs=("variance_hrs","sum"), htd_start=("htd_start","first"))
    wl_df["previous_htd"] = wl_df.apply(
        lambda r: max(0.0, (float(r["htd_start"]) if r["htd_start"] else 0.0) - r["hours_this_period"]), axis=1)
    wl_df["hours_to_date"] = wl_df.apply(
        lambda r: (float(r["htd_start"]) if r["htd_start"] else 0.0), axis=1)

    def get_scope_wl(row):
        ptype = str(row.get("project_type","") or ""); pname = str(row.get("project","") or "")
        if "premium" in ptype.strip().lower() and pname:
            _nums = _re_constants.findall(r"(?<![\d])(10|20)(?![\d])", pname)
            if _nums: return float(_nums[0])
        _pm = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in ptype.strip().lower()]
        return max(_pm, key=lambda x: len(x[0]))[1] if _pm else 0

    wl_df["scope_h"] = wl_df.apply(get_scope_wl, axis=1)
    wl_df["burn_pct"] = wl_df.apply(
        lambda r: (float(r["htd_start"]) if r["htd_start"] else 0) / r["scope_h"] if r["scope_h"] > 0 else None, axis=1)
    def _wl_status(r):
        s_h = r["scope_h"] or 0; htd = (float(r["htd_start"]) if r["htd_start"] else 0)
        if s_h > 0 and htd > s_h:   return "OVERRUN"
        if s_h > 0 and htd == s_h:  return "AT LIMIT"
        if s_h == 0 and htd > 0:    return "REVIEW"
        if (r["burn_pct"] or 0) >= 0.9: return "REVIEW"
        return "ON TRACK"
    wl_df["status"] = wl_df.apply(_wl_status, axis=1)
    watchlist = wl_df[wl_df["status"].isin(["OVERRUN","AT LIMIT","REVIEW"])].sort_values(
        "burn_pct", ascending=False, na_position="last")

    r_idx = 3
    for _, row in watchlist.iterrows():
        status = row["status"]
        status_bg = "FDECED" if status == "OVERRUN" else "FEF9E7"
        burn_val = row["burn_pct"] if row["burn_pct"] is not None else "—"
        pm_name = proj_pm.get(row["project"], "")
        start_dt = proj_start.get(row["project"])
        _htd_wl = float(row["previous_htd"]) + float(row["hours_this_period"])
        _tot_ov = _htd_wl - row["scope_h"] if row["scope_h"] and row["scope_h"] > 0 else "—"
        _ps_reg_wl = proj_ps_region.get(row["project"], "")
        vals = [row["project"], row["project_type"], _ps_reg_wl, pm_name,
                row["scope_h"] or "—", row["previous_htd"], _htd_wl, _tot_ov,
                burn_val, row["variance_hrs"], status]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%","#,##0.00",None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_wl.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, status_bg if c_idx == 11 else status_bg, fmt=fmt, bold=(c_idx == 9),
                       align="right" if c_idx in (5,6,7,8) else "center" if c_idx == 9 else "left")
        r_idx += 1

    r_idx += 1
    unconf_title_cell = ws_wl.cell(row=r_idx, column=1, value="FF: NO SCOPE DEFINED (Hours at Risk)")
    unconf_title_cell.font  = Font(name="Manrope", bold=True, size=11, color="FFFFFF")
    unconf_title_cell.fill  = hdr_fill("E67E22")
    ws_wl.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(wlh))
    r_idx += 1
    unconf_df = df[df["credit_tag"] == "UNCONFIGURED"].groupby(
        ["project","project_type"], as_index=False).agg(hours=("hours","sum")).sort_values("hours", ascending=False)
    for _, row in unconf_df.iterrows():
        bg = "FEF3E2"
        vals = [row["project"], row["project_type"], proj_cust_region.get(row["project"],""),
                proj_pm.get(row["project"],""), "—", "—", "—", row["hours"], "FF: NO SCOPE DEFINED"]
        fmts = [None,None,None,None,None,None,None,"#,##0.00",None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_wl.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, bg, fmt=fmt, align="right" if c_idx == 8 else "left")
        r_idx += 1

    ws7 = wb.create_sheet("Skipped Rows")
    ws7.sheet_properties.tabColor = "E74C3C"
    ws7.freeze_panes = "A3"
    skh = ["Employee","Project","Project Type","Billing Type","Date","Hours","Reason"]
    skw = [20,35,20,14,14,10,45]
    write_title(ws7, "SKIPPED ROWS — Not Included in Utilization Calculations", len(skh))
    style_header(ws7, 2, skh, "C0392B")
    for i, w in enumerate(skw, 1):
        ws7.column_dimensions[get_column_letter(i)].width = w
    skip_cols = ["employee","project","project_type","billing_type","date","hours","notes"]
    skipped = df[df["credit_tag"] == "SKIPPED"]
    if len(skipped) > 0:
        for r_idx, (_, row) in enumerate(skipped.iterrows(), 3):
            for c_idx, col in enumerate(skip_cols, 1):
                val  = row.get(col, "")
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                fmt  = "YYYY-MM-DD" if col == "date" else "#,##0.00" if col == "hours" else None
                style_cell(cell, "FDECED", fmt=fmt,
                           align="right" if col == "hours" else "center" if col == "date" else "left")
    else:
        ws7.cell(row=3, column=1, value="No skipped rows — all entries were processed.")

    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.sheet_properties.tabColor = "1e2c63"
    ws_dash.sheet_view.showGridLines = False

    def dash_label(ws, row, col, text, size=10, bold=False, color="808080"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        return c

    def dash_value(ws, row, col, value, fmt=None, size=18, bold=True, color="1e2c63"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        if fmt: c.number_format = fmt
        return c

    def dash_section(ws, row, col, text, ncols=4):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Manrope", size=11, bold=True, color="FFFFFF")
        c.fill = hdr_fill(NAVY)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ncols-1)
        return c

    def rag_cell(ws, row, col, value, fmt=None, status="green"):
        colors = {"green":"EAF9F1","yellow":"FEF9E7","red":"FDECED"}
        txt    = {"green":"2ECC71","yellow":"F39C12","red":"E74C3C"}
        c = ws.cell(row=row, column=col, value=value)
        c.font  = Font(name="Manrope", size=14, bold=True, color=txt.get(status,"000000"))
        c.fill  = PatternFill("solid", fgColor=colors.get(status,"FFFFFF"))
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt: c.number_format = fmt
        return c

    for col, w in [(1,3),(2,22),(3,18),(4,18),(5,18),(6,18),(7,18),(8,3)]:
        ws_dash.column_dimensions[get_column_letter(col)].width = w
    for row in range(1, 45):
        ws_dash.row_dimensions[row].height = 18

    tc = ws_dash.cell(row=2, column=2, value="Professional Services — Utilization Credit Report")
    tc.font = Font(name="Manrope", size=16, bold=True, color="FFFFFF")
    tc.fill = hdr_fill(NAVY)
    ws_dash.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws_dash.row_dimensions[2].height = 30

    if "date" in df.columns:
        max_dt   = pd.to_datetime(df["date"], errors="coerce").max()
        date_str = max_dt.strftime("%d %B %Y") if pd.notna(max_dt) else "—"
    else:
        date_str = "—"
    sc = ws_dash.cell(row=3, column=2, value=f"Data through {date_str}")
    sc.font = Font(name="Manrope", size=10, color="808080")
    ws_dash.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    kc = ws_dash.cell(row=4, column=2, value="This report calculates Utilization Credits from NetSuite time detail exports. T&M projects: full credit for all hours logged. Fixed Fee projects: credit up to scoped hours; hours beyond scope tracked as overrun (excluded from credits). Internal time: excluded from utilization, tracked as Admin Hours. Util % = Utilization Credits / Hours This Period.")
    kc.font = Font(name="Manrope", size=9, italic=True, color="808080")
    kc.alignment = Alignment(wrap_text=True)
    ws_dash.merge_cells(start_row=4, start_column=2, end_row=4, end_column=7)
    ws_dash.row_dimensions[4].height = 30

    dash_section(ws_dash, 6, 2, "KEY METRICS", ncols=6)
    ws_dash.row_dimensions[5].height = 22
    hours_tp_d   = df[df["credit_tag"] != "SKIPPED"]["hours"].sum()
    credit_hrs_d = df[df["credit_tag"].isin(["CREDITED","PARTIAL"])]["credit_hrs"].sum()
    overrun_hrs_d = df[df["credit_tag"].isin(["OVERRUN", "PARTIAL"])]["variance_hrs"].sum()
    admin_hrs_d  = df[df["billing_type"].str.lower()=="internal"]["hours"].sum() if "billing_type" in df.columns else 0
    util_pct_d   = credit_hrs_d / hours_tp_d if hours_tp_d > 0 else 0
    util_status_d = "green" if util_pct_d >= 0.70 else "yellow" if util_pct_d >= 0.60 else "red"

    for i, (label, value, fmt, status) in enumerate([
        ("Hours This Period", hours_tp_d, "#,##0.00", None),
        ("Utilization Credits", credit_hrs_d, "#,##0.00", None),
        ("Util % (target 70%)", util_pct_d, "0.0%", util_status_d),
        ("FF Overrun Hrs", overrun_hrs_d, "#,##0.00", None),
        ("Admin Hrs", admin_hrs_d, "#,##0.00", None),
        ("Projects This Period", df[df["billing_type"].fillna("").str.lower() != "internal"].groupby(["project","project_type"]).ngroups, "#,##0", None),
    ]):
        col = 2 + i
        dash_label(ws_dash, 7, col, label)
        if status: rag_cell(ws_dash, 8, col, value, fmt=fmt, status=status)
        else: dash_value(ws_dash, 8, col, value, fmt=fmt, size=14)
    ws_dash.row_dimensions[8].height = 28

    dash_section(ws_dash, 10, 2, "UTILIZATION BY PS REGION", ncols=6)
    ws_dash.row_dimensions[9].height = 22
    for ci, hdr in enumerate(["PS Region","Hours This Period","Credit Hrs","Util %","FF Overrun Hrs","Admin Hrs"], 2):
        c = ws_dash.cell(row=11, column=ci, value=hdr)
        c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
        c.fill = hdr_fill(TEAL)

    ps_base_d = df[df["credit_tag"] != "SKIPPED"]
    ps_sum_d  = ps_base_d.groupby("ps_region").agg(hours=("hours","sum"), credit=("credit_hrs","sum"), overrun=("variance_hrs","sum"))
    ps_admin_d = df[df["billing_type"].str.lower()=="internal"].groupby("ps_region")["hours"].sum() if "billing_type" in df.columns else pd.Series(dtype=float)
    ps_avail_d = {}
    _seen_emp_period = set()
    for _emp2, _grp2 in df.groupby("employee"):
        _loc2 = emp_region.get(_emp2,""); _ps2 = PS_REGION_MAP.get(_loc2,"Other")
        for _p2 in _grp2["period"].unique():
            if (_emp2, _p2) not in _seen_emp_period:
                _seen_emp_period.add((_emp2, _p2))
                ps_avail_d[_ps2] = ps_avail_d.get(_ps2,0) + (get_avail_hours(_loc2,_p2) or 0)

    for ri, reg in enumerate(["APAC","EMEA","NOAM","Other"], 12):
        if reg not in ps_sum_d.index: continue
        _row = ps_sum_d.loc[reg]
        _adm = float(ps_admin_d.get(reg,0)) if reg in ps_admin_d.index else 0
        _util= _row["credit"] / _row["hours"] if _row["hours"] > 0 else None
        _bg  = bgs[ri % 2]
        _util_color = ("E74C3C" if _util<0.60 else "2ECC71" if _util>=0.70 else "F39C12") if _util is not None else "808080"
        _dash_ps_vals = [
            (2, reg, None, False, "000000"),
            (3, _row["hours"], "#,##0.00", False, "000000"),
            (4, _row["credit"], "#,##0.00", False, "000000"),
            (5, _util if _util is not None else "—", "0.0%" if _util is not None else None, True, _util_color),
            (6, _row["overrun"], "#,##0.00", False, "000000"),
            (7, _adm, "#,##0.00", False, "000000"),
        ]
        for ci2, val2, fmt2, bold2, color2 in _dash_ps_vals:
            _c = ws_dash.cell(row=ri, column=ci2, value=val2)
            _c.font = Font(name="Manrope", size=10, bold=bold2, color=color2)
            _c.fill = PatternFill("solid", fgColor=_bg)
            _c.border = thin_border()
            if fmt2: _c.number_format = fmt2

    dash_section(ws_dash, 17, 2, "WATCH LIST SUMMARY", ncols=6)
    ws_dash.row_dimensions[17].height = 22
    n_overrun  = len(wl_df[wl_df["status"]=="OVERRUN"]) if len(wl_df) > 0 else len(df[df["credit_tag"]=="OVERRUN"]["project"].unique())
    _wl_at_risk = wl_df[(wl_df["burn_pct"].notna()) & (wl_df["burn_pct"]>=0.9) & (wl_df["status"]!="OVERRUN")]
    n_at_risk  = len(_wl_at_risk["project"].unique()) if len(_wl_at_risk) > 0 else 0
    n_unconf   = len(df[df["credit_tag"]=="UNCONFIGURED"]["project"].unique())
    unconf_hrs_d = df[df["credit_tag"]=="UNCONFIGURED"]["hours"].sum()
    for i, (label, value, fmt, status) in enumerate([
        ("Projects in Overrun", n_overrun, "#,##0", "red" if n_overrun>0 else "green"),
        ("Projects (≥90% burn)", n_at_risk, "#,##0", "yellow" if n_at_risk>0 else "green"),
        ("FF: No Scope Defined Projects", n_unconf, "#,##0", "yellow" if n_unconf>0 else "green"),
        ("FF: No Scope Defined Hours", unconf_hrs_d, "#,##0.00", "yellow" if unconf_hrs_d>0 else "green"),
    ]):
        col = 2 + i
        dash_label(ws_dash, 18, col, label)
        rag_cell(ws_dash, 19, col, value, fmt=fmt, status=status)
    ws_dash.row_dimensions[19].height = 28

    dash_section(ws_dash, 21, 2, "EMPLOYEES BELOW 60% UTILIZATION — Action Required", ncols=6)
    ws_dash.row_dimensions[21].height = 22
    for ci, hdr in enumerate(["Employee","Location","PS Region","Period","Util %","Credit Hrs"], 2):
        _c = ws_dash.cell(row=22, column=ci, value=hdr)
        _c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
        _c.fill = hdr_fill(TEAL)

    _low_rows = []
    for _, _erow in emp_sum.iterrows():
        _emp3  = _erow["employee"]
        if any(_emp3.lower().startswith(ex.lower()) for ex in UTIL_EXEMPT_EMPLOYEES): continue
        _loc3  = emp_region.get(_emp3,""); _ps3 = PS_REGION_MAP.get(_loc3,"Other")
        _p3    = _erow["period"]; _avl3 = get_avail_hours(_loc3, _p3) or 0
        _util3 = _erow["credit_hrs"] / _avl3 if _avl3 > 0 else 0
        if _util3 < 0.60 and _avl3 > 0:
            _low_rows.append((_emp3, _loc3, _ps3, _p3, _util3, _erow["credit_hrs"]))

    for ri, (_e,_l,_ps,_per,_u,_c) in enumerate(sorted(_low_rows, key=lambda x:x[4])[:15], 23):
        for ci2, (val2,fmt2) in enumerate([(_e,None),(_l,None),(_ps,None),(_per,None),(_u,"0.0%"),(_c,"#,##0.00")], 2):
            _cv = ws_dash.cell(row=ri, column=ci2, value=val2)
            _cv.font  = Font(name="Manrope", size=10, color="E74C3C" if ci2==6 else "000000")
            _cv.fill  = PatternFill("solid", fgColor="FDECED")
            _cv.border = thin_border()
            if fmt2: _cv.number_format = fmt2

    sheet_order = [
        "Dashboard", "Project Count", "SUMMARY - By Employee", "FF By Project Utilization",
        "FF Overrun by Project Type", "By Customer Region (WIP)", "By PS Region",
        "Watch List", "Non-Billable", "Task Analysis", "FF Project Type Analysis",
        "Skipped Rows", "PROCESSED_DATA",
    ]
    existing  = [s for s in sheet_order if s in wb.sheetnames]
    remaining = [s for s in wb.sheetnames if s not in existing]
    wb._sheets = [wb[s] for s in existing + remaining]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════
# PAGE
# ════════════════════════════════════════════════════════

import streamlit as st
import pandas as pd
import io




def get_avail_hours(region, period):
    """Look up available hours for a region/period."""
    region_clean = str(region).strip()
    for r, months in AVAIL_HOURS.items():
        if r.lower() == region_clean.lower():
            return months.get(str(period), None)
    return None


def calc_consultant_util(ns_rows, month_key, scope_map, avail_hours):
    """
    Calculate util % for a single consultant for a given month.
    T&M full credit + FF capped at scope, keyed by project_id to prevent
    same-customer multi-product bleed.
    """
    import re as _re_util
    import pandas as _pd_util

    empty = dict(util_pct=None, util_hrs=0.0, overrun_pct=None, overrun_hrs=0.0,
                 admin_pct=None, admin_hrs=0.0, tm_hrs=0.0, ff_credit=0.0, ff_overrun=0.0)

    if ns_rows is None or ns_rows.empty or not avail_hours:
        return empty

    ns_rows = ns_rows.copy()
    if "date" in ns_rows.columns:
        ns_rows["date"] = _pd_util.to_datetime(ns_rows["date"], errors="coerce")

    month_ns = ns_rows[ns_rows["date"].dt.strftime("%Y-%m") == month_key].copy()
    if month_ns.empty:
        return empty

    bt_col = "billing_type" if "billing_type" in month_ns.columns else None
    if bt_col:
        _bt = month_ns[bt_col].fillna("").str.strip().str.lower()
        admin_hrs = round(float(month_ns[_bt == "internal"]["hours"].sum() or 0), 2)
        tm_hrs    = round(float(month_ns[_bt == "t&m"]["hours"].sum() or 0), 2)
        ff_rows   = month_ns[(_bt != "internal") & (_bt != "t&m")].copy()
    else:
        admin_hrs = 0.0
        tm_hrs    = round(float(month_ns["hours"].sum() or 0), 2)
        ff_rows   = _pd_util.DataFrame()

    ff_credit = 0.0
    ff_overrun = 0.0

    if not ff_rows.empty and "project" in ff_rows.columns:
        ff_rows = ff_rows.sort_values(["project", "date"])

        # Pre-compute prior HTD per project, keyed by both project_id and project name
        prior_htd = {}
        if "hours_to_date" in month_ns.columns:
            _pid_col = "project_id" if "project_id" in month_ns.columns else "project"
            for _proj_key, _grp in month_ns.groupby(_pid_col):
                _proj_n = " ".join(str(_grp["project"].iloc[0]).strip().split()) \
                    if _pid_col == "project_id" else " ".join(str(_proj_key).strip().split())
                _pid_str = str(_proj_key).strip().replace(".0", "") \
                    if _pid_col == "project_id" else None
                try:
                    _max_htd    = float(_grp["hours_to_date"].dropna().astype(float).max() or 0)
                    _period_hrs = float(_grp["hours"].dropna().astype(float).sum() or 0)
                    _val = max(0.0, _max_htd - _period_hrs)
                    prior_htd[_proj_n] = _val
                    if _pid_str and _pid_str not in ("nan", "None", ""):
                        prior_htd[_pid_str] = _val
                except Exception:
                    prior_htd[_proj_n] = 0.0
                    if _pid_str and _pid_str not in ("nan", "None", ""):
                        prior_htd[_pid_str] = 0.0

        _con = {}
        for _, _r in ff_rows.iterrows():
            _proj  = " ".join(str(_r.get("project", "")).split())
            _ptype = str(_r.get("project_type", "")).strip()
            _hrs   = float(_r.get("hours", 0) or 0)
            if _hrs <= 0:
                continue

            _ptype_lower = _ptype.lower()
            if "premium" in _ptype_lower:
                _sku = str(_r.get("time_item_sku", "") or "")
                _sku_nums = _re_util.findall(r"IMPL(\d+)", _sku.upper())
                if _sku_nums:
                    _sc = float(_sku_nums[0])
                else:
                    _name_nums = _re_util.findall(r"(?<!\d)(10|20)(?!\d)", str(_r.get("project", "")))
                    _sc = float(_name_nums[0]) if _name_nums else None
            else:
                _m  = [(k, float(v)) for k, v in scope_map.items()
                       if k.strip().lower() in _ptype_lower]
                _sc = max(_m, key=lambda x: len(x[0]))[1] if _m else None

            if _sc is None:
                continue

            # Key on project_id if available
            _pid_raw = str(_r.get("project_id", "") or "").strip().replace(".0", "")
            _con_key = _pid_raw if _pid_raw and _pid_raw not in ("nan", "None", "") else _proj

            if _con_key not in _con:
                _con[_con_key] = prior_htd.get(_con_key, prior_htd.get(_proj, 0.0))

            _used = _con[_con_key]
            _rem  = _sc - _used
            if _rem <= 0:
                ff_overrun += _hrs
            elif _hrs <= _rem:
                ff_credit += _hrs
                _con[_con_key] = _used + _hrs
            else:
                ff_credit += _rem
                ff_overrun += _hrs - _rem
                _con[_con_key] = _sc

    ff_credit  = round(ff_credit, 2)
    ff_overrun = round(ff_overrun, 2)
    util_hrs   = round(tm_hrs + ff_credit, 2)
    avail      = float(avail_hours)

    return dict(
        util_pct    = round(util_hrs   / avail * 100, 1) if avail else None,
        util_hrs    = util_hrs,
        overrun_pct = round(ff_overrun / avail * 100, 1) if avail else None,
        overrun_hrs = ff_overrun,
        admin_pct   = round(admin_hrs  / avail * 100, 1) if avail else None,
        admin_hrs   = admin_hrs,
        tm_hrs      = tm_hrs,
        ff_credit   = ff_credit,
        ff_overrun  = ff_overrun,
    )
