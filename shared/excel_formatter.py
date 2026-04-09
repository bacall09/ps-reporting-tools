"""
Zone & Co brand Excel formatter for PS Revenue Report.
- Dashboard tab: full Zone branding (Manrope, navy/teal)
- Data tabs: clean Arial finance format with autofilter
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import io
import re

# ── Brand colours ──────────────────────────────────────────────────────────────
NAVY   = "1E2C63"
TEAL   = "08A9B7"
BLUE   = "4472C4"
WHITE  = "FFFFFF"
GREY   = "808080"
LGREY  = "F2F2F2"
DGREY  = "64748B"
RED    = "E74C3C"
RED_BG = "FDECED"

# Month column pattern — e.g. "Jan 2026", "Feb 2026"
_MONTH_RE = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}$')

def _is_month_col(name: str) -> bool:
    return bool(_MONTH_RE.match(str(name).strip()))

def _mfont(name="Manrope", size=10, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)

def _afont(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


# ══════════════════════════════════════════════════════════════════════════════
# DATA SHEET FORMATTER  (Arial, autofilter, USD on month cols)
# ══════════════════════════════════════════════════════════════════════════════

def format_data_sheet(ws, title_text):
    """Finance-standard data sheet: Arial, blue header, autofilter, USD month cols."""
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # ── Title banner row 1 ────────────────────────────────────────────────────
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font      = Font(name="Arial", size=12, bold=True, color=WHITE)
    cell.fill      = _fill(NAVY)
    cell.alignment = _left()
    ws.row_dimensions[1].height = 20
    for c in range(2, max_col + 1):
        ws.cell(row=1, column=c).fill = _fill(NAVY)

    # ── Header row 2 — blue ───────────────────────────────────────────────────
    header_values = []
    if max_row >= 2:
        for cell in ws[2]:
            header_values.append(str(cell.value or ""))
            cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
            cell.fill      = _fill(BLUE)
            cell.alignment = _center()
        ws.row_dimensions[2].height = 18

    # ── Autofilter on header row ──────────────────────────────────────────────
    if max_row >= 2 and max_col >= 1:
        ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}2"

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Identify month columns for USD formatting ─────────────────────────────
    month_col_indices = set()
    for i, hdr in enumerate(header_values):
        if _is_month_col(hdr) or hdr in ("Rev Amount", "Total Carve",
                                          "YTD", "QTD", "MTD", "Full Month"):
            month_col_indices.add(i + 1)  # 1-indexed

    # ── Data rows — alternating, Arial 10pt ──────────────────────────────────
    for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=max_row)):
        bg = _fill(LGREY) if r_idx % 2 == 1 else _fill(WHITE)
        for cell in row:
            cell.font  = Font(name="Arial", size=10)
            cell.fill  = bg
            cell.alignment = _left()
            # USD number format on month/value columns
            if cell.column in month_col_indices:
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '$#,##0.00_);($#,##0.00)'
                        cell.alignment = _center()
                    except (ValueError, TypeError):
                        pass

    # ── Auto-width columns (capped at 45) ────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 45)


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD BUILDER  (Manrope, Zone brand)
# ══════════════════════════════════════════════════════════════════════════════

def _mfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def format_metric_card(ws, row, col, label, value, alert=False):
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = _mfont(size=10, color=GREY)
    val = ws.cell(row=row+1, column=col, value=value)
    if alert:
        val.font      = _mfont(size=14, bold=True, color=RED)
        val.fill      = _mfill(RED_BG)
        val.alignment = _center()
    else:
        val.font = _mfont(size=14, bold=True, color=NAVY)

def format_section_header(ws, row, col, text, n_cols=6):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = _mfont(size=11, bold=True, color=WHITE)
    cell.fill      = _mfill(NAVY)
    cell.alignment = _left()
    ws.row_dimensions[row].height = 20
    for c in range(col+1, col+n_cols):
        ws.cell(row=row, column=c).fill = _mfill(NAVY)

def format_table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col+i, value=h)
        cell.font      = _mfont(size=9, bold=True, color=WHITE)
        cell.fill      = _mfill(BLUE)
        cell.alignment = _center()
    ws.row_dimensions[row].height = 18

def build_dashboard(wb, metrics: dict, date_str: str, blurb: str):
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    for col in ["C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 18

    # Header banner
    cell = ws.cell(row=2, column=2, value="Professional Services — Revenue Report")
    cell.font = _mfont(size=16, bold=True, color=WHITE)
    cell.fill = _mfill(NAVY)
    cell.alignment = _left()
    ws.row_dimensions[2].height = 28
    for c in range(3, 9):
        ws.cell(row=2, column=c).fill = _mfill(NAVY)

    sub = ws.cell(row=3, column=2, value=f"Data through {date_str}")
    sub.font = _mfont(size=10, color=GREY)

    bl = ws.cell(row=4, column=2, value=blurb)
    bl.font = _mfont(size=9, color=GREY)
    bl.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 36

    # Executive Summary
    format_section_header(ws, row=5, col=2, text="EXECUTIVE SUMMARY", n_cols=7)
    exec_metrics = [
        ("YTD Revenue",         metrics.get("ytd"),      False),
        ("QTD Revenue",         metrics.get("qtd"),      False),
        ("MTD Revenue",         metrics.get("mtd"),      False),
        ("Full Month Forecast", metrics.get("full_mo"),  False),
        ("Run Rate (ARR)",      metrics.get("run_rate"), False),
        ("MoM Growth",          metrics.get("mom"),      False),
    ]
    for i, (label, value, alert) in enumerate(exec_metrics):
        format_metric_card(ws, row=6, col=2+i, label=label, value=value, alert=alert)
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 22

    # Revenue by Type
    format_section_header(ws, row=9, col=2, text="REVENUE BY TYPE", n_cols=7)
    type_metrics = [
        ("FF Revenue YTD",      metrics.get("ff_ytd"),      False),
        ("T&M Revenue YTD",     metrics.get("tm_ytd"),      False),
        ("Reconcile Carve YTD", metrics.get("recon_ytd"),   False),
        ("FF Projects",         metrics.get("ff_projects"), False),
        ("T&M Projects",        metrics.get("tm_projects"), False),
        ("Carve Flags",         metrics.get("flag_count"),
         (metrics.get("flag_count") or 0) > 0),
    ]
    for i, (label, value, alert) in enumerate(type_metrics):
        format_metric_card(ws, row=10, col=2+i, label=label, value=value, alert=alert)
    ws.row_dimensions[10].height = 16
    ws.row_dimensions[11].height = 22

    # Analyst Detail
    format_section_header(ws, row=13, col=2,
                          text="ANALYST DETAIL — Revenue by Month (USD)", n_cols=7)
    trend_data = metrics.get("trend_rows", [])
    if trend_data:
        hdrs = list(trend_data[0].keys())
        format_table_headers(ws, row=14, headers=hdrs, start_col=2)
        for r_idx, row_data in enumerate(trend_data):
            fill = _mfill(LGREY) if r_idx % 2 == 1 else _mfill(WHITE)
            for c_idx, val in enumerate(row_data.values()):
                cell = ws.cell(row=15+r_idx, column=2+c_idx, value=val)
                cell.font  = _mfont(size=10)
                cell.fill  = fill
        last_trend_row = 15 + len(trend_data)
    else:
        last_trend_row = 15

    # By Region + By Product
    format_section_header(ws, row=last_trend_row+1, col=2,
                          text="BY REGION — YTD", n_cols=3)
    format_section_header(ws, row=last_trend_row+1, col=5,
                          text="BY PRODUCT — YTD", n_cols=4)
    region_rows  = metrics.get("region_rows", [])
    product_rows = metrics.get("product_rows", [])
    if region_rows:
        hdrs_r = list(region_rows[0].keys())
        format_table_headers(ws, row=last_trend_row+2, headers=hdrs_r, start_col=2)
        for r_idx, rd in enumerate(region_rows):
            fill = _mfill(LGREY) if r_idx % 2 == 1 else _mfill(WHITE)
            for c_idx, val in enumerate(rd.values()):
                cell = ws.cell(row=last_trend_row+3+r_idx, column=2+c_idx, value=val)
                cell.font = _mfont(size=10); cell.fill = fill
    if product_rows:
        hdrs_p = list(product_rows[0].keys())
        format_table_headers(ws, row=last_trend_row+2, headers=hdrs_p, start_col=5)
        for r_idx, pd_row in enumerate(product_rows):
            fill = _mfill(LGREY) if r_idx % 2 == 1 else _mfill(WHITE)
            for c_idx, val in enumerate(pd_row.values()):
                cell = ws.cell(row=last_trend_row+3+r_idx, column=5+c_idx, value=val)
                cell.font = _mfont(size=10); cell.fill = fill
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def apply_zone_formatting(excel_bytes: bytes, metrics: dict,
                           date_str: str, blurb: str) -> bytes:
    wb = load_workbook(io.BytesIO(excel_bytes))

    # Build Dashboard first
    build_dashboard(wb, metrics, date_str, blurb)

    # Format all data sheets with finance style
    skip = {"Dashboard"}
    for sheet_name in wb.sheetnames:
        if sheet_name in skip:
            continue
        ws = wb[sheet_name]
        # Insert two blank rows at top for title banner + header
        ws.insert_rows(1)
        ws.insert_rows(1)
        format_data_sheet(ws, sheet_name)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
