"""
Capacity Outlook
Upload SS DRS + NS Unassigned Projects to project consultant availability
and surface demand from unassigned (closed) deals.
"""
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Capacity Outlook", page_icon=None, layout="wide")

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY   = "1e2c63"
WHITE  = "FFFFFF"
LTGRAY = "F2F2F2"
RED    = "E74C3C"
GREEN  = "27AE60"
AMBER  = "F39C12"
BLUE   = "4472C4"

# ── Employee config (mirrors other pages) ─────────────────────────────────────
EMPLOYEE_ROLES = {
    # ── Project Managers (no product delivery) ────────────────────────────────
    "Barrio, Nairobi":        {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Hughes, Madalyn":        {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Porangada, Suraj":       {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Cadelina":               {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    # ── Solution Architects ───────────────────────────────────────────────────
    "Bell, Stuart":           {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "DiMarco, Nicole R":      {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "Murphy, Conor":          {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},
    # ── Developer ─────────────────────────────────────────────────────────────
    "Church, Jason G":        {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    # ── Consultants ───────────────────────────────────────────────────────────
    "Arestarkhov, Yaroslav":  {"role": "Consultant",         "products": ["Billing", "Capture"],                                                                   "learning": []},
    "Carpen, Anamaria":       {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Centinaje, Rhodechild":  {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},
    "Cooke, Ellen":           {"role": "Consultant",         "products": ["Billing", "Payroll"],                                                                   "learning": []},
    "Dolha, Madalina":        {"role": "Consultant",         "products": ["Capture", "Reconcile", "CC Statement Import", "Reconcile PSP", "e-Invoicing"],          "learning": []},
    "Finalle-Newton, Jesse":  {"role": "Solution Architect", "products": ["Reporting"],                                                                            "learning": []},
    "Gardner, Cheryll L":     {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Hopkins, Chris":         {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Ickler, Georganne":      {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Isberg, Eric":           {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Jordanova, Marija":      {"role": "Consultant",         "products": ["Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],     "learning": []},
    "Lappin, Thomas":         {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": ["Capture", "Reconcile"]},
    "Longalong, Santiago":    {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": ["Billing"]},
    "Mohammad, Manaan":       {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Morris, Lisa":           {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "NAQVI, SYED":            {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Olson, Austin D":        {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Pallone, Daniel":        {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Raykova, Silvia":        {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Selvakumar, Sajithan":   {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Snee, Stefanie J":       {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Swanson":                {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},
    "Tuazon, Carol":          {"role": "Consultant",         "products": ["Payroll", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],       "learning": []},
    "Zoric, Ivan":            {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},
    "Dunn, Steven":           {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Law, Brandon":           {"role": "Developer",          "products": ["Reporting"],                                                                                  "learning": []},
    "Quiambao, Generalyn":    {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
        "Cruz, Daniel":           {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "Payments", "e-Invoicing", "SFTP Connector", "CC Statement Import"],                                                                                  "learning": []},
    # ── Leavers (historical) ──────────────────────────────────────────────────
    "Alam, Laisa":            {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Chan, Joven":            {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Cloete, Bronwyn":        {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Eyong, Eyong":           {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Hamilton, Julie C":      {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Hernandez, Camila":      {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Rushbrook, Emma C":      {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Strauss, John W":        {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
}

EMPLOYEE_LOCATION = {
    "Arestarkhov, Yaroslav":  ("Czech Republic",      None,       None),
    "Barrio, Nairobi":        ("USA",                 None,       None),
    "Bell, Stuart":           ("USA",                 None,       None),
    "Cadelina":               ("Manila (PH)",         "2026-03",  None),
    "Carpen, Anamaria":       ("Spain",               None,       None),
    "Centinaje, Rhodechild":  ("Manila (PH)",         None,       None),
    "Church, Jason G":        ("USA",                 None,       None),
    "Cooke, Ellen":           ("Northern Ireland",    None,       None),
    "Cruz, Daniel":           ("Manila (PH)",         None,       None),
    "DiMarco, Nicole R":      ("USA",                 None,       None),
    "Dolha, Madalina":        ("Faroe Islands",       None,       None),
    "Finalle-Newton, Jesse":  ("USA",                 None,       None),
    "Gardner, Cheryll L":     ("USA",                 None,       None),
    "Hopkins, Chris":         ("USA",                 None,       None),
    "Hughes, Madalyn":        ("USA",                 None,       None),
    "Ickler, Georganne":      ("USA",                 None,       None),
    "Isberg, Eric":           ("USA",                 None,       None),
    "Jordanova, Marija":      ("North Macedonia",     None,       None),
    "Lappin, Thomas":         ("Northern Ireland",    None,       None),
    "Longalong, Santiago":    ("Manila (PH)",         None,       None),
    "Mohammad, Manaan":       ("Canada",              None,       None),
    "Morris, Lisa":           ("Sydney (NSW)",        None,       None),
    "Murphy, Conor":          ("USA",                 None,       None),
    "NAQVI, SYED":            ("Canada",              None,       None),
    "Olson, Austin D":        ("USA",                 None,       None),
    "Pallone, Daniel":        ("Sydney (NSW)",        None,       None),
    "Porangada, Suraj":       ("USA",                 None,       None),
    "Raykova, Silvia":        ("Netherlands",         None,       None),
    "Selvakumar, Sajithan":   ("Canada",              None,       None),
    "Snee, Stefanie J":       ("USA",                 None,       None),
    "Stone, Matt":            ("USA",                 None,       None),
    "Swanson":                ("USA",                 None,       None),
    "Tuazon, Carol":          ("Manila (PH)",         None,       None),
    "Zoric, Ivan":            ("Serbia",              None,       None),
    "Alam, Laisa":            ("USA",                 None,       "2025-12"),
    "Chan, Joven":            ("Manila (PH)",         None,       "2025-12"),
    "Cloete, Bronwyn":        ("Netherlands",         None,       "2026-02"),
    "Eyong, Eyong":           ("USA",                 None,       "2025-12"),
    "Hamilton, Julie C":      ("USA",                 None,       "2026-01"),
    "Hernandez, Camila":      ("USA",                 None,       "2025-12"),
    "Rushbrook, Emma C":      ("Wales",               None,       "2025-12"),
    "Dunn, Steven":            ("USA",                 None,       None),
    "Law, Brandon":            ("USA",                 None,       None),
    "Quiambao, Generalyn":     ("Manila (PH)",         None,       None),
    "Strauss, John W":        ("USA",                 None,       "2026-03"),
}

PS_REGION_OVERRIDE = {
    "NAQVI, SYED":       "EMEA",
    "Cruz, Daniel":      "NOAM",
    "Chan, Joven":       "NOAM",
    "Rushbrook, Emma C": "EMEA",
}
PS_REGION_MAP = {
    "Sydney (NSW)":     "APAC",
    "Manila (PH)":      "APAC",
    "UK":               "EMEA",
    "Wales":            "EMEA",
    "Spain":            "EMEA",
    "Netherlands":      "EMEA",
    "Northern Ireland": "EMEA",
    "Faroe Islands":    "EMEA",
    "North Macedonia":  "EMEA",
    "Czech Republic":   "EMEA",
    "Serbia":           "EMEA",
    "USA":              "NOAM",
    "Canada":           "NOAM",
}

# ── Phase duration table — end week per phase per product type ─────────────────
# Values = week number from project start when phase completes
PHASE_END_WEEKS = {
    "Approvals": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     6,
        "05. prep for go-live":        6,
        "06. go-live":                 8,
        "08. ready for support transition": 8,
    },
    "Capture": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     6,
        "05. prep for go-live":        6,
        "06. go-live":                 8,
        "08. ready for support transition": 8,
    },
    "Capture & e-Invoicing": {
        "01. requirements and design": 1,
        "02. configuration":           4,
        "03. enablement/training":     3,
        "04. uat":                     9,
        "05. prep for go-live":        6,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "Payments": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     8,
        "05. prep for go-live":        8,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "Reconcile": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     7,
        "05. prep for go-live":        8,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "Reconcile PSP": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     7,
        "05. prep for go-live":        8,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "e-Invoicing": {
        "01. requirements and design": 1,
        "02. configuration":           4,
        "03. enablement/training":     3,
        "04. uat":                     9,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "SFTP Connector": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     7,
        "05. prep for go-live":        8,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
    "CC Statement Import": {
        "01. requirements and design": 1,
        "02. configuration":           2,
        "03. enablement/training":     3,
        "04. uat":                     7,
        "05. prep for go-live":        8,
        "06. go-live":                 10,
        "08. ready for support transition": 10,
    },
}

# Default total duration (weeks) for unknown product types
DEFAULT_DURATION_WEEKS = 10

# Phase weight for parallel load assessment
PHASE_WEIGHTS = {
    "00. onboarding":                    1.0,
    "01. requirements and design":       1.0,
    "02. configuration":                 2.0,
    "03. enablement/training":           2.5,
    "04. uat":                           3.0,
    "05. prep for go-live":              1.0,
    "06. go-live":                       3.0,
    "07. data migration":                1.0,
    "08. ready for support transition":  0.5,
    "09. phase 2 scoping":               1.0,
    "10. complete/pending final billing": 0.0,
    "11. on hold":                       0.25,
    "12. ps review":                     0.25,
}
PARALLEL_CONFLICT_THRESHOLD = 5.0  # sum of phase weights across concurrent projects

# FF scope map (mirrors utilization report)
FF_SCOPE_MAP = {
    # ZoneApps FF products only — Billing/Reporting/Payroll are T&M (scope from NS)
    # Keys sorted longest-first so more specific entries match before general ones
    "capture & e-invoicing": 30,
    "reconcile premium":     30,
    "approvals premium":     35,
    "reconcile 2.0":         20,
    "reconcile psp":         17,
    "cc statement import":    6,
    "sftp connector":        12,
    "e-invoicing":           15,
    "approvals":             17,
    "reconcile":             17,
    "capture":               20,
    "payments":              20,
}

HORIZON_MONTHS = 6  # how many months forward to project

# ── Helper functions ───────────────────────────────────────────────────────────
def _emp_location(name):
    v = EMPLOYEE_LOCATION.get(name)
    if v is None:
        return None
    return v[0] if isinstance(v, tuple) else v

def _emp_ps_region(name):
    if name in PS_REGION_OVERRIDE:
        return PS_REGION_OVERRIDE[name]
    loc = _emp_location(name)
    return PS_REGION_MAP.get(loc, "Unknown") if loc else "Unknown"

def _emp_active(name, period_str):
    v = EMPLOYEE_LOCATION.get(name)
    if v is None or not isinstance(v, tuple):
        return True
    _, start, end = v
    try:
        period = str(period_str)[:7]
        if start and period < start:
            return False
        if end and period > end:
            return False
    except Exception:
        pass
    return True

def _is_delivery_role(name):
    """True if the employee owns product delivery (not a pure PM)."""
    info = EMPLOYEE_ROLES.get(name, {})
    return info.get("role") not in ("Project Manager",)

def _resolve_ff_scope(project_type):
    """Return scoped hours for a FF project type."""
    pt = (project_type or "").strip().lower()
    for key in sorted(FF_SCOPE_MAP.keys(), key=len, reverse=True):
        if key in pt:
            return FF_SCOPE_MAP[key]
    return None

def _project_end_date(start_date, project_type):
    """Return estimated project completion date from start date + phase duration table."""
    pt = (project_type or "").strip()
    phase_map = PHASE_END_WEEKS.get(pt)
    if phase_map:
        total_weeks = max(phase_map.values())
    else:
        # Try case-insensitive match
        for k, v in PHASE_END_WEEKS.items():
            if k.lower() == pt.lower():
                total_weeks = max(v.values())
                break
        else:
            total_weeks = DEFAULT_DURATION_WEEKS
    return start_date + timedelta(weeks=total_weeks)

def _current_phase_end_date(start_date, current_phase, project_type):
    """Return estimated end date of the current phase."""
    pt = (project_type or "").strip()
    phase_map = None
    for k, v in PHASE_END_WEEKS.items():
        if k.lower() == pt.lower():
            phase_map = v
            break
    if phase_map is None:
        phase_map = {}

    phase_key = (current_phase or "").strip().lower()
    # Find this phase's end week
    phase_end_wk = phase_map.get(phase_key)
    if phase_end_wk is None:
        # Use total duration as fallback
        phase_end_wk = max(phase_map.values()) if phase_map else DEFAULT_DURATION_WEEKS

    return start_date + timedelta(weeks=phase_end_wk)

def _months_range(n=HORIZON_MONTHS):
    """Return list of (year, month) tuples for the next n months from today."""
    today = date.today()
    months = []
    for i in range(n):
        d = today + relativedelta(months=i)
        months.append((d.year, d.month))
    return months

def _month_label(year, month):
    return datetime(year, month, 1).strftime("%b %Y")

def _territory_to_ps_region(territory):
    """Map a territory string to NOAM / EMEA / APAC."""
    t = (territory or "").strip().upper()
    if "NOAM" in t or "NORTH AM" in t or "US" in t or "CANADA" in t:
        return "NOAM"
    if "EMEA" in t:
        return "EMEA"
    if "APAC" in t or "ASIA" in t or "PACIFIC" in t:
        return "APAC"
    return t if t else "Unknown"

def fmt_hrs(n):
    return f"{n:,.2f}".rstrip("0").rstrip(".") if "." in f"{n:,.2f}" else f"{n:,}"

# ── SS DRS loader ──────────────────────────────────────────────────────────────
SS_COL_MAP = {
    "project id":        "project_id",
    "project name":      "project_name",
    "name":              "project_name",
    "project":           "project_name",
    "consultant":        "consultant",
    "assigned to":       "consultant",
    "resource":          "consultant",
    "phase":             "phase",
    "project phase":     "phase",
    "milestone":         "phase",
    "project type":      "project_type",
    "type":              "project_type",
    "billing type":      "billing_type",
    "status":            "status",
    "project status":    "status",
    "territory":         "territory",
    "project manager":   "consultant",
    "pm":                "project_manager",
    "status":            "status",
    "project status":    "status",
    "start date":        "start_date",
    "project start date":"start_date",
    "client health":     "client_health",
    "risk":              "risk",
    "risk flag":         "risk",
}

def load_ss(file):
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    df.columns = [c.strip().lower() for c in df.columns]
    rename = {}
    for col in df.columns:
        if col in SS_COL_MAP:
            rename[col] = SS_COL_MAP[col]
    df = df.rename(columns=rename)
    # Normalise phase
    if "phase" in df.columns:
        df["phase"] = df["phase"].astype(str).str.strip().str.lower()
    # Parse start_date
    if "start_date" in df.columns:
        df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
    # Keep all billing types — both FF and T&M consume consultant capacity
    # (T&M unassigned demand is handled separately via NS report)
    # Exclude inactive rows — phase-based OR status-based (mirrors WHS logic)
    inactive_phases  = {"10. complete/pending final billing", "11. on hold", "12. ps review"}
    onhold_statuses  = {"on-hold", "on hold", "onhold", "on_hold"}
    phase_inactive   = df["phase"].isin(inactive_phases) if "phase" in df.columns else pd.Series(False, index=df.index)
    status_inactive  = df["status"].astype(str).str.strip().str.lower().isin(onhold_statuses) if "status" in df.columns else pd.Series(False, index=df.index)
    inactive_mask    = phase_inactive | status_inactive
    dropped_df       = df[inactive_mask].copy()
    # Tag reason for each dropped row
    if not dropped_df.empty:
        def _drop_reason(row):
            reasons = []
            if "phase" in row and str(row["phase"]).strip().lower() in inactive_phases:
                reasons.append(f"phase: {row['phase']}")
            if "status" in row and str(row["status"]).strip().lower() in onhold_statuses:
                reasons.append(f"status: {row['status']}")
            return " | ".join(reasons) if reasons else "unknown"
        dropped_df["Filter Reason"] = dropped_df.apply(_drop_reason, axis=1)
    df = df[~inactive_mask]
    return df, dropped_df


# ── SFDC Closed Won loader ─────────────────────────────────────────────────────
SFDC_COL_MAP = {
    "opportunity name":   "opp_name",
    "opportunity":        "opp_name",
    "account name":       "account_name",
    "account":            "account_name",
    "close date":         "close_date",
    "closed date":        "close_date",
    "stage":              "stage",
    "territory":          "territory",
    "region":             "territory",
    "product":            "project_type",
    "products":           "project_type",
    "project type":       "project_type",
    "product family":     "project_type",
    "ps sow hours":       "ps_hours",
    "ps hours":           "ps_hours",
    "sow hours":          "ps_hours",
}

def load_sfdc(file):
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    df.columns = [str(col).strip().lower() for col in df.columns]
    rename = {col: SFDC_COL_MAP[col] for col in df.columns if col in SFDC_COL_MAP}
    df = df.rename(columns=rename)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]
    # Sanitise strings
    for _scol in ["opp_name", "account_name", "stage", "territory", "project_type"]:
        if _scol in df.columns:
            df[_scol] = df[_scol].fillna("").astype(str).str.strip()
    # Parse dates
    if "close_date" in df.columns:
        df["close_date"] = pd.to_datetime(df["close_date"], errors="coerce")
    if "ps_hours" in df.columns:
        df["ps_hours"] = pd.to_numeric(df["ps_hours"], errors="coerce")
    # Derive PS region
    if "territory" in df.columns:
        df["ps_region"] = df["territory"].apply(_territory_to_ps_region)
    # Derive product family
    if "project_type" in df.columns:
        df["product_family"] = df["project_type"].apply(_ns_pt_to_family)
    # Display name mirrors NS format: Account Name : Opp Name
    df["project_name"] = df.apply(
        lambda r: f"{r.get('account_name', '')} : {r.get('opp_name', '')}".strip(" :"),
        axis=1
    )
    return df

# ── NS Unassigned Projects loader ──────────────────────────────────────────────
NS_COL_MAP = {
    "project id":          "project_id",
    "project":             "project_name",
    "project name":        "project_name",
    "name":                "project_name",
    "customer : project":  "project_name",
    "customer":            "customer",
    "customer name":       "customer",
    "account":             "customer",
    "account name":        "customer",
    "territory":           "territory",
    "billing country":     "billing_country",
    "status":              "status",
    "billing type":        "billing_type",
    "project type":        "project_type",
    "signed date":         "signed_date",
    "project outreach":    "outreach_date",
    "start date":          "start_date",
    "t&m scope":           "tm_scope",
    "orig estimated work": "tm_scope",
}

def load_ns_unassigned(file):
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    df.columns = [str(c).strip().lower() for c in df.columns]
    rename = {}
    for col in df.columns:
        if col in NS_COL_MAP:
            rename[col] = NS_COL_MAP[col]
    df = df.rename(columns=rename)
    # Drop duplicate columns (e.g. both "project id" and "internal id" map to project_id)
    df = df.loc[:, ~df.columns.duplicated(keep="first")]
    # If project_name still not resolved, scan for likely column
    if "project_name" not in df.columns:
        for _candidate in df.columns:
            if "project" in _candidate and "type" not in _candidate and "id" not in _candidate:
                df = df.rename(columns={_candidate: "project_name"})
                break
    # Fill NaN in string columns before any string operations
    for _scol in ["project_name", "customer", "project_type", "billing_type", "territory",
                  "billing_country", "status", "project_id"]:
        if _scol in df.columns:
            df[_scol] = df[_scol].fillna("").astype(str).str.strip()
    # Parse dates
    for dcol in ["signed_date", "outreach_date", "start_date"]:
        if dcol in df.columns:
            df[dcol] = pd.to_datetime(df[dcol], errors="coerce")
    # Numeric T&M scope
    if "tm_scope" in df.columns:
        df["tm_scope"] = pd.to_numeric(df["tm_scope"], errors="coerce")

    # Derive scoped hours — T&M from NS column, FF from lookup table
    df["scoped_hours"] = df.apply(lambda r: (
        r.get("tm_scope") if str(r.get("billing_type", "")).strip().lower()
        in ("time & material", "time and material", "t&m", "tm")
        else _resolve_ff_scope(str(r.get("project_type", "")))
    ), axis=1)
    # Derive PS region — territory first, fall back to billing_country if blank
    def _get_ps_region(row):
        t = str(row.get("territory", "")).strip()
        if t and t.lower() not in ("", "nan", "none"):
            region = _territory_to_ps_region(t)
            if region != "Unknown":
                return region
        # Fallback: billing country
        bc = str(row.get("billing_country", "")).strip()
        return _territory_to_ps_region(bc) if bc and bc.lower() not in ("", "nan", "none") else "Unknown"

    df["ps_region"] = df.apply(_get_ps_region, axis=1)
    return df

# ── Projection engine ─────────────────────────────────────────────────────────
# WHS bands — mirrors Workload Health Score page exactly
def _whs_band(score):
    if score <= 25:  return "Low",    "#C6EFCE", "#276221"
    if score <= 60:  return "Medium", "#FFEB9C", "#9C6500"
    return                  "High",   "#FFC7CE", "#9C0006"

def _client_health_mult(responsiveness, sentiment):
    r = str(responsiveness).strip().lower() if pd.notna(responsiveness) else ""
    s = str(sentiment).strip().lower() if pd.notna(sentiment) else ""
    if ("negative" in s or "unresponsive" in r) and ("negative" in s and "unresponsive" in r):
        return 1.3
    if "negative" in s or "unresponsive" in r or "not responding" in r:
        return 1.15
    return 1.0

def _risk_mult(risk):
    r = str(risk).strip().lower() if pd.notna(risk) else ""
    if "high" in r:   return 1.2
    if "medium" in r: return 1.1
    return 1.0

# ── Phase sequence helper ─────────────────────────────────────────────────────
# Ordered phase list for sequence progression
PHASE_SEQUENCE = [
    "00. onboarding",
    "01. requirements and design",
    "02. configuration",
    "03. enablement/training",
    "04. uat",
    "05. prep for go-live",
    "06. go-live",
    "07. data migration",
    "08. ready for support transition",
    "09. phase 2 scoping",
    "10. complete/pending final billing",
]

def _projected_phase(current_phase, project_type, start_date, target_month):
    """
    Given a project's current phase, product type, and start date,
    estimate which phase the project will be in during target_month (year, month tuple).
    Uses PHASE_END_WEEKS table. Returns projected phase string.
    """
    from datetime import date
    today = date.today()

    # Find phase duration map — fuzzy match since SS project_type may include
    # prefixes like "ZoneApp: Approvals" vs PHASE_END_WEEKS key "Approvals"
    pt = (project_type or "").strip().lower()
    for prefix in ("zoneapp: ", "zonebill: ", "zonepay: ", "za - "):
        if pt.startswith(prefix):
            pt = pt[len(prefix):]
            break
    phase_map = None
    for k, v in PHASE_END_WEEKS.items():
        if k.lower() == pt:
            phase_map = v
            break
    if phase_map is None:
        for k, v in PHASE_END_WEEKS.items():
            if k.lower() in pt or pt in k.lower():
                phase_map = v
                break

    if phase_map is None or start_date is None:
        return current_phase  # fallback: hold current phase

    # Target date = first day of target month
    target_date = date(target_month[0], target_month[1], 1)

    # Normalise start_date to plain date — handle all types including numpy/datetime
    try:
        sd = pd.to_datetime(start_date)
        if pd.isna(sd):
            return current_phase
        start_date = sd.date()
    except Exception:
        return current_phase

    weeks_elapsed = (target_date - start_date).days / 7.0

    # Find which phase we'll be in at that week count
    # Build sorted list of (end_week, phase_name)
    phase_ends = sorted(
        [(wk, ph) for ph, wk in phase_map.items()],
        key=lambda x: x[0]
    )
    projected = "10. complete/pending final billing"  # default if past all phases
    for end_wk, ph in phase_ends:
        if weeks_elapsed <= end_wk:
            projected = ph
            break

    return projected


def project_consultant_availability(ss_df, months):
    """
    Compute WHS score per consultant per month using phase-sequence forecasting.
    For each project, projects which phase it will be in during each future month
    using the PHASE_END_WEEKS duration table + project start date.
    Score = sum of (projected_phase_weight * client_health_mult * risk_mult)
    across all active projects for that month.
    Bands: Low<=25, Medium<=60, High>60.
    """
    results   = {}
    conflicts = []

    if ss_df is None or ss_df.empty:
        return results, conflicts
    if "consultant" not in ss_df.columns:
        return results, conflicts

    complete_phases = {"10. complete/pending final billing", "12. ps review"}

    # Build list of active project rows per consultant
    consultant_projects = {}  # name -> list of project dicts

    for _, row in ss_df.iterrows():
        consultant = str(row.get("consultant", "")).strip()
        if not consultant or consultant.lower() in ("nan", "none", "", "0", "unassigned"):
            continue
        if not _is_delivery_role(consultant):
            continue

        current_phase = str(row.get("phase", "")).strip().lower()
        project_type  = str(row.get("project_type", "")).strip()
        start_date    = row.get("start_date")
        ch_mult       = _client_health_mult(row.get("client_health", ""), row.get("risk", ""))
        risk_mult     = _risk_mult(row.get("risk", ""))

        if consultant not in consultant_projects:
            consultant_projects[consultant] = []
        consultant_projects[consultant].append({
            "current_phase": current_phase,
            "project_type":  project_type,
            "start_date":    start_date,
            "ch_mult":       ch_mult,
            "risk_mult":     risk_mult,
        })

    # For each consultant, compute projected WHS score per month
    for consultant, projects in consultant_projects.items():
        results[consultant] = {}
        for mo in months:
            monthly_score   = 0.0
            active_proj_cnt = 0
            for proj in projects:
                proj_phase = _projected_phase(
                    proj["current_phase"],
                    proj["project_type"],
                    proj["start_date"],
                    mo,
                )
                if proj_phase in complete_phases:
                    continue  # project complete in this month — no load
                weight  = PHASE_WEIGHTS.get(proj_phase, 1.0)
                monthly_score   += round(weight * proj["ch_mult"] * proj["risk_mult"], 2)
                active_proj_cnt += 1

            band, _, _ = _whs_band(monthly_score)
            status = "free" if band == "Low" else "partial" if band == "Medium" else "busy"
            results[consultant][mo] = (status, round(monthly_score, 1), active_proj_cnt)

    return results, conflicts


# Maps NS project type keywords -> EMPLOYEE_ROLES product family names
NS_PRODUCT_FAMILY_MAP = [
    (["bill"],                          "Billing"),
    (["rpt", "report"],                 "Reporting"),
    (["payroll", "zep"],               "Payroll"),
    (["zoneapp", "capture", "approvals", "reconcile", "payments",
       "e-invoicing", "sftp", "cc statement", "za -"],  "Apps"),
]

def _ns_pt_to_family(project_type):
    """Map NS/SFDC project type or product field to family. Keyword-based, order matters."""
    pt = str(project_type).lower()
    # Payroll — zonepay alone = Payroll; zonepayments = Apps
    if "zonepay" in pt and "payment" not in pt:
        return "Payroll"
    if "payroll" in pt or "zep" in pt:
        return "Payroll"
    # Billing
    if "bill" in pt or "zonebill" in pt:
        return "Billing"
    # Reporting
    if "rpt" in pt or "report" in pt or "zonerpt" in pt:
        return "Reporting"
    # Apps — ZoneApp products + common SFDC product names
    if "zoneapp" in pt or "za -" in pt:
        return "Apps"
    if "zonepayments" in pt or ("payments" in pt and "payroll" not in pt):
        return "Apps"
    if any(kw in pt for kw in ["capture", "approvals", "reconcile",
                                "e-invoicing", "einvoic", "sftp",
                                "cc statement", "zone app", "apps"]):
        return "Apps"
    return "Other"


def _consultant_handles_family(consultant, family):
    """Return True if consultant's products include the given product family."""
    info = EMPLOYEE_ROLES.get(consultant, {})
    products = info.get("products", [])
    if not products:
        return False
    if "All" in products:
        return True
    pl = [p.lower() for p in products]
    if family == "Billing":
        return any("bill" in p for p in pl)
    if family == "Reporting":
        return any("rpt" in p or "report" in p for p in pl)
    if family == "Payroll":
        return any("payroll" in p or "zep" in p for p in pl)
    if family == "Apps":
        return any(kw in p for p in pl
                   for kw in ["capture", "approvals", "reconcile", "payments",
                               "e-invoicing", "sftp", "cc statement"])
    return False

def get_available_consultants(availability, months, product_type, ps_region, start_month):
    """
    Return list of consultants who:
    1. Have product_type in their EMPLOYEE_ROLES products list
    2. Are in the same PS region
    3. Have WHS score <= 60 (Medium or below) in start_month
    Sorted by score ascending (lowest load first).
    """
    from datetime import date
    today_period = date.today().strftime("%Y-%m")
    candidates = []
    for consultant, months_map in availability.items():
        # Active check — exclude leavers
        if not _emp_active(consultant, today_period):
            continue
        # Exclude Developers — matching logic TBD, removed from suggestions for now
        info_check = EMPLOYEE_ROLES.get(consultant, {})
        if info_check.get("role") == "Developer":
            continue
        # Product match — map NS project type to family, then check consultant
        family = _ns_pt_to_family(product_type)
        if not _consultant_handles_family(consultant, family):
            continue
        # Region match
        if _emp_ps_region(consultant) != ps_region:
            continue
        # WHS score in start month
        val = months_map.get(start_month)
        if val is None:
            continue
        status, score, _ = val if isinstance(val, tuple) else (val, 999, 0)
        if score <= 60:
            candidates.append((consultant, score))

    candidates.sort(key=lambda x: x[1])
    return candidates

def build_demand_summary(ns_df, months):
    """Summarise unassigned project demand by region and month (based on planned start date)."""
    if ns_df is None or ns_df.empty:
        return {}

    demand = {}  # region -> {month_tuple: {"count": n, "hours": h, "projects": []}}
    for _, row in ns_df.iterrows():
        region = str(row.get("ps_region", "Unknown")).strip()
        start_dt = row.get("start_date")
        if pd.isna(start_dt) or start_dt is None:
            # Use outreach date as proxy
            start_dt = row.get("outreach_date")
        if pd.isna(start_dt) or start_dt is None:
            continue
        start_d = start_dt.date() if hasattr(start_dt, "date") else start_dt
        mo_key = (start_d.year, start_d.month)
        if mo_key not in [m for m in months]:
            continue
        scoped_hrs = row.get("scoped_hours") or 0
        project_name = str(row.get("project_name", "")).strip()
        project_type = str(row.get("project_type", "")).strip()

        if region not in demand:
            demand[region] = {}
        if mo_key not in demand[region]:
            demand[region][mo_key] = {"count": 0, "hours": 0, "projects": []}
        demand[region][mo_key]["count"] += 1
        demand[region][mo_key]["hours"] += scoped_hrs
        demand[region][mo_key]["projects"].append(f"{project_name} ({project_type})")

    return demand

# ── Excel export ───────────────────────────────────────────────────────────────
def _status(val):
    """Unpack status from availability tuple (status, score, count) or plain string."""
    return val[0] if isinstance(val, tuple) else val

def build_excel(availability, conflicts, ns_df, months, ss_df):
    wb = Workbook()
    navy_fill   = PatternFill("solid", fgColor=NAVY)
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    amber_fill  = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    blue_fill   = PatternFill("solid", fgColor="BDD7EE")
    gray_fill   = PatternFill("solid", fgColor=LTGRAY)
    white_fill  = PatternFill("solid", fgColor=WHITE)
    hdr_font    = Font(bold=True, color=WHITE, name="Calibri", size=11)
    bold_font   = Font(bold=True, name="Calibri", size=10)
    std_font    = Font(name="Calibri", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_row(ws, row, values, fill=None):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hdr_font
            c.fill = fill or navy_fill
            c.alignment = center
            c.border = border

    def data_cell(ws, row, col, val, fill=None, bold=False, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = bold_font if bold else std_font
        if fill:
            c.fill = fill
        c.alignment = align or left
        c.border = border
        return c

    month_labels = [_month_label(y, m) for y, m in months]

    # ── Tab 1: Unassigned Projects ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Unassigned Projects")
    hdr_row(ws4, 1, ["Project", "Project Type", "Billing Type", "Territory", "PS Region",
                     "Signed Date", "Outreach Deadline", "Planned Start", "Scoped Hours", "Urgency"])
    for col_i, w in enumerate([35, 22, 15, 15, 10, 14, 18, 14, 14, 12], 1):
        ws4.column_dimensions[get_column_letter(col_i)].width = w

    if ns_df is not None and not ns_df.empty:
        today_d = date.today()
        for row_i, (_, r) in enumerate(ns_df.iterrows(), 2):
            outreach = r.get("outreach_date")
            outreach_d = outreach.date() if pd.notna(outreach) and hasattr(outreach, "date") else None
            days_to_outreach = (outreach_d - today_d).days if outreach_d else None
            urgency = ("Overdue" if days_to_outreach is not None and days_to_outreach < 0
                      else "This Week" if days_to_outreach is not None and days_to_outreach <= 7
                      else "This Month" if days_to_outreach is not None and days_to_outreach <= 30
                      else "Upcoming")
            urg_fill = (red_fill if urgency == "Overdue" else
                       amber_fill if urgency in ("This Week", "This Month") else white_fill)

            vals = [
                r.get("project_name", ""),
                r.get("project_type", ""),
                r.get("billing_type", ""),
                r.get("territory", ""),
                r.get("ps_region", ""),
                r.get("signed_date").strftime("%Y-%m-%d") if pd.notna(r.get("signed_date")) else "",
                outreach_d.strftime("%Y-%m-%d") if outreach_d else "",
                r.get("start_date").strftime("%Y-%m-%d") if pd.notna(r.get("start_date")) else "",
                fmt_hrs(r.get("scoped_hours") or 0),
                urgency,
            ]
            for col_i, val in enumerate(vals, 1):
                fill = urg_fill if col_i == 10 else white_fill
                data_cell(ws4, row_i, col_i, val, fill=fill,
                         align=center if col_i in (5, 6, 7, 8, 9, 10) else left)



    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── UI ─────────────────────────────────────────────────────────────────────────
def main():
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;600;700&display=swap" rel="stylesheet">
        <style>
            html, body, [class*="css"] { font-family: 'Manrope', sans-serif !important; }
            h1, h2, h3, .stMarkdown, .stDataFrame, label, button { font-family: 'Manrope', sans-serif !important; }
        </style>
        <div style='background-color:#1e2c63;padding:24px 32px;border-radius:8px;margin-bottom:24px;font-family:Manrope,sans-serif'>
            <h1 style='color:white;margin:0;font-size:28px;font-family:Manrope,sans-serif'>Capacity Outlook</h1>
            <p style='color:#aac4d0;margin:6px 0 0 0;font-size:14px;font-family:Manrope,sans-serif'>
                Project consultant availability over the next 6 months — active FF projects from Smartsheets + unassigned closed deals from NetSuite
            </p>
        </div>
    """, unsafe_allow_html=True)

    with st.expander("Exclusions & Limitations", expanded=False):
        st.markdown("""
        <style>
        .excl-list { font-family: Manrope, sans-serif; font-size: 13px; color: #4a5568; line-height: 1.8; padding-left: 4px; }
        .excl-list li { margin-bottom: 4px; }
        </style>
        <ul class='excl-list'>
            <li><b>Project Managers</b> are excluded from capacity calculations — they do not own product delivery.</li>
            <li><b>Projection accuracy</b> depends on SS project start dates being kept current by consultants and PMs.</li>
            <li><b>T&M active projects</b> are excluded from the projection engine — only FF active projects are used. T&M demand is captured via the NS Unassigned Projects report.</li>
            <li><b>Learning products</b> (Lappin: Capture/Reconcile; Longalong: Billing) are surfaced as future capacity only and not included in current assignment matching.</li>
            <li><b>Parallel conflicts</b> are flagged when two concurrent projects have a combined phase weight of 5.0 or higher (e.g. UAT + Go-Live simultaneously).</li>
        </ul>
        """, unsafe_allow_html=True)

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Step 1 — Upload Smartsheets DRS Export")
        st.caption("Required: Project Name, Consultant, Phase, Project Type, Billing Type, Start Date")
        ss_file = st.file_uploader("Drop SS DRS file here", type=["xlsx", "xls", "csv"], key="ss_cap_p3")
    with col2:
        st.subheader("Step 2 — Upload NS Unassigned Projects")
        st.caption("Required: Project, Territory, Billing Type, Project Type, Signed Date, Project Outreach, Start Date, T&M Scope")
        st.markdown("[Open NS Unassigned Projects Report ↗](https://3838224.app.netsuite.com/app/common/search/searchresults.nl?searchid=68439&whence=)")
        ns_file = st.file_uploader("Drop NS Unassigned Projects file here", type=["xlsx", "xls", "csv"], key="ns_unassigned_p3")

    ss_df         = None
    ss_raw_df     = None  # unfiltered — for project counts
    ss_dropped_df = None  # rows excluded by inactive filter
    ns_df         = None

    if ss_file:
        try:
            ss_df, ss_dropped_df = load_ss(ss_file)
            # Raw load for counts (rewind file)
            ss_file.seek(0)
            ss_raw_df = pd.read_excel(ss_file) if not ss_file.name.endswith(".csv") else pd.read_csv(ss_file)
            ss_raw_df.columns = [c.strip().lower() for c in ss_raw_df.columns]
            ss_raw_df = ss_raw_df.rename(columns={
                col: SS_COL_MAP[col] for col in ss_raw_df.columns if col in SS_COL_MAP
            })
            st.success(f"Loaded {len(ss_df):,} active project rows from SS DRS.")
            if ss_dropped_df is not None and not ss_dropped_df.empty:
                with st.expander(f"⚠️ {len(ss_dropped_df):,} rows excluded by filter — click to review"):
                    drop_cols = ["project_name", "consultant", "phase", "status", "Filter Reason"]
                    show_drop = ss_dropped_df[
                        [col for col in drop_cols if col in ss_dropped_df.columns]
                    ].rename(columns={
                        "project_name": "Project",
                        "consultant":   "Consultant",
                        "phase":        "Phase",
                        "status":       "Status",
                    })
                    st.dataframe(show_drop, hide_index=True, use_container_width=True)
        except Exception as e:
            st.error(f"Error loading SS file: {e}")

    if ns_file:
        try:
            ns_df = load_ns_unassigned(ns_file)
            st.success(f"Loaded {len(ns_df):,} unassigned projects from NS.")
        except Exception as e:
            st.error(f"Error loading NS file: {e}")

    # Step 3 — SFDC Closed Won
    st.markdown("---")
    st.subheader("Step 3 — Upload SFDC Closed Won PS Opps (Optional)")
    st.caption("Required: Opportunity Name, Account Name, Close Date, Territory, Product, PS SOW Hours")
    sfdc_col1, sfdc_col2 = st.columns([2, 1])
    with sfdc_col1:
        sfdc_file = st.file_uploader("Drop SFDC Closed Won file here", type=["xlsx", "xls", "csv"], key="sfdc_cap_p3")
    with sfdc_col2:
        sfdc_buffer_weeks = st.slider("Weeks from Close Date to PS Start", min_value=1, max_value=12, value=4,
                                       help="Estimated lag between SFDC close date and PS project kick-off")

    sfdc_df = None
    if sfdc_file:
        try:
            sfdc_df = load_sfdc(sfdc_file)
            st.success(f"Loaded {len(sfdc_df):,} closed won opps from SFDC.")
        except Exception as e:
            st.error(f"Error loading SFDC file: {e}")

    if not ss_file and not ns_file:
        st.info("Upload SS DRS to project current capacity. Add NS Unassigned Projects to overlay demand.")
        return

    months = _months_range(HORIZON_MONTHS)
    month_labels = [_month_label(y, m) for y, m in months]

    availability, conflicts = project_consultant_availability(ss_df, months)



    # ── Metrics row ────────────────────────────────────────────────────────────
    today_str = datetime.today().strftime("%Y-%m")
    active_consultants = sum(
        1 for name, info in EMPLOYEE_ROLES.items()
        if info.get("role") not in ("Project Manager",)
        and _emp_active(name, today_str)
        and not info.get("util_exempt")
    )
    # Project counts from SS
    # SS has one row per consultant per project — dedupe by project_id using project_name fallback
    _total_projects, _active_projects, _on_hold_projects = 0, 0, 0
    if ss_raw_df is not None and not ss_raw_df.empty:
        _id_col = "project_id" if "project_id" in ss_raw_df.columns else "project_name"
        _complete_phases = {"10. complete/pending final billing", "12. ps review"}
        _hold_phase      = "11. on hold"
        _onhold_statuses = {"on-hold", "on hold", "onhold", "on_hold"}

        # Per-project: collect all phases and statuses
        _grp = ss_raw_df.groupby(_id_col)
        _total_projects   = _grp.ngroups
        _on_hold_projects = 0
        _complete_count   = 0
        for _, rows in _grp:
            phases   = set(rows["phase"].astype(str).str.strip().str.lower().unique()) if "phase" in rows.columns else set()
            statuses = set(rows["status"].astype(str).str.strip().str.lower().unique()) if "status" in rows.columns else set()
            if statuses & _onhold_statuses or phases.issubset({_hold_phase}):
                _on_hold_projects += 1
            elif phases.issubset(_complete_phases):
                _complete_count += 1
        _active_projects = _total_projects - _on_hold_projects - _complete_count
    elif ss_df is not None and not ss_df.empty:
        _id_col = "project_id" if "project_id" in ss_df.columns else "project_name"
        _active_projects = ss_df[_id_col].nunique() if _id_col in ss_df.columns else len(ss_df)
        _total_projects  = _active_projects
        _on_hold_projects = 0
    currently_busy = sum(
        1 for c, months_map in availability.items()
        if _status(months_map.get(months[0], "free")) == "busy"
    )
    available_now = sum(
        1 for c, months_map in availability.items()
        if _status(months_map.get(months[0], "free")) in ("free", "partial")
    )
    unassigned_count = len(ns_df) if ns_df is not None else 0
    conflict_count = len(conflicts)


    # ── Product family counts from NS unassigned ──────────────────────────────
    def _classify_product(pt):
        pt_l = str(pt).lower()
        # zonepayments = Apps; zonepay alone = Payroll
        if "zonepayments" in pt_l or ("payments" in pt_l and "zonepay" not in pt_l):
            return "Apps"
        if "zoneapp" in pt_l or "za -" in pt_l:          return "Apps"
        if "bill" in pt_l:                                return "Billing"
        if "rpt" in pt_l or "report" in pt_l:             return "Reporting"
        if "zonepay" in pt_l or "payroll" in pt_l or "zep" in pt_l: return "Payroll"
        return "Other"

    # Cross-reference NS unassigned against SS DRS — if a project has a real
    # consultant assigned in SS, it's already staffed (NS lags by ~1 week)
    _staffed_in_ss = set()
    _ss_project_names = set()   # lowercased project names from SS — for SFDC matching
    _ss_account_names = set()   # lowercased account portion (before " : ") for fuzzy match
    if ss_df is not None and "consultant" in ss_df.columns:
        _invalid = {"", "nan", "none", "0", "unassigned"}
        _ss_active = ss_df[
            ~ss_df["consultant"].astype(str).str.strip().str.lower().isin(_invalid)
        ]
        if "project_id" in ss_df.columns:
            _staffed_in_ss = set(_ss_active["project_id"].astype(str).str.strip().unique())
        if "project_name" in ss_df.columns:
            # Use ALL SS rows — we want to know if the project exists
            _ss_project_names = set(
                ss_df["project_name"].astype(str).str.strip().str.lower().unique()
            ) - {"", "nan", "none"}
            # Build set of project names that are on-hold or closed in SS
            _closed_phases   = {"10. complete/pending final billing", "11. on hold", "12. ps review"}
            _onhold_statuses = {"on-hold", "on hold", "onhold", "on_hold"}
            _ss_closed_mask  = pd.Series([False] * len(ss_df), index=ss_df.index)
            if "phase" in ss_df.columns:
                _ss_closed_mask |= ss_df["phase"].astype(str).str.strip().str.lower().isin(_closed_phases)
            if "status" in ss_df.columns:
                _ss_closed_mask |= ss_df["status"].astype(str).str.strip().str.lower().isin(_onhold_statuses)
            _ss_closed_project_names = set(
                ss_df[_ss_closed_mask]["project_name"].astype(str).str.strip().str.lower().unique()
            ) - {"", "nan", "none"}
        # Customer/account names from SS — prefer explicit customer col, fallback to project name
        _ss_customer_col = "customer" if "customer" in ss_df.columns else "project_name"
        _ss_account_names = set(
            ss_df[_ss_customer_col].astype(str).str.strip().str.lower().unique()
        ) - {"", "nan", "none"}

    # Filter NS unassigned to true unassigned only
    if ns_df is not None and not ns_df.empty and "project_id" in ns_df.columns:
        _true_unassigned = ns_df[
            ~ns_df["project_id"].astype(str).str.strip().isin(_staffed_in_ss)
        ]
    else:
        _true_unassigned = ns_df if ns_df is not None else pd.DataFrame()

    # ── SFDC cross-reference and merge ────────────────────────────────────────
    # Build NS lookup sets for SFDC cross-reference
    _ns_project_names = set()
    _ns_customer_names = set()
    if not _true_unassigned.empty:
        if "project_name" in _true_unassigned.columns:
            _ns_project_names = set(
                _true_unassigned["project_name"].astype(str).str.strip().str.lower().unique()
            )
        # Prefer explicit customer col; fall back to project_name (current state)
        _ns_cust_col = "customer" if "customer" in _true_unassigned.columns else "project_name"
        _ns_customer_names = set(
            _true_unassigned[_ns_cust_col].astype(str).str.strip().str.lower().unique()
        ) - {"", "nan", "none"}

    _sfdc_rows = []
    if sfdc_df is not None and not sfdc_df.empty:
        from datetime import timedelta
        for _, row in sfdc_df.iterrows():
            opp_name     = str(row.get("opp_name", "")).strip()
            account_name = str(row.get("account_name", "")).strip()
            proj_name    = str(row.get("project_name", "")).strip()

            # Exact match: opp name in NS project names
            opp_lower = opp_name.lower()
            acct_lower = account_name.lower()

            # Exact exclude: opp name == NS project name OR SS project name
            ns_exact = opp_lower in _ns_project_names
            ss_exact = opp_lower in _ss_project_names if _ss_project_names else False
            exact_match = ns_exact or ss_exact

            # Fuzzy flag:
            # NS — account name == NS customer column (explicit customer field)
            ns_fuzzy = (not exact_match and acct_lower and
                        acct_lower in _ns_customer_names)
            # SS — account name found as substring within any SS project name
            # e.g. "3SI Security Systems" in "3SI Security Systems - ZB Optimization T&M"
            ss_fuzzy = (not exact_match and acct_lower and
                        any(acct_lower in ss_name for ss_name in _ss_project_names)
                        ) if _ss_project_names else False
            fuzzy_match = ns_fuzzy or ss_fuzzy

            if exact_match:
                continue  # already in NS or SS — skip

            # Check if fuzzy-matched SS project is on-hold or closed
            ss_closed_flag = (
                fuzzy_match and acct_lower and
                "_ss_closed_project_names" in dir() and
                any(acct_lower in ss_name for ss_name in _ss_closed_project_names)
            )

            # Derive estimated start date from close date + buffer
            close_dt = row.get("close_date")
            if pd.notna(close_dt):
                est_start = pd.Timestamp(close_dt) + timedelta(weeks=sfdc_buffer_weeks)
            else:
                est_start = None

            # Derive scoped hours — use PS SOW Hours; SFDC is T&M so no FF lookup
            ps_hours = row.get("ps_hours")
            if pd.isna(ps_hours) or not ps_hours:
                ps_hours = None  # show blank rather than FF estimate for T&M

            _sfdc_rows.append({
                "project_name":  proj_name,
                "customer":      account_name,
                "project_type":  str(row.get("project_type", "")),
                "billing_type":  "Time & Material",
                "territory":     str(row.get("territory", "")),
                "ps_region":     str(row.get("ps_region", "Unknown")),
                "signed_date":   close_dt,
                "start_date":    est_start,
                "outreach_date": None,
                "scoped_hours":  ps_hours,
                "source":        ("🔒 On Hold / Closed in SS" if ss_closed_flag
                                 else "⚠️ Possible Duplicate" if fuzzy_match
                                 else "SFDC"),
            })

    # Capture NS count before merge
    _ns_total = len(_true_unassigned) if _true_unassigned is not None else 0

    # Tag NS rows with source
    if not _true_unassigned.empty:
        _true_unassigned = _true_unassigned.copy()
        _true_unassigned["source"] = "NS"

    # Merge NS + SFDC
    if _sfdc_rows:
        _sfdc_frame = pd.DataFrame(_sfdc_rows)
        _true_unassigned = pd.concat(
            [_true_unassigned, _sfdc_frame], ignore_index=True
        ) if not _true_unassigned.empty else _sfdc_frame
    _ns_apps      = 0
    _ns_billing   = 0
    _ns_reporting = 0
    _ns_payroll   = 0
    _ns_billing_hrs   = 0.0
    _ns_reporting_hrs = 0.0
    _ns_payroll_hrs   = 0.0
    if not _true_unassigned.empty and "project_type" in _true_unassigned.columns:
        _families         = _true_unassigned["project_type"].apply(_classify_product)
        _ns_apps          = (_families == "Apps").sum()
        _ns_billing       = (_families == "Billing").sum()
        _ns_reporting     = (_families == "Reporting").sum()
        _ns_payroll       = (_families == "Payroll").sum()
        if "scoped_hours" in _true_unassigned.columns:
            _hrs = pd.to_numeric(_true_unassigned["scoped_hours"], errors="coerce").fillna(0)
            _ns_billing_hrs   = _hrs[_families == "Billing"].sum()
            _ns_reporting_hrs = _hrs[_families == "Reporting"].sum()
            _ns_payroll_hrs   = _hrs[_families == "Payroll"].sum()
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    def metric_card(label, value, sub=None, pill_color=None, hrs=None):
        if pill_color and sub:
            bottom = f"<div style='display:inline-block;margin-top:6px;padding:2px 10px;border-radius:999px;background:{pill_color}22;color:{pill_color};font-size:12px'>{sub}</div>"
        elif sub:
            bottom = f"<div style='font-size:12px;color:#888;margin-top:2px'>{sub}</div>"
        else:
            bottom = ""
        hrs_line = (f"<div style='font-size:16px;font-weight:600;color:#5a6a8a;"
                    f"font-family:Manrope,sans-serif;margin-top:2px;line-height:1.2'>{hrs}</div>"
                    ) if hrs else ""
        return (
            "<div style='background:#f8f9fb;border-radius:8px;padding:16px 20px;border:1px solid #e8eaed'>"
            f"<div style='font-size:12px;color:#a0a0a0;font-family:Manrope,sans-serif;margin-bottom:4px'>{label}</div>"
            f"<div style='font-size:28px;font-weight:700;color:#1e2c63;font-family:Manrope,sans-serif;line-height:1.1'>{value}</div>"
            f"{hrs_line}"
            f"{bottom}"
            "</div>"
        )

    _sfdc_total = len(_sfdc_rows) if "_sfdc_rows" in dir() and _sfdc_rows else 0
    _combined_total = _ns_total + _sfdc_total
    _metric_sub = f"NS: {_ns_total} · SFDC: {_sfdc_total}" if _sfdc_total > 0 else ""

    def _fmt_hrs(h):
        return f"{int(h):,} hrs" if h == int(h) else f"{h:,.1f} hrs"

    with m1: st.markdown(metric_card("Unassigned Projects", _combined_total,
                _metric_sub, "#27AE60" if _metric_sub else ("#E74C3C" if _combined_total > 0 else None)), unsafe_allow_html=True)
    with m2: st.markdown(metric_card("Apps Projects", _ns_apps), unsafe_allow_html=True)
    with m3: st.markdown(metric_card("Billing Projects", _ns_billing,
                hrs=_fmt_hrs(_ns_billing_hrs) if _ns_billing_hrs > 0 else None), unsafe_allow_html=True)
    with m4: st.markdown(metric_card("Reporting Projects", _ns_reporting,
                hrs=_fmt_hrs(_ns_reporting_hrs) if _ns_reporting_hrs > 0 else None), unsafe_allow_html=True)
    with m5: st.markdown(metric_card("Payroll Projects", _ns_payroll,
                hrs=_fmt_hrs(_ns_payroll_hrs) if _ns_payroll_hrs > 0 else None), unsafe_allow_html=True)
    with m6: st.markdown(metric_card("Available Consultants", available_now, "", "#27AE60" if available_now > 0 else "#E74C3C"), unsafe_allow_html=True)


    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

    # ── Tabs ───────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "Unassigned Projects",
        "Capacity Heatmap",
        "Consultant Detail",
        "Regional Rollup",
    ])

    # ── Tab 1: Heatmap ─────────────────────────────────────────────────────────
    with tab1:
        st.markdown("#### Unassigned Projects — Closed Deals Awaiting Staffing")
        st.caption("Suggested Consultants matched by: product type · PS region · WHS score ≤ Medium (≤60) in planned start month")
        if ns_df is None or ns_df.empty:
            st.info("Upload NS Unassigned Projects file to view demand.")
        else:
            today_d = date.today()
            # Use true unassigned (NS minus already-staffed in SS)
            display_df = _true_unassigned.copy() if not _true_unassigned.empty else ns_df.copy()

            # Urgency from outreach deadline
            if "outreach_date" in display_df.columns:
                def urgency(d):
                    if pd.isna(d): return "Unknown"
                    diff = (d.date() - today_d).days if hasattr(d, "date") else (d - today_d).days
                    return "Overdue" if diff < 0 else "This Week" if diff <= 7 else "This Month" if diff <= 30 else "Upcoming"
                display_df["Urgency"] = display_df["outreach_date"].apply(urgency)

            # Strip time from date columns
            for _dcol in ["signed_date", "outreach_date", "start_date"]:
                if _dcol in display_df.columns:
                    display_df[_dcol] = pd.to_datetime(display_df[_dcol], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")

            # Suggested consultants — match by product + region + WHS score in start month
            def _suggest(row):
                if not availability:
                    return "Upload SS DRS"
                pt       = str(row.get("project_type", "")).strip()
                region   = str(row.get("ps_region", "")).strip()
                start_dt = row.get("start_date")
                # Determine which month the project starts
                if start_dt and str(start_dt) not in ("", "nan", "NaT"):
                    try:
                        sd = pd.to_datetime(start_dt)
                        start_mo = (sd.year, sd.month)
                    except:
                        start_mo = months[0]
                else:
                    start_mo = months[0]
                # Use closest available month in our range
                if start_mo not in months:
                    future = [m for m in months if m >= start_mo]
                    start_mo = future[0] if future else months[-1]
                candidates = get_available_consultants(availability, months, pt, region, start_mo)
                if not candidates:
                    return "No match"
                # Return top 3 with scores
                return ", ".join(f"{name} ({score})" for name, score in candidates[:3])

            display_df["Suggested Consultants"] = display_df.apply(_suggest, axis=1)

            col_rename = {
                "project_name":   "Project",
                "project_type":   "Project Type",
                "billing_type":   "Billing Type",
                "territory":      "Territory",
                "ps_region":      "PS Region",
                "signed_date":    "Signed Date",
                "start_date":     "Planned Start",
                "outreach_date":  "Outreach Deadline",
                "scoped_hours":   "Scoped Hours",
            }
            # Explicit column order: Project first, dates in sequence, Urgency, Suggested last
            ordered_cols = [
                "source", "customer", "project_name", "project_type", "billing_type",
                "territory", "ps_region",
                "signed_date", "start_date", "outreach_date",
                "scoped_hours",
            ]
            col_rename["source"]   = "Source"
            col_rename["customer"] = "Account / Customer"
            show_cols = (
                [col for col in ordered_cols if col in display_df.columns]
                + (["Urgency"] if "Urgency" in display_df.columns else [])
                + ["Suggested Consultants"]
            )
            display_df = display_df[show_cols].rename(columns=col_rename)

            def color_urgency(val):
                if val == "Overdue":    return "background-color:#FFC7CE;color:#9C0006"
                if val == "This Week":  return "background-color:#FFEB9C;color:#9C6500"
                if val == "This Month": return "background-color:#FFEB9C;color:#9C6500"
                return ""

            def color_suggest(val):
                if val == "No match":       return "background-color:#FFC7CE;color:#9C0006"
                if val == "Upload SS DRS":  return "background-color:#FFEB9C;color:#9C6500"
                return "background-color:#C6EFCE;color:#276221"

            def color_source(val):
                v = str(val)
                if v == "NS":                       return "background-color:#E8F4FD;color:#1A5276"
                if v == "SFDC":                     return "background-color:#EBF5EB;color:#1E8449"
                if "On Hold" in v or "Closed" in v: return "background-color:#F2F2F2;color:#666666"
                if "Possible Duplicate" in v:       return "background-color:#FEFDE0;color:#9C6500"
                return ""

            style_cols = {}
            if "Source" in display_df.columns:
                style_cols["Source"] = color_source
            if "Urgency" in display_df.columns:
                style_cols["Urgency"] = color_urgency
            style_cols["Suggested Consultants"] = color_suggest

            styled_ns = display_df.style
            for col, fn in style_cols.items():
                styled_ns = styled_ns.applymap(fn, subset=[col])
            st.dataframe(styled_ns, hide_index=True, use_container_width=True)

    with tab2:
        st.markdown("#### Consultant Availability — Next 6 Months")
        st.caption("Green = Available · Amber = Light load (low-weight phase) · Red = Busy")

        st.caption("Workload score = phase weight × client health × risk, summed across active projects per month.  "
                   "Scores are projected forward using the phase duration table.  "
                   "Low ≤ 25 · Medium ≤ 60 · High > 60")

        if not availability:
            st.info("No active project data found. Upload SS DRS to generate heatmap.")
        else:
            heatmap_data = []
            for consultant in sorted(availability.keys()):
                region = _emp_ps_region(consultant)
                row_data = {"Consultant": consultant, "Region": region}
                # Current month score (first month)
                first_val = availability[consultant].get(months[0], ("free", 0, 0))
                score = first_val[1] if isinstance(first_val, tuple) else 0
                proj_count = first_val[2] if isinstance(first_val, tuple) else 0
                row_data["WHS Score"] = score
                row_data["Projects"] = proj_count
                for mo, label in zip(months, month_labels):
                    val = availability[consultant].get(mo, ("free", 0, 0))
                    status, sc, _ = val if isinstance(val, tuple) else (val, 0, 0)
                    band, _, _ = _whs_band(sc)
                    row_data[label] = f"{band} ({sc})"
                heatmap_data.append(row_data)

            hm_df = pd.DataFrame(heatmap_data)

            def color_cell(val):
                v = str(val)
                if v.startswith("Low"):    return "background-color:#C6EFCE;color:#276221"
                if v.startswith("Medium"): return "background-color:#FFEB9C;color:#9C6500"
                if v.startswith("High"):   return "background-color:#FFC7CE;color:#9C0006"
                return ""

            styled = hm_df.style.applymap(color_cell, subset=month_labels)
            st.dataframe(styled, hide_index=True, use_container_width=True)

    # ── Tab 2: Consultant Detail ───────────────────────────────────────────────
    with tab3:
        st.markdown("#### Consultant Capacity Detail")
        detail_rows = []
        for name, info in sorted(EMPLOYEE_ROLES.items(), key=lambda x: x[0]):
            if info.get("role") == "Project Manager":
                continue
            if not _emp_active(name, today_str):
                continue
            region = _emp_ps_region(name)
            role = info.get("role", "Consultant")
            products = ", ".join(info.get("products", []) or ["All"])
            learning = ", ".join(info.get("learning", []))

            projects, free_from = "", "Available Now"
            if ss_df is not None and "consultant" in ss_df.columns:
                emp_rows = ss_df[ss_df["consultant"].astype(str).str.strip() == name]
                if not emp_rows.empty:
                    proj_list = emp_rows["project_name"].dropna().unique() if "project_name" in emp_rows.columns else []
                    projects = str(len(proj_list))
            # Get WHS score from availability
            whs_score = ""
            if name in availability:
                first_val = availability[name].get(months[0], ("free", 0, 0))
                if isinstance(first_val, tuple):
                    band, _, _ = _whs_band(first_val[1])
                    whs_score = f"{band} ({first_val[1]})"

            detail_rows.append({
                "Consultant": name,
                "Region": region,
                "Role": role,
                "WHS Score": whs_score,
                "Active Projects (#)": int(projects) if str(projects).isdigit() else 0,
                "Enabled Products": products,
                "Learning (Future)": learning,
            })

        if detail_rows:
            detail_df = pd.DataFrame(detail_rows)

            # Colour WHS Score cell only — not full row (avoids dark mode readability issues)
            def color_whs_cell(val):
                v = str(val)
                if v.startswith("Low"):    return "background-color:#C6EFCE;color:#276221"
                if v.startswith("Medium"): return "background-color:#FFEB9C;color:#9C6500"
                if v.startswith("High"):   return "background-color:#FFC7CE;color:#9C0006"
                return ""

            styled_detail = detail_df.style.applymap(color_whs_cell, subset=["WHS Score"])
            st.dataframe(styled_detail, hide_index=True, use_container_width=True)
        else:
            st.info("Upload SS DRS to populate consultant detail.")

    # ── Tab 3: Regional Rollup ─────────────────────────────────────────────────
    with tab4:
        st.markdown("#### Regional Capacity vs Demand")
        demand = build_demand_summary(ns_df, months) if ns_df is not None else {}

        for region in ["NOAM", "EMEA", "APAC"]:
            st.markdown(f"**{region}**")
            region_consultants = [c for c in availability if _emp_ps_region(c) == region]
            reg_demand = demand.get(region, {})

            rollup_rows = [
                {"Metric": "Consultants Available (Free/Light)"} | {
                    lbl: sum(1 for c in region_consultants
                             if availability[c].get(mo, "free") in ("free", "partial"))
                    for mo, lbl in zip(months, month_labels)
                },
                {"Metric": "Unassigned Projects (Demand)"} | {
                    lbl: reg_demand.get(mo, {}).get("count", 0)
                    for mo, lbl in zip(months, month_labels)
                },
                {"Metric": "Scoped Hours Demand"} | {
                    lbl: round(reg_demand.get(mo, {}).get("hours", 0), 1)
                    for mo, lbl in zip(months, month_labels)
                },
            ]
            rollup_df = pd.DataFrame(rollup_rows)

            def highlight_rollup(row):
                colors = [""] * len(row)
                if row["Metric"] == "Consultants Available (Free/Light)":
                    for i, col in enumerate(month_labels, 1):
                        v = row.get(col, 0)
                        colors[i] = "background-color:#C6EFCE;color:#276221" if v > 2 else "background-color:#FFEB9C;color:#9C6500" if v > 0 else "background-color:#FFC7CE;color:#9C0006"
                elif row["Metric"] == "Unassigned Projects (Demand)":
                    for i, col in enumerate(month_labels, 1):
                        v = row.get(col, 0)
                        colors[i] = "background-color:#FFC7CE;color:#9C0006" if v > 2 else "background-color:#FFEB9C;color:#9C6500" if v > 0 else "background-color:#C6EFCE;color:#276221"
                return colors

            styled_rollup = rollup_df.style.apply(highlight_rollup, axis=1)
            st.dataframe(styled_rollup, hide_index=True, use_container_width=True)

            # Drill-down: show unassigned projects making up the demand for this region
            if _true_unassigned is not None and not _true_unassigned.empty:
                region_unassigned = _true_unassigned[
                    _true_unassigned.get("ps_region", pd.Series(dtype=str)) == region
                ] if "ps_region" in _true_unassigned.columns else _true_unassigned[
                    _true_unassigned["ps_region"] == region
                ] if "ps_region" in _true_unassigned.columns else pd.DataFrame()

                if not region_unassigned.empty:
                    with st.expander(f"📋 {len(region_unassigned)} unassigned project(s) in {region} — click to view"):
                        drill_cols = ["project_name", "project_type", "start_date",
                                      "outreach_date", "scoped_hours"]
                        drill_rename = {
                            "project_name":  "Project",
                            "project_type":  "Project Type",
                            "start_date":    "Planned Start",
                            "outreach_date": "Outreach Deadline",
                            "scoped_hours":  "Scoped Hours",
                        }
                        drill_df = region_unassigned[
                            [col for col in drill_cols if col in region_unassigned.columns]
                        ].copy()
                        for _dcol in ["start_date", "outreach_date"]:
                            if _dcol in drill_df.columns:
                                drill_df[_dcol] = pd.to_datetime(
                                    drill_df[_dcol], errors="coerce"
                                ).dt.strftime("%Y-%m-%d").fillna("")
                        st.dataframe(
                            drill_df.rename(columns=drill_rename),
                            hide_index=True, use_container_width=True
                        )
            st.markdown("<div style='margin-bottom:12px'></div>", unsafe_allow_html=True)

    # ── Tab 1: Unassigned Projects ─────────────────────────────────────────────
    # ── Export ─────────────────────────────────────────────────────────────────
    st.markdown("---")
    if st.button("Download Capacity Report (Excel)", type="primary"):
        excel_buf = build_excel(availability, conflicts, ns_df, months, ss_df)
        st.download_button(
            label="Save Excel",
            data=excel_buf,
            file_name=f"capacity_outlook_{datetime.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

main()
