"""
Workload Health Score
Upload a Smartsheets DRS export + NetSuite time detail to generate
a weighted workload score per consultant across active FF projects.
"""
import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="FF Workload Score", page_icon=None, layout="wide")

# ── Constants ─────────────────────────────────────────────────────────────────
NAVY     = "1e2c63"
TEAL     = "4472C4"
WHITE    = "FFFFFF"
LTGRAY   = "F2F2F2"
RED      = "E74C3C"
GREEN    = "27AE60"
AMBER    = "F39C12"

# ── Employee / Region config (mirrors utilization report) ─────────────────────
EMPLOYEE_ROLES = {
    # Structure: name -> {role, products, learning, util_exempt (optional)}
    # roles: Consultant | Project Manager | Solution Architect | Developer
    # products: assignable product types
    # learning: in-training — surface as future capacity only, not current
    # util_exempt: True = excluded from utilization targets

    # ── Project Managers (no product delivery) ────────────────────────────────
    "Barrio, Nairobi":          {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Hughes, Madalyn":          {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Porangada, Suraj":       {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Cadelina":             {"role": "Project Manager",    "products": [],                                                                                      "learning": []},

    # ── Solution Architects ───────────────────────────────────────────────────
    "Bell, Stuart":            {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "DiMarco, Nicole R":         {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "Murphy, Conor":          {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},

    # ── Developer ─────────────────────────────────────────────────────────────
    "Church, Jason G":      {"role": "Developer",          "products": ["All"],                                                                                  "learning": [], "util_exempt": True},

    # ── Consultants ───────────────────────────────────────────────────────────
    "Arestarkhov, Yaroslav":     {"role": "Consultant",         "products": ["Billing", "Capture"],                                                                   "learning": []},
    "Carpen, Anamaria":          {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Centinaje, Rhodechild":       {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},
    "Cooke, Ellen":           {"role": "Consultant",         "products": ["Billing", "Payroll"],                                                                   "learning": []},
    "Dolha, Madalina":           {"role": "Consultant",         "products": ["Capture", "Reconcile", "CC Statement Import", "Reconcile PSP", "e-Invoicing"],          "learning": []},
    "Finalle-Newton, Jesse":  {"role": "Solution Architect", "products": ["Reporting"],                                                                            "learning": []},
    "Gardner, Cheryll L":         {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Hopkins, Chris":         {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Ickler, Georganne":          {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Isberg, Eric":          {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Jordanova, Marija":       {"role": "Consultant",         "products": ["Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],     "learning": []},
    "Lappin, Thomas":          {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": ["Capture", "Reconcile"]},
    "Longalong, Santiago":       {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": ["Billing"]},
    "Mohammad, Manaan":        {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Morris, Lisa":          {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "NAQVI, SYED":          {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Olson, Austin D":           {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Pallone, Daniel":         {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Raykova, Silvia":         {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Selvakumar, Sajithan":      {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Snee, Stefanie J":            {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Swanson":         {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},
    "Tuazon, Carol":          {"role": "Consultant",         "products": ["Payroll", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],       "learning": []},
    "Zoric, Ivan":           {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},

    "Dunn, Steven":           {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Law, Brandon":           {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Quiambao, Generalyn":    {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
        "Cruz, Daniel":           {"role": "Consultant",         "products": ["All"],                                                                                  "learning": []},
    # ── Leavers (historical data only) ────────────────────────────────────────
    "Alam, Laisa":          {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Chan, Joven":          {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Cloete, Bronwyn":      {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Eyong, Eyong":         {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Hamilton, Julie C":    {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Hernandez, Camila":    {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Rushbrook, Emma C":    {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Strauss, John W":      {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
}

# ── Employee roster — {name: (location, start_date, end_date)}
# start/end as "YYYY-MM" strings or None (None = no limit).
EMPLOYEE_LOCATION = {
    #  Name                        Location              Start       End
    # ── Active employees ──────────────────────────────────────────────────────
    "Arestarkhov, Yaroslav":  ("Czech Republic",      None,       None),
    "Barrio, Nairobi":        ("USA",                 None,       None),
    "Bell, Stuart":           ("USA",                 None,       None),
    "Cadelina":               ("Manila (PH)",         "2026-03",  None),
    "Carpen, Anamaria":       ("Spain",               None,       None),
    "Centinaje, Rhodechild":  ("Manila (PH)",         None,       None),
    "Church, Jason G":        ("USA",                 None,       None),
    "Cooke, Ellen":           ("Northern Ireland",    None,       None),
    "Cruz, Daniel":           ("Manila (PH)",         None,       None),
    "DiMarco, Nicole R":      ("USA",                 None,       None),
    "Dolha, Madalina":        ("Faroe Islands",       None,       None),
    "Finalle-Newton, Jesse":  ("USA",                 None,       None),
    "Gardner, Cheryll L":     ("USA",                 None,       None),
    "Hopkins, Chris":         ("USA",                 None,       None),
    "Hughes, Madalyn":        ("USA",                 None,       None),
    "Ickler, Georganne":      ("USA",                 None,       None),
    "Isberg, Eric":           ("USA",                 None,       None),
    "Jordanova, Marija":      ("North Macedonia",     None,       None),
    "Lappin, Thomas":         ("Northern Ireland",    None,       None),
    "Longalong, Santiago":    ("Manila (PH)",         None,       None),
    "Mohammad, Manaan":       ("Canada",              None,       None),
    "Morris, Lisa":           ("Sydney (NSW)",        None,       None),
    "Murphy, Conor":          ("USA",                 None,       None),
    "NAQVI, SYED":            ("Canada",              None,       None),
    "Olson, Austin D":        ("USA",                 None,       None),
    "Pallone, Daniel":        ("Sydney (NSW)",        None,       None),
    "Porangada, Suraj":       ("USA",                 None,       None),
    "Raykova, Silvia":        ("Netherlands",         None,       None),
    "Selvakumar, Sajithan":   ("Canada",              None,       None),
    "Snee, Stefanie J":       ("USA",                 None,       None),
    "Stone, Matt":            ("USA",                 None,       None),
    "Swanson":                ("USA",                 None,       None),
    "Tuazon, Carol":          ("Manila (PH)",         None,       None),
    "Zoric, Ivan":            ("Serbia",              None,       None),
    "Dunn, Steven":           ("USA",                 None,       None),
    "Law, Brandon":           ("USA",                 None,       None),
    "Quiambao, Generalyn":    ("Manila (PH)",         None,       None),
        # ── Leavers ───────────────────────────────────────────────────────────────
    "Alam, Laisa":            ("USA",                 None,       "2025-12"),
    "Chan, Joven":            ("Manila (PH)",         None,       "2025-12"),
    "Cloete, Bronwyn":        ("Netherlands",         None,       "2026-02"),
    "Eyong, Eyong":           ("USA",                 None,       "2025-12"),
    "Hamilton, Julie C":      ("USA",                 None,       "2026-01"),
    "Hernandez, Camila":      ("USA",                 None,       "2025-12"),
    "Rushbrook, Emma C":      ("Wales",               None,       "2025-12"),
    "Strauss, John W":        ("USA",                 None,       "2026-03"),
}

def _emp_location(name):
    """Return location string regardless of tuple or plain string."""
    v = EMPLOYEE_LOCATION.get(name)
    if v is None:
        return None
    return v[0] if isinstance(v, tuple) else v
PS_REGION_OVERRIDE = {
    "NAQVI, SYED":  "EMEA",
    "Cruz, Daniel": "NOAM",
    "Chan, Joven":    "NOAM",   # Manila-based but reports into NOAM
    "Rushbrook, Emma C": "EMEA",  # Wales → EMEA
}
PS_REGION_MAP = {
    "Sydney (NSW)":     "APAC",
    "Manila (PH)":      "APAC",
    "UK":               "EMEA",
    "Wales":            "EMEA",
    "Spain":            "EMEA",
    "Netherlands":      "EMEA",
    "Northern Ireland": "EMEA",
    "Faroe Islands":    "EMEA",
    "North Macedonia":  "EMEA",
    "Czech Republic":   "EMEA",
    "Serbia":           "EMEA",
    "USA":              "NOAM",
    "Canada":           "NOAM",
}

# ── Phase weights ─────────────────────────────────────────────────────────────
PHASE_WEIGHTS = {
    "00. onboarding":                    1.0,
    "01. requirements and design":       1.0,
    "02. configuration":                 2.0,
    "03. enablement/training":           2.5,
    "04. uat":                           3.0,
    "05. prep for go-live":              1.0,
    "06. go-live":                       3.0,
    "07. data migration":                1.0,
    "08. ready for support transition":  0.5,
    "09. phase 2 scoping":               1.0,
    "10. complete/pending final billing": 0.0,
    "11. on hold":                       0.25,
    "12. ps review":                     0.25,
}

# Phases excluded from active workload score
INACTIVE_PHASES = {"10. complete/pending final billing", "12. ps review"}

# ── Workload thresholds ───────────────────────────────────────────────────────
def _emp_active(name, period_str):
    """
    Return True if the employee should be included for the given period.
    period_str: "YYYY-MM" or "YYYYMM" or a pandas Period string.
    """
    v = EMPLOYEE_LOCATION.get(name)
    if v is None or not isinstance(v, tuple):
        return True
    _, start, end = v
    try:
        period = str(period_str)[:7]
        if start and period < start:
            return False
        if end and period > end:
            return False
    except Exception:
        pass
    return True

def _emp_role(name):
    """Return role string for an employee."""
    v = EMPLOYEE_ROLES.get(name)
    if v is None:
        return None
    return v["role"] if isinstance(v, dict) else v

def _emp_products(name, include_learning=False):
    """Return list of assignable products for an employee."""
    v = EMPLOYEE_ROLES.get(name)
    if v is None or not isinstance(v, dict):
        return []
    products = list(v.get("products", []))
    if include_learning:
        products += v.get("products_learning", [])
    return products

def _emp_util_exempt(name):
    """Return True if employee is exempt from utilization targets."""
    v = EMPLOYEE_ROLES.get(name)
    if isinstance(v, dict):
        return v.get("util_exempt", False)
    return False


def workload_level(score):
    if score <= 25:   return "Low",    GREEN
    if score <= 60:   return "Medium", AMBER
    return "High", RED

# ── Client health multiplier ──────────────────────────────────────────────────
def client_health_multiplier(responsiveness, sentiment):
    r = str(responsiveness).strip().lower() if pd.notna(responsiveness) else ""
    s = str(sentiment).strip().lower() if pd.notna(sentiment) else ""
    if ("negative" in s or "unresponsive" in r) and ("negative" in s and "unresponsive" in r):
        return 1.3
    if "negative" in s or "unresponsive" in r or "not responding" in r:
        return 1.15
    return 1.0

# ── Risk multiplier ───────────────────────────────────────────────────────────
def risk_multiplier(risk_level):
    r = str(risk_level).strip().lower() if pd.notna(risk_level) else ""
    if "high"   in r: return 1.2
    if "medium" in r: return 1.1
    return 1.0

# ── Employee → PS Region lookup ───────────────────────────────────────────────
def get_ps_region(name):
    if not name or str(name).strip().lower() in ("", "nan", "none"):
        return "Unknown"
    name = str(name).strip()
    if name in PS_REGION_OVERRIDE:
        return PS_REGION_OVERRIDE[name]
    loc = _emp_location(name)
    if not loc:
        last = name.split(",")[0].strip()
        loc = next((_emp_location(k) for k, v in EMPLOYEE_LOCATION.items() if k.split(",")[0].strip().lower() == last.lower()), None)
    return PS_REGION_MAP.get(loc, "Unknown") if loc else "Unknown"

# ── Phase weight lookup ───────────────────────────────────────────────────────
def get_phase_weight(phase):
    if not phase or str(phase).strip().lower() in ("", "nan", "none"):
        return 1.0, "Undefined"
    p = str(phase).strip().lower()
    if p in PHASE_WEIGHTS:
        return PHASE_WEIGHTS[p], str(phase).strip()
    return 1.0, str(phase).strip()

# ── SS column map ─────────────────────────────────────────────────────────────
SS_COL_MAP = {
    "project name":          "project_name",
    "project id":            "project_id",
    "overall rag":           "rag",
    "start date":            "start_date",
    "go live date":          "go_live_date",
    "% complete":            "pct_complete",
    "project type":          "project_type",
    "status":                "status",
    "project phase":         "phase",
    "client responsiveness": "client_responsiveness",
    "client sentiment":      "client_sentiment",
    "risk level":            "risk_level",
    "schedule health":       "schedule_health",
    "resource health":       "resource_health",
    "scope health":          "scope_health",
    "territory":             "territory",
    "actual hours":          "actual_hours",
    "budgeted hours":        "budgeted_hours",
    "budget":                "budget",
    "change order":          "change_order",
    "partner name":          "partner_name",
    "on hold reason":        "on_hold_reason",
    "project manager":       "project_manager",
    "billing type":          "billing_type",
    "billing":               "billing_type",
}

MILESTONE_COLS = [
    "Intro. Email Sent", "Standard Config Start", "Enablement Session",
    "Session #1", "Session #2", "Prod Cutover", "Hypercare Start",
    "Close Out Remaining Tasks", "UAT Signoff", "Transition to Support",
]

# NS column map (subset needed for PM join)
NS_COL_MAP = {
    "employee":        "employee",
    "name":            "employee",
    "project":         "project",
    "project name":    "project",
    "project id":      "project_id",
    "billing type":    "billing_type",
    "project manager": "project_manager",
    "date":            "date",
    "hours":           "hours",
    "quantity":        "hours",
}

# ── Excel helpers ─────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def style_cell(cell, bg, bold=False, align="left", font_color="000000", size=10):
    cell.fill = hex_fill(bg)
    cell.font = Font(bold=bold, color=font_color, name="Manrope", size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

def write_title(ws, title, ncols, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    style_cell(c, NAVY, bold=True, align="left", font_color=WHITE, size=12)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.append([""] * ncols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sc = ws.cell(2, 1, subtitle)
        style_cell(sc, "EBF5FB", bold=False, align="left", font_color="1A5276", size=9)
        ws.row_dimensions[2].height = 16

def style_header(ws, row, headers, bg=None):
    bg = bg or TEAL
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        style_cell(c, bg, bold=True, align="center", font_color=WHITE, size=10)
        ws.row_dimensions[row].height = 22

def term_age_flag(start_date_val):
    """Return ('expired'|'near_expiry'|None) based on project start date vs today.
    expired    = start date > 12 months ago (past PS term limit)
    near_expiry = start date >= 10 months ago (approaching limit)
    """
    try:
        if not start_date_val or (hasattr(start_date_val, '__class__') and str(start_date_val) in ('', 'nan', 'None', 'NaT')):
            return None
        sd = pd.to_datetime(start_date_val, errors="coerce")
        if pd.isna(sd):
            return None
        today = pd.Timestamp.today().normalize()
        months_old = (today.year - sd.year) * 12 + (today.month - sd.month)
        if months_old >= 12:
            return "expired"
        elif months_old >= 10:
            return "near_expiry"
        return None
    except Exception:
        return None


def rag_color(rag):
    r = str(rag).strip().lower() if pd.notna(rag) else ""
    if "red"    in r: return "FDECED"
    if "amber"  in r or "yellow" in r: return "FEF9E7"
    if "green"  in r: return "EAF9F1"
    return LTGRAY

def level_color(level):
    if level == "High":   return "FDECED"
    if level == "Medium": return "FEF9E7"
    if level == "Low":    return "EAF9F1"
    return LTGRAY

def border_thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Data loading ──────────────────────────────────────────────────────────────
def load_ss(file):
    """Load and normalise Smartsheets export."""
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)

    df.columns = df.columns.str.strip()

    # Rename via SS_COL_MAP
    rename = {c: SS_COL_MAP[c.lower()] for c in df.columns if c.lower() in SS_COL_MAP}
    df = df.rename(columns=rename)

    # Capture milestone columns
    milestone_present = [c for c in df.columns if c in MILESTONE_COLS]

    # Ensure project_id is string for joining
    if "project_id" in df.columns:
        df["project_id"] = df["project_id"].astype(str).str.strip()

    return df, milestone_present


def load_ns(file):
    """Load NetSuite time entries — used for stale project detection."""
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        # Read first sheet only — NS exports often have multiple tabs
        xl = pd.ExcelFile(file)
        df = pd.read_excel(xl, sheet_name=xl.sheet_names[0])

    df.columns = df.columns.str.strip()
    rename = {col: NS_COL_MAP[col.lower()] for col in df.columns if col.lower() in NS_COL_MAP}
    df = df.rename(columns=rename)

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if "hours" in df.columns:
        df["hours"] = pd.to_numeric(df["hours"], errors="coerce").fillna(0)
    if "project_id" in df.columns:
        df["project_id"] = df["project_id"].astype(str).str.strip()
    if "employee" in df.columns:
        df["employee"] = df["employee"].astype(str).str.strip()

    ns_min_date = df["date"].min() if "date" in df.columns else None
    return df, ns_min_date


def build_stale_projects(ss_df, ns_df):
    """
    Cross-reference SS active projects against NS time entries.
    Join key: project_id. Employee name sourced from SS 'Project Manager' column.
    Thresholds: <14d = ok, 14-29d = yellow, 30-59d = orange, 60+ = red.
    """
    if ns_df is None or ss_df is None:
        return pd.DataFrame()
    if "date" not in ns_df.columns:
        return pd.DataFrame()
    if "project_id" not in ss_df.columns:
        return pd.DataFrame()

    today = pd.Timestamp.today().normalize()

    # Last time entry per project_id in NS
    ns_join = "project_id" if "project_id" in ns_df.columns else "project"
    last_entry = (
        ns_df[ns_df["date"].notna()]
        .groupby(ns_join)["date"]
        .max()
        .reset_index()
        .rename(columns={ns_join: "project_id", "date": "last_entry"})
    )
    last_entry["project_id"] = last_entry["project_id"].astype(str).str.strip().str.split(".").str[0]

    # SS active FF projects — exclude complete/on-hold
    _complete_ph = {"10. complete/pending final billing", "12. ps review"}
    _onhold_st   = {"on-hold", "on hold", "onhold", "on_hold"}
    _tm_types    = {"t&m", "time & material", "time and material"}

    _phase = ss_df["phase"].astype(str).str.strip().str.lower() if "phase" in ss_df.columns else pd.Series("", index=ss_df.index)
    _stat  = ss_df["status"].astype(str).str.strip().str.lower() if "status" in ss_df.columns else pd.Series("", index=ss_df.index)
    _bill  = ss_df["billing_type"].astype(str).str.strip().str.lower() if "billing_type" in ss_df.columns else pd.Series("", index=ss_df.index)
    _type  = ss_df["project_type"].astype(str).str.strip().str.lower() if "project_type" in ss_df.columns else pd.Series("", index=ss_df.index)

    # Exclude complete/on-hold; exclude T&M whether in billing_type or project_type
    _tm_by_bill = _bill.isin(_tm_types)
    _tm_by_type = _type.str.contains("t&m|time.*material", na=False, regex=True)

    active_ss = ss_df[
        ~_phase.isin(_complete_ph) &
        ~_stat.isin(_onhold_st) &
        ~(_tm_by_bill | _tm_by_type)
    ].copy()
    active_ss["project_id"] = active_ss["project_id"].astype(str).str.strip().str.split(".").str[0]  # strip decimals e.g. "12345.0"

    if active_ss.empty:
        return pd.DataFrame()

    # Join on project_id only — employee name comes from SS Project Manager column
    merged = active_ss.merge(last_entry, on="project_id", how="left")
    merged["days_since"] = (today - merged["last_entry"]).dt.days

    def _flag(days):
        if pd.isna(days): return "⚫ No Entry"
        if days < 14:     return ""
        if days < 30:     return "🟡 14d+"
        if days < 60:     return "🟠 30d+"
        return "🔴 60d+"

    merged["Staleness"] = merged["days_since"].apply(_flag)
    stale = merged[merged["Staleness"] != ""].copy()

    if stale.empty:
        return pd.DataFrame()

    # Source consultant name from SS Project Manager column
    _pm_col   = "project_manager" if "project_manager" in stale.columns else None
    _name_col = "project_name" if "project_name" in stale.columns else "project_id"
    _phase_col = "phase" if "phase" in stale.columns else None

    display = pd.DataFrame()
    display["Consultant"] = stale[_pm_col].astype(str) if _pm_col else "—"
    display["Project"]    = stale[_name_col].astype(str)
    display["Phase"]      = stale[_phase_col].astype(str) if _phase_col else ""
    display["Last Entry"] = stale["last_entry"].dt.strftime("%Y-%m-%d").where(stale["last_entry"].notna(), "—")
    display["Days Since"] = stale["days_since"].where(stale["days_since"].notna(), -1).astype(int).replace(-1, "—")
    display["Staleness"]  = stale["Staleness"]

    _sort_order = {"🔴 60d+": 0, "🟠 30d+": 1, "🟡 14d+": 2, "⚫ No Entry": 3}
    display["_sort"] = display["Staleness"].map(_sort_order).fillna(9)
    display = display.sort_values(["_sort", "Consultant"]).drop(columns=["_sort"]).reset_index(drop=True)

    return display


# ── Scoring engine ────────────────────────────────────────────────────────────
def score_projects(ss_df, ns_df):
    """
    Join SS + NS, compute per-project weighted score (FF only, excl inactive).
    Returns scored DataFrame.
    """
    df = ss_df.copy()

    # T&M already filtered at load_ss — no further filtering needed here

    # Exclude inactive phases from score (still kept in data, score = 0)
    def is_active(phase):
        if not phase or str(phase).strip().lower() in ("", "nan", "none"):
            return True
        return str(phase).strip().lower() not in INACTIVE_PHASES

    # PM comes from SS directly — NS no longer used for PM lookup
    # Normalise: use "project_manager" col from SS if present
    _ss_pm_col = next((col for col in ["project_manager", "consultant"] if col in df.columns), None)
    if _ss_pm_col and _ss_pm_col != "project_manager":
        df["project_manager"] = df[_ss_pm_col]
    elif _ss_pm_col is None:
        df["project_manager"] = None
    df["pm_flag"] = df["project_manager"].isna()

    # Phase weight
    df["phase_weight"] = df["phase"].apply(lambda p: get_phase_weight(p)[0])

    # Multipliers
    resp_col = "client_responsiveness" if "client_responsiveness" in df.columns else None
    sent_col = "client_sentiment"      if "client_sentiment"      in df.columns else None
    risk_col = "risk_level"            if "risk_level"            in df.columns else None

    df["client_health_mult"] = df.apply(
        lambda r: client_health_multiplier(
            r[resp_col] if resp_col else None,
            r[sent_col] if sent_col else None
        ), axis=1
    )
    df["risk_mult"] = df[risk_col].apply(risk_multiplier) if risk_col else 1.0

    # Active = not on hold by phase OR status
    # Active statuses: In Progress, Awarded, Pending (anything except On-hold variants)
    def is_active_row(row):
        phase_inactive = not is_active(row.get("phase", ""))
        status_val     = str(row.get("status", "")).strip().lower()
        status_onhold  = status_val in ("on-hold", "on hold", "onhold", "on_hold")
        return not phase_inactive and not status_onhold

    df["active"] = df.apply(is_active_row, axis=1)

    # Total projects = all FF rows (excl Complete/Pending Final Billing)
    complete_phases = {"10. complete/pending final billing"}
    df["total_project"] = df["phase"].apply(
        lambda p: str(p).strip().lower() not in complete_phases if pd.notna(p) else True
    )

    # T&M projects kept for project counts but scored 0 (demand tracked via NS)
    _tm_mask = df["project_type"].str.lower().str.contains("t&m|time.*material", na=False, regex=True)         if "project_type" in df.columns else pd.Series(False, index=df.index)
    df["weighted_score"] = df.apply(
        lambda r: round(r["phase_weight"] * r["client_health_mult"] * r["risk_mult"], 2)
        if (r["active"] and not _tm_mask.loc[r.name]) else 0.0, axis=1
    )

    # PS Region from employee lookup (consultant = project_manager from NS)
    df["ps_region"] = df["project_manager"].apply(get_ps_region)
    df["role"] = df["project_manager"].apply(
        lambda n: EMPLOYEE_ROLES.get(str(n).strip(), {}).get("role", "Consultant") if pd.notna(n) else ""
    )

    return df


def build_consultant_summary(scored_df, ss_df=None):
    """Aggregate scored projects by consultant."""
    # Scoring groupby — uses project_manager from NS join, pm_flag rows score 0
    active = scored_df[scored_df["active"]].copy()

    grp = active.groupby("project_manager").agg(
        total_score=("weighted_score", "sum"),
        ps_region=("ps_region", "first"),
        role=("role", "first"),
    ).reset_index()

    grp["total_score"] = grp["total_score"].round(1)
    grp["workload_level"] = grp["total_score"].apply(lambda s: workload_level(s)[0])
    grp = grp.sort_values("total_score", ascending=False).reset_index(drop=True)

    # Project counts — computed from SS directly (consultant col) to match page 3
    # This avoids undercounting from pm_flag / T&M exclusions on the NS join side
    # Determine which column holds the consultant/PM name in ss_df
    _pm_col = next((col for col in ["project_manager", "consultant"] if col in ss_df.columns), None) if ss_df is not None else None
    if ss_df is not None and _pm_col is not None:
        _complete_ph = {"10. complete/pending final billing", "12. ps review"}
        _onhold_st   = {"on-hold", "on hold", "onhold", "on_hold"}
        _tm_types    = {"t&m", "time & material", "time and material"}
        _id_col = "project_id" if "project_id" in ss_df.columns else "project_name"
        _total_map  = {}
        _active_map = {}
        for _cons, _grp in ss_df.groupby(_pm_col):
            _billing = (_grp["billing_type"].astype(str).str.strip().str.lower()
                        if "billing_type" in _grp.columns else pd.Series("", index=_grp.index))
            _phase   = (_grp["phase"].astype(str).str.strip().str.lower()
                        if "phase" in _grp.columns else pd.Series("", index=_grp.index))
            _stat    = (_grp["status"].astype(str).str.strip().str.lower()
                        if "status" in _grp.columns else pd.Series("", index=_grp.index))
            _ff           = ~_billing.isin(_tm_types)
            _not_complete = ~_phase.isin(_complete_ph)
            _not_onhold   = ~_stat.isin(_onhold_st)
            _total_map[str(_cons).strip()]  = int(_grp[_ff & _not_complete][_id_col].nunique())
            _active_map[str(_cons).strip()] = int(_grp[_ff & _not_complete & _not_onhold][_id_col].nunique())
        grp["total_project_count"]  = grp["project_manager"].map(_total_map).fillna(0).astype(int)
        grp["active_project_count"] = grp["project_manager"].map(_active_map).fillna(0).astype(int)
    else:
        # Fallback to scored_df counts if ss_df not available
        total = scored_df[scored_df["total_project"]].copy() if "total_project" in scored_df.columns else scored_df.copy()
        active_counts = active.groupby("project_manager")["project_id"].nunique()
        total_counts  = total.groupby("project_manager")["project_id"].nunique()
        grp["active_project_count"] = grp["project_manager"].map(active_counts).fillna(0).astype(int)
        grp["total_project_count"]  = grp["project_manager"].map(total_counts).fillna(0).astype(int)

    # Flag missing PM
    missing = scored_df[scored_df["pm_flag"] == True]["project_id"].nunique()
    return grp, missing


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(scored_df, consultant_df, missing_pm_count, as_of, ns_min_date=None):
    def _safe(v):
        """Coerce value to a type openpyxl can write."""
        if v is None or (isinstance(v, float) and (v != v)):
            return ""
        try:
            import numpy as np
            if isinstance(v, np.integer):  return int(v)
            if isinstance(v, np.floating): return float(v) if v == v else ""
            if isinstance(v, np.bool_):    return bool(v)
        except ImportError:
            pass
        import pandas as pd
        if isinstance(v, pd.Timedelta): return str(v)
        if hasattr(pd, "NA") and v is pd.NA: return ""
        return v
    wb = Workbook()
    wb.remove(wb.active)

    # ── Tab 1: Dashboard ──────────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.sheet_properties.tabColor = NAVY
    ws_dash.sheet_view.showGridLines = False

    def _hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def dash_label(ws, row, col, text, size=10, bold=False, color="808080"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        return c

    def dash_value(ws, row, col, value, fmt=None, size=18, bold=True, color=NAVY):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        if fmt: c.number_format = fmt
        return c

    def dash_section(ws, row, col, text, ncols=6):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Manrope", size=11, bold=True, color="FFFFFF")
        c.fill = _hdr_fill(NAVY)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ncols-1)
        return c

    def rag_cell_d(ws, row, col, value, fmt=None, status="green"):
        colors = {"green": "EAF9F1", "yellow": "FEF9E7", "red": "FDECED"}
        txt    = {"green": "2ECC71", "yellow": "F39C12", "red": "E74C3C"}
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Manrope", size=14, bold=True, color=txt.get(status, "000000"))
        c.fill      = PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt: c.number_format = fmt
        return c

    # Column widths — buffer cols A and H
    for col, w in [(1,3),(2,22),(3,18),(4,18),(5,18),(6,18),(7,18),(8,3)]:
        ws_dash.column_dimensions[get_column_letter(col)].width = w
    for row in range(1, 55):
        ws_dash.row_dimensions[row].height = 18

    # Title
    tc = ws_dash.cell(row=2, column=2, value="Professional Services — FF Workload Score")
    tc.font = Font(name="Manrope", size=16, bold=True, color="FFFFFF")
    tc.fill = _hdr_fill(NAVY)
    ws_dash.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws_dash.row_dimensions[2].height = 30

    sc = ws_dash.cell(row=3, column=2, value=f"Data as of {as_of}  ·  Fixed Fee active projects only  ·  T&M excluded")
    sc.font = Font(name="Manrope", size=10, color="808080")
    ws_dash.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)

    kc = ws_dash.cell(row=4, column=2,
        value="Each active FF project is scored: Phase Weight × Client Health Multiplier × Risk Multiplier. "
              "Thresholds: Low 1–25 pts · Medium 26–60 pts · High 61+ pts (flag to Director). "
              "Active = not On Hold by phase or status. Total Projects includes On Hold, excludes Complete/Pending Final Billing.")
    kc.font      = Font(name="Manrope", size=9, italic=True, color="808080")
    kc.alignment = Alignment(wrap_text=True)
    ws_dash.merge_cells(start_row=4, start_column=2, end_row=4, end_column=7)
    ws_dash.row_dimensions[4].height = 30

    # ── Key Metrics ───────────────────────────────────────────────────────────
    total_consultants = len(consultant_df)
    high_count   = len(consultant_df[consultant_df["workload_level"] == "High"])
    medium_count = len(consultant_df[consultant_df["workload_level"] == "Medium"])
    low_count    = len(consultant_df[consultant_df["workload_level"] == "Low"])
    active_projects = scored_df[scored_df["active"]]["project_id"].nunique()
    total_projects  = scored_df[scored_df["total_project"]]["project_id"].nunique() if "total_project" in scored_df.columns else active_projects

    dash_section(ws_dash, 6, 2, "KEY METRICS", ncols=6)
    ws_dash.row_dimensions[5].height = 22

    for i, (label, value, fmt, status) in enumerate([
        ("Consultants Scored", total_consultants, "#,##0", None),
        ("High Workload",      high_count,   "#,##0", "red"    if high_count   > 0 else "green"),
        ("Medium Workload",    medium_count, "#,##0", "yellow" if medium_count > 0 else "green"),
        ("Low Workload",       low_count,    "#,##0", "green"),
        ("Total FF Projects",  total_projects,    "#,##0", None),
        ("Active Projects",    active_projects,   "#,##0", None),
    ]):
        col = 2 + i
        dash_label(ws_dash, 7, col, label)
        if status:
            rag_cell_d(ws_dash, 8, col, value, fmt=fmt, status=status)
        else:
            dash_value(ws_dash, 8, col, value, fmt=fmt, size=14)
    ws_dash.row_dimensions[8].height = 28

    # ── Workload by PS Region ─────────────────────────────────────────────────
    dash_section(ws_dash, 10, 2, "WORKLOAD BY PS REGION", ncols=6)
    ws_dash.row_dimensions[9].height = 22

    for ci, hdr in enumerate(["PS Region", "Total Projects", "Active Projects",
                               "Avg Weighted Score", "High Workload", "Consultants"], 2):
        c = ws_dash.cell(row=11, column=ci, value=hdr)
        c.font      = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
        c.fill      = _hdr_fill(TEAL)
        c.alignment = Alignment(horizontal="center")

    region_data = []
    for region in ["APAC", "EMEA", "NOAM"]:
        con_r    = consultant_df[consultant_df["ps_region"] == region]
        if len(con_r) == 0: continue
        scored_r = scored_df[scored_df["ps_region"] == region]
        r_total  = scored_r[scored_r["total_project"]]["project_id"].nunique() if "total_project" in scored_df.columns else 0
        r_active = scored_r[scored_r["active"]]["project_id"].nunique()
        r_avg    = round(con_r["total_score"].mean(), 1)
        r_high   = len(con_r[con_r["workload_level"] == "High"])
        r_cons   = len(con_r)
        region_data.append((region, r_total, r_active, r_avg, r_high, r_cons))

    for ri, (reg, r_total, r_active, r_avg, r_high, r_cons) in enumerate(region_data, 12):
        avg_status = "red" if r_avg > 60 else "yellow" if r_avg > 25 else "green"
        for ci, val in enumerate([reg, r_total, r_active, r_avg, r_high, r_cons], 2):
            c = ws_dash.cell(row=ri, column=ci, value=_safe(val))
            if ci == 5:  # Avg Weighted Score — RAG
                colors_d = {"red": ("E74C3C","FDECED"), "yellow": ("F39C12","FEF9E7"), "green": ("2ECC71","EAF9F1")}
                fc, bc = colors_d[avg_status]
                c.font = Font(name="Manrope", size=11, bold=True, color=fc)
                c.fill = PatternFill("solid", fgColor=bc)
            elif ci == 6 and r_high > 0:
                c.font = Font(name="Manrope", size=11, bold=True, color="E74C3C")
                c.fill = PatternFill("solid", fgColor="FDECED")
            else:
                c.font = Font(name="Manrope", size=10, color="1e2c63")
                c.fill = PatternFill("solid", fgColor=LTGRAY if ri % 2 == 0 else WHITE)
            c.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
            c.border    = border_thin()
        ws_dash.row_dimensions[ri].height = 18

    next_row = 12 + len(region_data) + 2

    # ── High Workload Consultants ─────────────────────────────────────────────
    high_cons = consultant_df[consultant_df["workload_level"] == "High"].sort_values("total_score", ascending=False)
    if len(high_cons) > 0:
        dash_section(ws_dash, next_row, 2, "HIGH WORKLOAD CONSULTANTS — Director Review Required", ncols=6)
        ws_dash.row_dimensions[next_row].height = 22
        next_row += 1
        for ci, hdr in enumerate(["Consultant", "PS Region", "Total Projects",
                                   "Active Projects", "Weighted Score", "Workload Level"], 2):
            c = ws_dash.cell(row=next_row, column=ci, value=hdr)
            c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=RED)
            c.alignment = Alignment(horizontal="center")
        next_row += 1
        for _, row in high_cons.iterrows():
            for ci, val in enumerate([
                row["project_manager"], row["ps_region"],
                row.get("total_project_count", 0), row.get("active_project_count", 0),
                row["total_score"], row["workload_level"]
            ], 2):
                c = ws_dash.cell(row=next_row, column=ci, value=_safe(val))
                c.font      = Font(name="Manrope", size=10, bold=(ci == 7), color="E74C3C" if ci == 7 else "1e2c63")
                c.fill      = PatternFill("solid", fgColor="FDECED")
                c.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
                c.border    = border_thin()
            ws_dash.row_dimensions[next_row].height = 16
            next_row += 1
        next_row += 1

    # No time booked flag
    if missing_pm_count > 0:
        if ns_min_date is not None and pd.notna(ns_min_date):
            no_time_since = (ns_min_date - pd.Timedelta(days=1)).strftime("%d %B %Y")
            flag_msg = f"⚠  {missing_pm_count} project(s) with no time booked since {no_time_since} — see 'No Time Booked This Period' tab."
        else:
            flag_msg = f"⚠  {missing_pm_count} project(s) with no time booked in this NS report period — see 'No Time Booked This Period' tab."
        ws_dash.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=7)
        fc = ws_dash.cell(next_row, 2, flag_msg)
        fc.font      = Font(name="Manrope", size=9, color="7D6608")
        fc.fill      = PatternFill("solid", fgColor="FEF9E7")
        fc.alignment = Alignment(wrap_text=True)

    ws_dash.freeze_panes = "B7"

    # ── Tab 2: By Consultant ──────────────────────────────────────────────────
    ws_con = wb.create_sheet("By Consultant")
    ws_con.sheet_properties.tabColor = TEAL

    con_headers = ["Consultant", "PS Region", "Total Projects", "Active Projects",
                   "Weighted Score", "Workload Level", "High Risk Projects", "Avg Score/Region"]
    con_widths   = [28, 14, 16, 16, 16, 16, 18, 18]
    write_title(ws_con, "WORKLOAD HEALTH SCORE — By Consultant", len(con_headers))
    style_header(ws_con, 2, con_headers, TEAL)
    ws_con.auto_filter.ref = f"A2:{get_column_letter(len(con_headers))}2"
    for i, w in enumerate(con_widths, 1):
        ws_con.column_dimensions[get_column_letter(i)].width = w

    # High risk count per consultant
    high_risk = scored_df[(scored_df["active"]) & (scored_df.get("risk_level", pd.Series(dtype=str)).str.lower().str.contains("high", na=False))]
    high_risk_count = high_risk.groupby("project_manager")["project_id"].nunique().to_dict() if "risk_level" in scored_df.columns else {}

    # Pre-calculate avg weighted score by PS region
    region_avg_score = consultant_df.groupby("ps_region")["total_score"].mean().round(1).to_dict()

    for r_idx, row in consultant_df.iterrows():
        r = 3 + r_idx
        level = row["workload_level"]
        bg = level_color(level)
        act  = row.get("active_project_count", 0)
        tot  = row.get("total_project_count", 0)
        avg  = region_avg_score.get(row["ps_region"], 0)
        hr   = high_risk_count.get(row["project_manager"], 0)
        vals = [row["project_manager"], row["ps_region"],
                tot, act, row["total_score"], level, hr, avg]
        for col, val in enumerate(vals, 1):
            c = ws_con.cell(r, col, _safe(val))
            # Col 9 = Avg Score/Region — highlight vs consultant's own score
            if col == 9:
                consultant_score = row["total_score"]
                diff = consultant_score - avg if avg > 0 else 0
                if diff > 5:       # meaningfully above regional avg — overloaded vs peers
                    c.font  = Font(name="Manrope", size=10, bold=True, color="E74C3C")
                    c.fill  = PatternFill("solid", fgColor="FDECED")
                elif diff < -5:    # meaningfully below regional avg — capacity available
                    c.font  = Font(name="Manrope", size=10, bold=True, color="27AE60")
                    c.fill  = PatternFill("solid", fgColor="EAF9F1")
                else:              # within ±5 of regional avg — on par
                    c.font  = Font(name="Manrope", size=10, color="555555")
                    c.fill  = PatternFill("solid", fgColor=LTGRAY)
                c.number_format = "0.0"
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_thin()
            else:
                style_cell(c, bg, align="center" if col > 1 else "left")
                c.border = border_thin()
        ws_con.row_dimensions[r].height = 16
    ws_con.freeze_panes = "A3"

    # ── Tab 3: At-Risk ────────────────────────────────────────────────────────
    ws_risk = wb.create_sheet("At-Risk Projects")
    ws_risk.sheet_properties.tabColor = RED

    at_risk_df = scored_df[
        (scored_df["active"]) & (
            (scored_df.get("risk_level",       pd.Series(dtype=str)).str.lower().str.contains("high",    na=False)) |
            (scored_df.get("schedule_health",  pd.Series(dtype=str)).str.lower().str.contains("behind",  na=False)) |
            (scored_df.get("rag",              pd.Series(dtype=str)).str.lower().str.contains("red",     na=False)) |
            (scored_df.get("client_sentiment", pd.Series(dtype=str)).str.lower().str.contains("negative",na=False))
        )
    ].copy()

    _flag_order = {"expired": 0, "near_expiry": 1, None: 2}
    at_risk_df["_sort_flag"] = at_risk_df["start_date"].apply(lambda s: _flag_order.get(term_age_flag(s), 2))
    at_risk_df = at_risk_df.sort_values(["_sort_flag", "weighted_score"], ascending=[True, False]).drop(columns=["_sort_flag"])

    risk_headers = ["Project", "Project ID", "Project Manager", "PS Region", "Phase",
                    "Weighted Score", "Risk Level", "Schedule Health", "RAG",
                    "Client Sentiment", "Client Responsiveness", "Start Date", "Go Live Date", "Term Flag"]
    risk_widths   = [38, 12, 24, 12, 28, 14, 12, 18, 10, 18, 22, 14, 14, 22]

    write_title(ws_risk, "AT-RISK PROJECTS — High Risk / Behind Schedule / Red RAG",
                len(risk_headers),
                f"{len(at_risk_df)} project(s) flagged")
    style_header(ws_risk, 3, risk_headers, RED)
    ws_risk.auto_filter.ref = f"A3:{get_column_letter(len(risk_headers))}3"
    for i, w in enumerate(risk_widths, 1):
        ws_risk.column_dimensions[get_column_letter(i)].width = w

    for r_idx, (_, row) in enumerate(at_risk_df.iterrows()):
        r = 4 + r_idx
        def gv(col): return row.get(col, "") if pd.notna(row.get(col, "")) else ""
        start_raw = gv("start_date")
        start_str = start_raw.strftime("%Y-%m-%d") if pd.notna(start_raw) and hasattr(start_raw, "strftime") else start_raw
        go_live_raw = gv("go_live_date")
        go_live_str = go_live_raw.strftime("%Y-%m-%d") if pd.notna(go_live_raw) and hasattr(go_live_raw, "strftime") else go_live_raw
        term_flag = term_age_flag(start_raw)
        term_label = ("⚠ Expired (>12 months)" if term_flag == "expired"
                      else "⚠ Near Expiry (10+ months)" if term_flag == "near_expiry"
                      else "")
        # Row background: expired = strong orange, near_expiry = amber, default = light red
        row_bg = ("FDEBD0" if term_flag == "expired"
                  else "FEF9E7" if term_flag == "near_expiry"
                  else "FEF0EE")
        vals = [
            gv("project_name"), gv("project_id"), gv("project_manager"), gv("ps_region"),
            gv("phase"), gv("weighted_score"), gv("risk_level"), gv("schedule_health"),
            gv("rag"), gv("client_sentiment"), gv("client_responsiveness"),
            start_str, go_live_str, term_label,
        ]
        for col, val in enumerate(vals, 1):
            c = ws_risk.cell(r, col, _safe(val))
            if col == 9:
                bg = rag_color(gv("rag"))
            elif col == 14 and term_flag:  # Term Flag column
                bg = "FDEBD0" if term_flag == "expired" else "FEF9E7"
                c.font = Font(name="Manrope", size=9, bold=True,
                              color="E67E22" if term_flag == "expired" else "F39C12")
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="left")
                c.border = border_thin()
                continue
            else:
                bg = row_bg
            style_cell(c, bg, align="center" if col > 2 else "left")
            c.border = border_thin()
        ws_risk.row_dimensions[r].height = 16
    ws_risk.freeze_panes = "A4"

    # ── Tab 5: Phase Distribution Matrix ─────────────────────────────────────
    ws_phase = wb.create_sheet("FF Phase Distribution")
    ws_phase.sheet_properties.tabColor = "1ABC9C"

    phase_order = [
        "00. Onboarding",
        "01. Requirements and Design",
        "02. Configuration",
        "03. Enablement/Training",
        "04. UAT",
        "05. Prep for Go-Live",
        "06. Go-Live",
        "07. Data Migration",
        "08. Ready for Support Transition",
        "09. Phase 2 Scoping",
        "10. Complete/Pending Final Billing",
        "11. On Hold",
        "12. PS Review",
    ]

    active_only = scored_df[scored_df["active"]].copy()
    if "phase" in active_only.columns and "project_manager" in active_only.columns:
        pivot = active_only.groupby(["project_manager", "phase"]).size().unstack(fill_value=0)
        existing_phases = [p for p in phase_order if p in pivot.columns]
        other_phases    = [p for p in pivot.columns if p not in phase_order]
        pivot = pivot[existing_phases + other_phases]
        # Total projects per consultant (all FF excl complete)
        if "total_project" in scored_df.columns:
            total_only = scored_df[scored_df["total_project"]]
            total_proj_counts = total_only.groupby("project_manager")["project_id"].nunique().rename("Total Projects")
        else:
            total_proj_counts = active_only.groupby("project_manager")["project_id"].nunique().rename("Total Projects")

        active_proj_counts = active_only.groupby("project_manager")["project_id"].nunique().rename("Active Projects")

        pivot["Total Projects"]  = pivot.index.map(total_proj_counts).fillna(0).astype(int)
        pivot["Active Projects"] = pivot.index.map(active_proj_counts).fillna(0).astype(int)
        pivot["Total Score"]     = active_only.groupby("project_manager")["weighted_score"].sum().round(1)
        pivot["Workload Level"]  = pivot["Total Score"].apply(lambda s: workload_level(s)[0])
        pivot = pivot.sort_values("Total Score", ascending=False)
        pivot = pivot.reset_index()

        phase_headers = ["Consultant"] + list(pivot.columns[1:])
        write_title(ws_phase, "PHASE DISTRIBUTION MATRIX — Active FF Projects by Consultant",
                    len(phase_headers))
        style_header(ws_phase, 2, phase_headers, "1ABC9C")
        ws_phase.auto_filter.ref = f"A2:{get_column_letter(len(phase_headers))}2"
        ws_phase.column_dimensions["A"].width = 28
        for i in range(2, len(phase_headers) + 1):
            ws_phase.column_dimensions[get_column_letter(i)].width = 14

        for r_idx, row_data in pivot.iterrows():
            r = 3 + r_idx
            level = row_data.get("Workload Level", "")
            bg    = level_color(level)
            for col, val in enumerate(row_data.tolist(), 1):
                c = ws_phase.cell(r, col, _safe(val))
                style_cell(c, bg, align="center" if col > 1 else "left")
                c.border = border_thin()
            ws_phase.row_dimensions[r].height = 16
        ws_phase.freeze_panes = "B3"

        # ── Weight key ────────────────────────────────────────────────────────
        key_start = 3 + len(pivot) + 3
        ws_phase.merge_cells(start_row=key_start, start_column=1,
                             end_row=key_start, end_column=len(phase_headers))
        kc = ws_phase.cell(key_start, 1, "PHASE WEIGHT KEY")
        kc.font = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
        kc.fill = PatternFill("solid", fgColor=NAVY)
        ws_phase.row_dimensions[key_start].height = 20

        key_headers = ["Phase", "Weight (pts)", "Rationale", "Workload Impact"]
        for ci, h in enumerate(key_headers, 1):
            c = ws_phase.cell(key_start + 1, ci, h)
            c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=TEAL)
            c.alignment = Alignment(horizontal="center")
        ws_phase.row_dimensions[key_start + 1].height = 16

        key_data = [
            ("00. Onboarding",                    1.0,  "Low daily effort; mostly scheduling",           "Low"),
            ("01. Requirements and Design",        1.0,  "Significant effort; analysis-intensive",        "Medium"),
            ("02. Configuration",                  2.0,  "Highest daily effort; deep client interaction", "High"),
            ("03. Enablement/Training",            2.5,  "Preparation-intensive; time-boxed",             "High"),
            ("04. UAT",                            3.0,  "Variable on client responsiveness",             "High"),
            ("05. Prep for Go-Live",               1.0,  "Coordination-heavy; moderate effort",           "Medium"),
            ("06. Go-Live",                        3.0,  "Short duration; high intensity",                "Medium"),
            ("07. Data Migration",                 1.0,  "Rare in FF; placeholder weight",                "Low"),
            ("08. Ready for Support Transition",   0.5,  "Handoff; low effort",                           "Low"),
            ("09. Phase 2 Scoping",                1.0,  "Light engagement; similar to Onboarding",       "Low"),
            ("10. Complete/Pending Final Billing", 0.0,  "No active delivery effort",                     "None"),
            ("11. On Hold",                        0.25, "Minimal effort; occupies mental overhead",      "Minimal"),
            ("12. PS Review",                      0.25, "Internal review; low active effort",            "Minimal"),
        ]
        level_colors = {"High": "FDECED", "Medium": "FEF9E7", "Low": "EAF9F1",
                        "Minimal": LTGRAY, "None": LTGRAY}

        for ki, (phase, weight, rationale, impact) in enumerate(key_data):
            r = key_start + 2 + ki
            bg = level_colors.get(impact, WHITE)
            for ci, val in enumerate([phase, weight, rationale, impact], 1):
                c = ws_phase.cell(r, ci, _safe(val))
                c.font = Font(name="Manrope", size=9,
                              color="1e2c63" if ci < 4 else
                              {"High":"E74C3C","Medium":"F39C12","Low":"27AE60"}.get(impact,"555555"))
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center" if ci == 2 else "left",
                                        wrap_text=(ci == 3))
                c.border = border_thin()
            ws_phase.row_dimensions[r].height = 16

        ws_phase.column_dimensions["A"].width = 32
        ws_phase.column_dimensions["B"].width = 14
        ws_phase.column_dimensions["C"].width = 42
        ws_phase.column_dimensions["D"].width = 14

    # ── Tab 6: No PM Assigned ─────────────────────────────────────────────────
    ws_nopm = wb.create_sheet("No Time Booked This Period")
    ws_nopm.sheet_properties.tabColor = "F39C12"

    no_pm_df = scored_df[scored_df["pm_flag"] == True].copy() if "pm_flag" in scored_df.columns else pd.DataFrame()

    # Date label for title and subtitle
    if ns_min_date is not None and pd.notna(ns_min_date):
        no_time_since = (ns_min_date - pd.Timedelta(days=1)).strftime("%d %B %Y")
        nopm_title    = f"NO TIME BOOKED SINCE {no_time_since.upper()} — Projects Not in NetSuite Time Report"
        nopm_subtitle = (f"These {len(no_pm_df)} project(s) had no time logged in NetSuite on or after "
                         f"{no_time_since}. Verify if active and confirm consultant assignment.")
    else:
        nopm_title    = "NO TIME BOOKED — Projects Not Found in NetSuite Time Report"
        nopm_subtitle = (f"These {len(no_pm_df)} project(s) were not found in the NetSuite time report. "
                         "Verify if active and confirm consultant assignment.")

    nopm_headers = ["Project", "Project ID", "Project Type", "Phase", "Territory",
                    "Status", "Overall RAG", "Start Date", "Go Live Date", "Term Flag"]
    nopm_widths  = [42, 12, 22, 30, 14, 14, 12, 14, 14, 22]

    write_title(ws_nopm, nopm_title, len(nopm_headers), nopm_subtitle)
    style_header(ws_nopm, 3, nopm_headers, "F39C12")
    ws_nopm.auto_filter.ref = f"A3:{get_column_letter(len(nopm_headers))}3"
    for i, w in enumerate(nopm_widths, 1):
        ws_nopm.column_dimensions[get_column_letter(i)].width = w

    if len(no_pm_df) > 0:
        # Sort: Expired first, Near Expiry second, then clean, then alpha within each
        _flag_order = {"⚠ Expired (>12 months)": 0, "⚠ Near Expiry (10+ months)": 1, "": 2}
        no_pm_df["_sort_flag"] = no_pm_df["start_date"].apply(
            lambda s: _flag_order.get(
                "⚠ Expired (>12 months)" if term_age_flag(s) == "expired"
                else "⚠ Near Expiry (10+ months)" if term_age_flag(s) == "near_expiry"
                else "", 2))
        no_pm_df = no_pm_df.sort_values(["_sort_flag", "project_name"]).drop(columns=["_sort_flag"])
        for r_idx, (_, row) in enumerate(no_pm_df.iterrows()):
            r = 4 + r_idx
            def gv(col): return row.get(col, "") if pd.notna(row.get(col, "")) else ""
            go_live = gv("go_live_date")
            go_live_str = go_live.strftime("%Y-%m-%d") if pd.notna(go_live) and hasattr(go_live, "strftime") else go_live
            start = gv("start_date")
            start_str = start.strftime("%Y-%m-%d") if pd.notna(start) and hasattr(start, "strftime") else start
            term_flag = term_age_flag(start)
            term_label = ("⚠ Expired (>12 months)" if term_flag == "expired"
                          else "⚠ Near Expiry (10+ months)" if term_flag == "near_expiry"
                          else "")
            row_bg = ("FDEBD0" if term_flag == "expired"
                      else "FEF9E7" if term_flag == "near_expiry"
                      else "F5F5F5" if r_idx % 2 == 0 else WHITE)
            vals = [gv("project_name"), gv("project_id"), gv("project_type"), gv("phase"),
                    gv("territory"), gv("status"), gv("rag"), start_str, go_live_str, term_label]
            for col, val in enumerate(vals, 1):
                c = ws_nopm.cell(r, col, _safe(val))
                if col == 10 and term_flag:  # Term Flag column
                    bg = "FDEBD0" if term_flag == "expired" else "FEF9E7"
                    c.font = Font(name="Manrope", size=9, bold=True,
                                  color="E67E22" if term_flag == "expired" else "F39C12")
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.alignment = Alignment(horizontal="left")
                    c.border = border_thin()
                    continue
                style_cell(c, row_bg, align="center" if col > 2 else "left")
                c.border = border_thin()
            ws_nopm.row_dimensions[r].height = 16
    else:
        ws_nopm.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(nopm_headers))
        nc = ws_nopm.cell(4, 1, "All projects have time logged in the NetSuite report.")
        style_cell(nc, "EAF9F1", font_color="27AE60")

    ws_nopm.freeze_panes = "A4"

    # ── Tab 7: Processed Data ─────────────────────────────────────────────────
    ws_raw = wb.create_sheet("Processed Data")
    raw_cols = [c for c in scored_df.columns]
    style_header(ws_raw, 1, raw_cols, NAVY)
    for r_idx, (_, row_data) in enumerate(scored_df.iterrows()):
        r = 2 + r_idx
        for col, val in enumerate(row_data.tolist(), 1):
            c = ws_raw.cell(r, col, str(val) if pd.notna(val) else "")
            c.font      = Font(size=9, name="Manrope")
            c.alignment = Alignment(horizontal="left")
        ws_raw.row_dimensions[r].height = 14

    # Tab order
    tab_order = ["Dashboard", "By Consultant", "FF Phase Distribution",
                 "At-Risk Projects", "No Time Booked This Period", "Processed Data"]
    wb._sheets = sorted(wb._sheets, key=lambda s: tab_order.index(s.title) if s.title in tab_order else 99)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit UI ──────────────────────────────────────────────────────────────
def main():
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;600;700&display=swap" rel="stylesheet">
        <style>
            html, body, [class*="css"] { font-family: 'Manrope', sans-serif !important; }
            h1, h2, h3, .stMarkdown, .stDataFrame, label, button { font-family: 'Manrope', sans-serif !important; }
        </style>
        <div style='background-color:#1e2c63;padding:24px 32px;border-radius:8px;margin-bottom:24px;font-family:Manrope,sans-serif'>
            <h1 style='color:white;margin:0;font-size:28px;font-family:Manrope,sans-serif'>FF Workload Score</h1>
            <p style='color:#aac4d0;margin:6px 0 0 0;font-size:14px;font-family:Manrope,sans-serif'>
                Weighted project scoring across active Fixed Fee engagements — by consultant and phase
            </p>
            <p style='color:#8ab0c0;margin:8px 0 0 0;font-size:12px;font-family:Manrope,sans-serif;line-height:1.6;'>
                Each active FF project is scored by <b>phase weight</b> × <b>client health multiplier</b> × <b>risk multiplier</b>.
                Thresholds: <b>Low</b> 1–25 pts &nbsp;·&nbsp; <b>Medium</b> 26–60 pts &nbsp;·&nbsp; <b>High</b> 61+ pts (flag to Director).
                T&amp;M projects and inactive phases (On Hold / Complete) are excluded from scoring.
            </p>
        </div>
    """, unsafe_allow_html=True)

    # Phase weight reference
    with st.expander("View phase weights & scoring model"):
        weight_df = pd.DataFrame([
            {"Phase": k.title(), "Weight (pts)": v}
            for k, v in PHASE_WEIGHTS.items()
        ])
        st.dataframe(weight_df, hide_index=True, use_container_width=True)

    # ── Exclusions & Limitations ──────────────────────────────────────────────
    with st.expander("Exclusions & Limitations", expanded=False):
        st.markdown("""
        <style>
        .excl-list { font-family: Manrope, sans-serif; font-size: 13px; color: #4a5568; line-height: 1.8; padding-left: 4px; }
        .excl-list li { margin-bottom: 4px; }
        </style>
        <ul class='excl-list'>
            <li><b>T&amp;M projects</b> are excluded from all scoring and project counts.</li>
            <li><b>Projects not found in the NetSuite time report</b> are treated as inactive for scoring purposes.</li>
            <li><b>Workload scores</b> do not account for hours per week or project complexity beyond phase and client signals.</li>
            <li><b>PS Region</b> is assigned from a maintained employee location lookup — update required when headcount changes.</li>
        </ul>
        """, unsafe_allow_html=True)

        st.markdown("<div style='margin-bottom:8px'></div>", unsafe_allow_html=True)

    # ── Uploads ───────────────────────────────────────────────────────────────
    st.subheader("Step 1 — Upload Smartsheets DRS Export")
    st.caption("Required columns: Project ID, Project Name, Project Phase, Project Type, Territory, Status")
    ss_file = st.file_uploader(
        "Drop your file here or click to browse",
        type=["xlsx", "xls", "csv"],
        key="ss_upload"
    )

    st.subheader("Step 2 — Upload NetSuite Time Detail Export")
    st.caption("Used to compare time data with the current project list to identify project manager assignments and projects without time entries for the period.")
    ns_file = st.file_uploader(
        "Drop your file here or click to browse",
        type=["xlsx", "xls", "csv"],
        key="ns_upload"
    )

    if not ss_file:
        st.info(" Upload your Smartsheets DRS export to continue.")
        return

    # ── Process ───────────────────────────────────────────────────────────────
    try:
        ss_df, milestone_cols = load_ss(ss_file)
        ns_df, ns_min_date = load_ns(ns_file) if ns_file else (None, None)
        scored_df = score_projects(ss_df, ns_df)
        consultant_df, missing_pm = build_consultant_summary(scored_df, ss_df=ss_df)
        stale_df = build_stale_projects(ss_df, ns_df) if ns_df is not None else pd.DataFrame()

        # ── TEMP DEBUG ───────────────────────────────────────────────────────
        with st.expander("🔍 Debug: SS columns & project count check"):
            st.write("**ss_df columns:**", list(ss_df.columns))
            _pm_debug = next((col for col in ["project_manager", "consultant"] if col in ss_df.columns), None)
            st.write("**PM col detected:**", _pm_debug)
            if _pm_debug:
                _sample = ss_df.groupby(_pm_debug).size().reset_index(name="Project Count")
                st.write("**Row counts per PM:**", _sample)
            st.write("**consultant_df:**", consultant_df[["project_manager","total_project_count","active_project_count"]].head(10))
        # ── END DEBUG ────────────────────────────────────────────────────────
        # Add stale project count per consultant
        if not stale_df.empty:
            _stale_counts = stale_df.groupby("Consultant").size().rename("stale_count")
            consultant_df["stale_count"] = consultant_df["project_manager"].map(_stale_counts).fillna(0).astype(int)
        else:
            consultant_df["stale_count"] = 0
        as_of = datetime.today().strftime("%B %d, %Y")
    except Exception as e:
        st.error(f"Error processing files: {e}")
        st.exception(e)
        return

        st.success(" Processing complete!")

    # ── Metrics ───────────────────────────────────────────────────────────────
    total    = len(consultant_df)
    high     = len(consultant_df[consultant_df["workload_level"] == "High"])
    medium   = len(consultant_df[consultant_df["workload_level"] == "Medium"])
    low      = len(consultant_df[consultant_df["workload_level"] == "Low"])
    active_projects = scored_df[scored_df["active"]]["project_id"].nunique()
    total_projects  = scored_df[scored_df["total_project"]]["project_id"].nunique() if "total_project" in scored_df.columns else active_projects

    st.markdown(f"<div style='font-size:13px;color:#a0a0a0;font-family:Manrope,sans-serif;margin-bottom:12px'>Data as of <strong style='color:#ffffff'>{as_of}</strong></div>", unsafe_allow_html=True)

    def metric_card(label, value, pill_txt=None, pill_fg=None):
        pill = ""
        if pill_txt and pill_fg:
            pill = f"<div style='display:inline-block;margin-top:6px;padding:2px 10px;border-radius:999px;background-color:{pill_fg}33;font-size:13px;font-family:Manrope,sans-serif;color:{pill_fg}'>&#8593; {pill_txt}</div>"
        return f"<div style='font-size:14px;color:#a0a0a0;font-family:Manrope,sans-serif;margin-bottom:4px'>{label}</div><div style='font-size:36px;font-weight:700;color:inherit;font-family:Manrope,sans-serif;line-height:1.1'>{value}</div>{pill}"

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1: st.markdown(metric_card("Consultants Scored",          f"{total:,}"),             unsafe_allow_html=True)
    with m2: st.markdown(metric_card("Consultant High Workload",    f"{high}",    "At or over capacity",  "#e74c3c"), unsafe_allow_html=True)
    with m3: st.markdown(metric_card("Consultant Medium Workload",  f"{medium}",  "Monitor for changes",  "#f39c12"), unsafe_allow_html=True)
    with m4: st.markdown(metric_card("Total FF Projects",           f"{total_projects:,}"),     unsafe_allow_html=True)
    with m5: st.markdown(metric_card("Active Projects",             f"{active_projects:,}", "Excl. On Hold", "#4472C4"), unsafe_allow_html=True)

    st.markdown("<div style='margin-top:20px'></div>", unsafe_allow_html=True)
    if missing_pm > 0:
        st.warning(f"{missing_pm} active FF project(s) have no Project Manager assigned in SS DRS.")

    st.markdown("---")

    # ── Metric definitions ────────────────────────────────────────────────────
    with st.expander("How projects are counted & what's excluded", expanded=False):
        st.markdown("""
**Total Projects** — All FF (Fixed Fee) projects assigned to a consultant in the SS DRS that are not in a complete phase
(`10. Complete/Pending Final Billing` or `12. PS Review`). T&M projects and unassigned projects are excluded.

**Active Projects** — Projects not in On Hold (`On-Hold`, `On Hold`) status in SS DRS.

**Stale Projects** — Active FF projects cross-referenced against NS time entries.
A project is flagged if no time has been booked within the NS report window:
- 🟡 **14d+** — No time booked in 14–29 days
- 🟠 **30d+** — No time booked in 30–59 days
- 🔴 **60d+** — No time booked in 60+ days
- ⚫ **No Entry** — No time booked in the NS report period

**Excluded from all counts:**
- T&M (Time & Material) projects — demand tracked separately via NS
- Complete / Pending Final Billing projects (phase 10)
- PS Review projects (phase 12)
- Projects with no consultant assigned
        """)

    # ── Preview tabs ──────────────────────────────────────────────────────────
    tab1, tab2, tab3 = st.tabs(["By Consultant", "At-Risk", "Stale Projects"])

    with tab1:
        display_con = consultant_df.rename(columns={
            "project_manager":      "Consultant",
            "ps_region":            "PS Region",
            "active_project_count": "Active Projects",
            "total_project_count":  "Total Projects",
            "total_score":          "Weighted Score",
            "workload_level":       "Workload Level",
            "stale_count":          "Stale Projects",
        })
        _show_cols = ["Consultant", "PS Region", "Total Projects", "Active Projects",
                      "Weighted Score", "Workload Level"]
        if not stale_df.empty:
            _show_cols.append("Stale Projects")
        st.dataframe(display_con[_show_cols], hide_index=True, use_container_width=True)

    with tab3:
        if stale_df.empty:
            if ns_df is None:
                st.info("Upload your NS time entries report to enable stale project detection.")
            else:
                st.success("No stale projects detected — all active projects have recent time entries.")
        else:
            st.markdown("#### Stale Projects — No Recent Time Booked")
            st.caption(
                "🟡 14d+ = no time in 14–29 days · "
                "🟠 30d+ = no time in 30–59 days · "
                "🔴 60d+ = no time in 60+ days · "
                "⚫ No Entry = no time booked in the NS report period"
            )
            # Summary counts — ordered lightest to most severe, then no entry
            _counts = stale_df["Staleness"].value_counts()
            _c1, _c2, _c3, _c4 = st.columns(4)
            with _c1: st.markdown(metric_card("14d+ No Time",  str(_counts.get("🟡 14d+",     0)), "14–29 days",        "#f39c12"), unsafe_allow_html=True)
            with _c2: st.markdown(metric_card("30d+ No Time",  str(_counts.get("🟠 30d+",     0)), "30–59 days",        "#e67e22"), unsafe_allow_html=True)
            with _c3: st.markdown(metric_card("60d+ No Time",  str(_counts.get("🔴 60d+",     0)), "60+ days",          "#e74c3c"), unsafe_allow_html=True)
            with _c4: st.markdown(metric_card("Not in NS",     str(_counts.get("⚫ No Entry", 0)), "No time booked", "#7f8c8d"), unsafe_allow_html=True)
            st.dataframe(stale_df, hide_index=True, use_container_width=True)

    with tab2:
        at_risk = scored_df[
            (scored_df["active"]) & (
                (scored_df.get("risk_level",      pd.Series(dtype=str)).str.lower().str.contains("high",   na=False)) |
                (scored_df.get("schedule_health", pd.Series(dtype=str)).str.lower().str.contains("behind", na=False)) |
                (scored_df.get("rag",             pd.Series(dtype=str)).str.lower().str.contains("red",    na=False))
            )
        ]
        if len(at_risk) == 0:
            st.success("No at-risk projects flagged.")
        else:
            risk_cols = ["project_name", "project_manager", "phase", "weighted_score",
                         "risk_level", "schedule_health", "rag", "go_live_date"]
            avail_r = [c for c in risk_cols if c in at_risk.columns]
            st.dataframe(at_risk[avail_r].sort_values("weighted_score", ascending=False),
                         hide_index=True, use_container_width=True)

    # ── Download ──────────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Step 3 — Generate Report")
    excel_buf = build_excel(scored_df, consultant_df, missing_pm, as_of, ns_min_date)
    fname = f"Workload_Health_Score_{datetime.today().strftime('%Y%m%d')}.xlsx"
    st.download_button(
label="⬇ Download Workload Health Score Report",
        data=excel_buf,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()