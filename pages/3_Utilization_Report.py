"""
PS Utilization Credit Report — Self-contained page
"""
import streamlit as st
import pandas as pd
import re as _re_constants
import io
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from shared.utils import (
    match_ff_task, thin_border, hdr_fill, row_fill, group_bg,
    style_header, style_cell, write_title, auto_detect_columns,
    assign_credits, _xl_val, build_excel,
)

st.session_state["current_page"] = "Utilization Report"

# ════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════

NAVY     = "1e2c63"
TEAL     = "4472C4"
WHITE    = "FFFFFF"
LTGRAY   = "F2F2F2"
MID_GRAY = "BDC3C7"

TAG_COLORS = {
    "CREDITED":     "EAF9F1",
    "OVERRUN":      "FDECED",
    "PARTIAL":      "FEF9E7",
    "NON-BILLABLE": "EBEDEE",
    "FF: NO SCOPE DEFINED": "F2F2F2",
}
TAG_BADGE = {
    "CREDITED":     "🟢",
    "OVERRUN":      "🔴",
    "PARTIAL":      "🟡",
    "NON-BILLABLE": "⚫",
    "FF: NO SCOPE DEFINED": "⚪",
}

PTO_KEYWORDS = ["vacation", "pto", "sick", "vacation/pto"]

EMPLOYEE_ROLES = {
    "Barrio, Nairobi":          {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Hughes, Madalyn":          {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Porangada, Suraj":         {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Cadelina":                 {"role": "Project Manager",    "products": [],                                                                                      "learning": []},
    "Bell, Stuart":             {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "DiMarco, Nicole R":        {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": []},
    "Murphy, Conor":            {"role": "Solution Architect", "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},
    "Church, Jason G":          {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Arestarkhov, Yaroslav":    {"role": "Consultant",         "products": ["Billing", "Capture"],                                                                   "learning": []},
    "Carpen, Anamaria":         {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Centinaje, Rhodechild":    {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},
    "Cooke, Ellen":             {"role": "Consultant",         "products": ["Billing", "Payroll"],                                                                   "learning": []},
    "Dolha, Madalina":          {"role": "Consultant",         "products": ["Capture", "Reconcile", "CC Statement Import", "Reconcile PSP", "e-Invoicing"],          "learning": []},
    "Finalle-Newton, Jesse":    {"role": "Solution Architect", "products": ["Reporting"],                                                                            "learning": []},
    "Gardner, Cheryll L":       {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Hopkins, Chris":           {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Ickler, Georganne":        {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Isberg, Eric":             {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Jordanova, Marija":        {"role": "Consultant",         "products": ["Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],     "learning": []},
    "Lappin, Thomas":           {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": ["Capture", "Reconcile"]},
    "Longalong, Santiago":      {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": ["Billing"]},
    "Mohammad, Manaan":         {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Morris, Lisa":             {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "NAQVI, SYED":              {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Olson, Austin D":          {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Pallone, Daniel":          {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Raykova, Silvia":          {"role": "Consultant",         "products": ["Capture", "Approvals", "e-Invoicing"],                                                  "learning": []},
    "Selvakumar, Sajithan":     {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile"],                                                    "learning": []},
    "Snee, Stefanie J":         {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Swanson, Patti":           {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": [], "util_exempt": True},
    "Tuazon, Carol":            {"role": "Consultant",         "products": ["Payroll", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"],       "learning": []},
    "Zoric, Ivan":              {"role": "Consultant",         "products": ["Capture", "Approvals", "Reconcile", "CC Statement Import", "Reconcile PSP", "SFTP Connector"], "learning": []},
    "Dunn, Steven":             {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Law, Brandon":             {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Quiambao, Generalyn":      {"role": "Developer",          "products": ["All"],                                                                                  "learning": []},
    "Cruz, Daniel":             {"role": "Consultant",         "products": ["All"],                                                                                  "learning": []},
    "Alam, Laisa":              {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Chan, Joven":              {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Cloete, Bronwyn":          {"role": "Consultant",         "products": ["Capture", "Approvals"],                                                                 "learning": []},
    "Eyong, Eyong":             {"role": "Consultant",         "products": ["Capture"],                                                                              "learning": []},
    "Hamilton, Julie C":        {"role": "Consultant",         "products": ["Reporting"],                                                                            "learning": []},
    "Hernandez, Camila":        {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
    "Rushbrook, Emma C":        {"role": "Consultant",         "products": ["Payroll"],                                                                              "learning": []},
    "Strauss, John W":          {"role": "Consultant",         "products": ["Billing"],                                                                              "learning": []},
}

UTIL_EXEMPT_EMPLOYEES = [
    k.lower() for k, v in EMPLOYEE_ROLES.items()
    if isinstance(v, dict) and v.get("util_exempt")
]

EMPLOYEE_LOCATION = {
    "Arestarkhov, Yaroslav":  ("Czech Republic",      None,       None),
    "Barrio, Nairobi":        ("USA",                 None,       None),
    "Bell, Stuart":           ("USA",                 None,       None),
    "Cadelina":               ("Manila (PH)",         "2026-03",  None),
    "Carpen, Anamaria":       ("Spain",               None,       None),
    "Centinaje, Rhodechild":  ("Manila (PH)",         None,       None),
    "Church, Jason G":        ("USA",                 None,       None),
    "Cooke, Ellen":           ("Northern Ireland",    None,       None),
    "Cruz, Daniel":           ("Manila (PH)",         None,       None),
    "DiMarco, Nicole R":      ("USA",                 None,       None),
    "Dolha, Madalina":        ("Faroe Islands",       None,       None),
    "Finalle-Newton, Jesse":  ("USA",                 None,       None),
    "Gardner, Cheryll L":     ("USA",                 None,       None),
    "Hopkins, Chris":         ("USA",                 None,       None),
    "Hughes, Madalyn":        ("USA",                 None,       None),
    "Ickler, Georganne":      ("USA",                 None,       None),
    "Isberg, Eric":           ("USA",                 None,       None),
    "Jordanova, Marija":      ("North Macedonia",     None,       None),
    "Lappin, Thomas":         ("Northern Ireland",    None,       None),
    "Longalong, Santiago":    ("Manila (PH)",         None,       None),
    "Mohammad, Manaan":       ("Canada",              None,       None),
    "Morris, Lisa":           ("Sydney (NSW)",        None,       None),
    "Murphy, Conor":          ("USA",                 None,       None),
    "NAQVI, SYED":            ("Canada",              None,       None),
    "Olson, Austin D":        ("USA",                 None,       None),
    "Pallone, Daniel":        ("Sydney (NSW)",        None,       None),
    "Porangada, Suraj":       ("USA",                 None,       None),
    "Raykova, Silvia":        ("Netherlands",         None,       None),
    "Selvakumar, Sajithan":   ("Canada",              None,       None),
    "Snee, Stefanie J":       ("USA",                 None,       None),
    "Stone, Matt":            ("USA",                 None,       None),
    "Swanson, Patti":         ("UK",                  None,       None),
    "Tuazon, Carol":          ("Manila (PH)",         None,       None),
    "Zoric, Ivan":            ("Serbia",              None,       None),
    "Dunn, Steven":           ("USA",                 None,       None),
    "Law, Brandon":           ("USA",                 None,       None),
    "Quiambao, Generalyn":    ("Manila (PH)",         None,       None),
    "Alam, Laisa":            ("USA",                 None,       "2025-12"),
    "Chan, Joven":            ("Manila (PH)",         None,       "2025-12"),
    "Cloete, Bronwyn":        ("Netherlands",         None,       "2026-02"),
    "Eyong, Eyong":           ("USA",                 None,       "2025-12"),
    "Hamilton, Julie C":      ("USA",                 None,       "2026-01"),
    "Hernandez, Camila":      ("USA",                 None,       "2025-12"),
    "Rushbrook, Emma C":      ("Wales",               None,       "2025-12"),
    "Strauss, John W":        ("USA",                 None,       "2026-03"),
}

def _emp_location(name):
    v = EMPLOYEE_LOCATION.get(name)
    if v is None: return None
    return v[0] if isinstance(v, tuple) else v

def _emp_active(name, period_str):
    v = EMPLOYEE_LOCATION.get(name)
    if v is None or not isinstance(v, tuple): return True
    _, start, end = v
    try:
        p = str(period_str).strip()
        if len(p) == 6 and "-" not in p: p = p[:4] + "-" + p[4:]
        if start and p < start[:7]: return False
        if end and p > end[:7]: return False
    except Exception: pass
    return True

PS_REGION_OVERRIDE = {
    "NAQVI, SYED":       "EMEA",
    "Cruz, Daniel":      "NOAM",
    "Chan, Joven":       "NOAM",
    "Rushbrook, Emma C": "EMEA",
}

PS_REGION_MAP = {
    "Sydney (NSW)":     "APAC",
    "Manila (PH)":      "APAC",
    "UK":               "EMEA",
    "Wales":            "EMEA",
    "Spain":            "EMEA",
    "Netherlands":      "EMEA",
    "Northern Ireland": "EMEA",
    "Faroe Islands":    "EMEA",
    "North Macedonia":  "EMEA",
    "Czech Republic":   "EMEA",
    "Serbia":           "EMEA",
    "USA":              "NOAM",
    "Canada":           "NOAM",
}

DEFAULT_SCOPE = {
    "Capture":                 20,
    "Approvals":               17,
    "Reconcile":               17,
    "PSP":                     18,
    "Payments":                30,
    "Reconcile 2.0":           20,
    "CC":                       6,
    "SFTP":                    12,
    "Premium - 10":            10,
    "Premium - 20":            20,
    "E-Invoicing":             15,
    "Capture and E-Invoicing": 30,
    "Additional Subsidiary":    2,
}

AVAIL_HOURS = {
    "Spain": {"2025-01":165.0,"2025-02":150.0,"2025-03":157.5,"2025-04":150.0,"2025-05":157.5,"2025-06":157.5,"2025-07":172.5,"2025-08":150.0,"2025-09":165.0,"2025-10":165.0,"2025-11":142.5,"2025-12":157.5,"2026-01":155.0,"2026-02":155.0,"2026-03":170.5,"2026-04":162.75,"2026-05":155.0,"2026-06":170.5,"2026-07":178.25,"2026-08":162.75,"2026-09":170.5,"2026-10":162.75,"2026-11":155.0,"2026-12":155.0},
    "UK": {"2025-01":165.0,"2025-02":150.0,"2025-03":157.5,"2025-04":150.0,"2025-05":150.0,"2025-06":157.5,"2025-07":172.5,"2025-08":150.0,"2025-09":165.0,"2025-10":172.5,"2025-11":150.0,"2025-12":157.5,"2026-01":157.5,"2026-02":150.0,"2026-03":165.0,"2026-04":150.0,"2026-05":142.5,"2026-06":165.0,"2026-07":172.5,"2026-08":150.0,"2026-09":165.0,"2026-10":165.0,"2026-11":157.5,"2026-12":157.5},
    "Northern Ireland": {"2025-01":165.0,"2025-02":150.0,"2025-03":150.0,"2025-04":150.0,"2025-05":150.0,"2025-06":157.5,"2025-07":165.0,"2025-08":150.0,"2025-09":165.0,"2025-10":172.5,"2025-11":150.0,"2025-12":157.5,"2026-01":157.5,"2026-02":150.0,"2026-03":157.5,"2026-04":150.0,"2026-05":142.5,"2026-06":165.0,"2026-07":165.0,"2026-08":150.0,"2026-09":165.0,"2026-10":165.0,"2026-11":157.5,"2026-12":157.5},
    "Netherlands": {"2025-01":176.0,"2025-02":160.0,"2025-03":168.0,"2025-04":152.0,"2025-05":160.0,"2025-06":160.0,"2025-07":184.0,"2025-08":168.0,"2025-09":176.0,"2025-10":184.0,"2025-11":160.0,"2025-12":168.0,"2026-01":168.0,"2026-02":160.0,"2026-03":176.0,"2026-04":152.0,"2026-05":144.0,"2026-06":176.0,"2026-07":184.0,"2026-08":168.0,"2026-09":176.0,"2026-10":176.0,"2026-11":168.0,"2026-12":176.0},
    "Faroe Islands": {"2025-01":176.0,"2025-02":160.0,"2025-03":168.0,"2025-04":152.0,"2025-05":160.0,"2025-06":152.0,"2025-07":176.0,"2025-08":168.0,"2025-09":176.0,"2025-10":184.0,"2025-11":160.0,"2025-12":168.0,"2026-01":168.0,"2026-02":160.0,"2026-03":176.0,"2026-04":144.0,"2026-05":144.0,"2026-06":176.0,"2026-07":168.0,"2026-08":168.0,"2026-09":176.0,"2026-10":176.0,"2026-11":168.0,"2026-12":168.0},
    "North Macedonia": {"2025-01":168.0,"2025-02":160.0,"2025-03":168.0,"2025-04":160.0,"2025-05":152.0,"2025-06":168.0,"2025-07":184.0,"2025-08":160.0,"2025-09":168.0,"2025-10":176.0,"2025-11":160.0,"2025-12":176.0,"2026-01":160.0,"2026-02":160.0,"2026-03":168.0,"2026-04":168.0,"2026-05":160.0,"2026-06":176.0,"2026-07":184.0,"2026-08":168.0,"2026-09":168.0,"2026-10":168.0,"2026-11":168.0,"2026-12":176.0},
    "Czech Republic": {"2025-01":176.0,"2025-02":160.0,"2025-03":168.0,"2025-04":160.0,"2025-05":160.0,"2025-06":168.0,"2025-07":168.0,"2025-08":168.0,"2025-09":168.0,"2025-10":176.0,"2025-11":152.0,"2025-12":160.0,"2026-01":168.0,"2026-02":160.0,"2026-03":176.0,"2026-04":160.0,"2026-05":152.0,"2026-06":176.0,"2026-07":176.0,"2026-08":168.0,"2026-09":168.0,"2026-10":168.0,"2026-11":160.0,"2026-12":168.0},
    "Serbia": {"2025-01":168.0,"2025-02":144.0,"2025-03":168.0,"2025-04":152.0,"2025-05":160.0,"2025-06":168.0,"2025-07":184.0,"2025-08":168.0,"2025-09":176.0,"2025-10":184.0,"2025-11":152.0,"2025-12":184.0,"2026-01":152.0,"2026-02":152.0,"2026-03":176.0,"2026-04":160.0,"2026-05":160.0,"2026-06":176.0,"2026-07":184.0,"2026-08":168.0,"2026-09":176.0,"2026-10":176.0,"2026-11":160.0,"2026-12":184.0},
    "Canada": {"2025-01":176.0,"2025-02":152.0,"2025-03":168.0,"2025-04":160.0,"2025-05":168.0,"2025-06":168.0,"2025-07":176.0,"2025-08":160.0,"2025-09":168.0,"2025-10":176.0,"2025-11":152.0,"2025-12":168.0,"2026-01":168.0,"2026-02":160.0,"2026-03":176.0,"2026-04":168.0,"2026-05":160.0,"2026-06":176.0,"2026-07":176.0,"2026-08":160.0,"2026-09":160.0,"2026-10":168.0,"2026-11":160.0,"2026-12":168.0},
    "USA": {"2025-01":168.0,"2025-02":152.0,"2025-03":168.0,"2025-04":176.0,"2025-05":168.0,"2025-06":160.0,"2025-07":176.0,"2025-08":168.0,"2025-09":168.0,"2025-10":184.0,"2025-11":144.0,"2025-12":176.0,"2026-01":160.0,"2026-02":152.0,"2026-03":176.0,"2026-04":176.0,"2026-05":160.0,"2026-06":168.0,"2026-07":176.0,"2026-08":168.0,"2026-09":168.0,"2026-10":168.0,"2026-11":152.0,"2026-12":176.0},
    "Sydney (NSW)": {"2025-01":159.6,"2025-02":152.0,"2025-03":159.6,"2025-04":136.8,"2025-05":167.2,"2025-06":152.0,"2025-07":174.8,"2025-08":152.0,"2025-09":167.2,"2025-10":167.2,"2025-11":152.0,"2025-12":159.6,"2026-01":152.0,"2026-02":152.0,"2026-03":167.2,"2026-04":144.4,"2026-05":159.6,"2026-06":159.6,"2026-07":174.8,"2026-08":152.0,"2026-09":167.2,"2026-10":159.6,"2026-11":159.6,"2026-12":159.6},
    "Manila (PH)": {"2025-01":176.0,"2025-02":152.0,"2025-03":168.0,"2025-04":152.0,"2025-05":168.0,"2025-06":160.0,"2025-07":184.0,"2025-08":160.0,"2025-09":176.0,"2025-10":184.0,"2025-11":136.0,"2025-12":160.0,"2026-01":168.0,"2026-02":152.0,"2026-03":176.0,"2026-04":152.0,"2026-05":160.0,"2026-06":168.0,"2026-07":184.0,"2026-08":152.0,"2026-09":176.0,"2026-10":176.0,"2026-11":152.0,"2026-12":144.0},
}

FF_TASKS = ["Configuration", "Enablement", "Training", "Post Go-live", "Project Management"]

def _emp_role(name):
    v = EMPLOYEE_ROLES.get(name)
    if v is None: return None
    return v["role"] if isinstance(v, dict) else v

def _emp_products(name, include_learning=False):
    v = EMPLOYEE_ROLES.get(name)
    if v is None or not isinstance(v, dict): return []
    products = list(v.get("products", []))
    if include_learning: products += v.get("products_learning", [])
    return products

def _emp_util_exempt(name):
    v = EMPLOYEE_ROLES.get(name)
    if isinstance(v, dict): return v.get("util_exempt", False)
    last = name.split(",")[0].strip() if "," in name else name.strip()
    v2 = EMPLOYEE_ROLES.get(last)
    if isinstance(v2, dict): return v2.get("util_exempt", False)
    return False

def get_avail_hours(region, period):
    region_clean = str(region).strip()
    for r, months in AVAIL_HOURS.items():
        if r.lower() == region_clean.lower():
            return months.get(str(period), None)
    return None

# ════════════════════════════════════════════════════════
# UTILS
# ════════════════════════════════════════════════════════

import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@st.cache_data(ttl=300, show_spinner=False)
def _run_utilization_engine(df_raw, period_start, period_end, df_drs=None):
    """
    Cross-user cached compute path. TTL 5 min — fresh enough for live data,
    cheap enough that two users hitting the same period in succession reuse work.
    DRS data is passed explicitly so it participates in the cache key.
    """
    import pandas as _pd
    _ts = _pd.Timestamp(period_start)
    _te = _pd.Timestamp(period_end) + _pd.Timedelta(days=1) - _pd.Timedelta(seconds=1)

    if "date" in df_raw.columns:
        _dated = _pd.to_datetime(df_raw["date"], errors="coerce")
        _mask = (_dated >= _ts) & (_dated <= _te)
        _df_period = df_raw[_mask].copy()
    else:
        _df_period = df_raw.copy()

    if _df_period.empty:
        return {"df": _df_period, "consumed": {}, "empty": True}

    if df_drs is not None and not df_drs.empty:
        if "project_id" in df_drs.columns and "project_name" in df_drs.columns:
            _drs_name_map = dict(zip(
                df_drs["project_id"].astype(str).str.strip(),
                df_drs["project_name"].astype(str).str.strip()
            ))
            if "project" in _df_period.columns:
                _df_period["_drs_project_name"] = _df_period["project"].astype(str).str.strip().map(
                    lambda pid: _drs_name_map.get(pid, ""))
                _drs_name_by_name = {v: v for v in df_drs["project_name"].astype(str).str.strip()}
                _empty_mask = _df_period["_drs_project_name"] == ""
                _df_period.loc[_empty_mask, "_drs_project_name"] = \
                    _df_period.loc[_empty_mask, "project"].map(
                        lambda n: next((v for v in _drs_name_by_name if n.lower() in v.lower() or v.lower() in n.lower()), ""))

    df, consumed, _skipped = assign_credits(_df_period, DEFAULT_SCOPE)
    return {"df": df, "consumed": consumed, "empty": False}


def _ns_signature(df_ns):
    if df_ns is None or df_ns.empty:
        return ("empty",)
    try:
        _max_d = pd.to_datetime(df_ns["date"], errors="coerce").max()
        return (len(df_ns), str(_max_d))
    except Exception:
        return (len(df_ns),)


def _proj_id_col(df):
    """Return the column name to use as the project ID for distinct counting."""
    for c in ("project_id", "project_internal_id", "project"):
        if c in df.columns:
            return c
    return None


def _billable_mask(df):
    """Boolean mask: rows that are billable (non-internal, non-skipped)."""
    if "billing_type" in df.columns and "credit_tag" in df.columns:
        return (df["billing_type"].fillna("").str.lower() != "internal") & (df["credit_tag"] != "SKIPPED")
    if "credit_tag" in df.columns:
        return df["credit_tag"].isin(["CREDITED", "PARTIAL", "OVERRUN", "UNCONFIGURED"])
    return pd.Series([True] * len(df), index=df.index)


def main():
    # ─────────────────────────────────────────────────────
    # Theme-aware styles using Streamlit's CSS variables
    # ─────────────────────────────────────────────────────
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700&display=swap" rel="stylesheet">
        <style>
            html, body, [class*="css"] { font-family: 'Manrope', sans-serif !important; }
            h1, h2, h3, .stMarkdown, .stDataFrame, label, button { font-family: 'Manrope', sans-serif !important; }

            /* Theme-aware card surfaces */
            /* Card surfaces — no explicit background; inherit page bg so cards
               flip with theme automatically (mirrors Portfolio Analytics pattern). */
            .util-card,
            .util-control-bar,
            .util-kpi,
            .util-legend,
            .util-callout,
            .util-table-header,
            .util-table-wrap {
                border: 1px solid rgba(128,128,128,0.25);
                color: inherit;
            }
            .util-card        { border-radius: 8px; padding: 14px; }
            .util-control-bar { border-radius: 10px; padding: 12px 16px; margin-bottom: 10px; }
            .util-kpi         { border-radius: 8px; padding: 14px; }
            .util-legend      { border-radius: 8px; padding: 10px 14px; margin-bottom: 8px;
                                display: flex; flex-wrap: wrap; gap: 8px;
                                align-items: center; font-size: 12px; }
            .util-callout     { border-radius: 8px; padding: 12px 14px; }
            .util-table-header{ border-radius: 8px 8px 0 0; padding: 10px 14px;
                                display: flex; justify-content: space-between; font-size: 12px; }
            .util-table-wrap  { border-top: none; border-radius: 0 0 8px 8px; overflow: hidden; }

            .util-meta-row {
                margin-top: 10px; padding-top: 10px;
                border-top: 1px solid rgba(128,128,128,0.18);
                display: flex; gap: 14px; font-size: 12px;
                opacity: 0.85;
                align-items: center; flex-wrap: wrap;
            }
            .util-live-dot {
                width: 7px; height: 7px; border-radius: 50%;
                background: #22c55e; display: inline-block; margin-right: 5px;
            }

            .util-kpi-label { font-size: 11px; opacity: 0.75; margin-bottom: 4px;
                              white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
            .util-kpi-value { font-size: 24px; font-weight: 600; line-height: 1.1;
                              font-variant-numeric: tabular-nums; }
            .util-kpi-sub   { font-size: 11px; opacity: 0.65; margin-top: 4px; }
            .util-kpi-pill  { display: inline-block; margin-top: 4px;
                              padding: 2px 9px; border-radius: 999px; font-size: 11px; }

            /* Pills (translucent so they work on either page bg) */
            .util-pill { padding: 3px 9px; border-radius: 999px;
                         font-size: 12px; white-space: nowrap; font-weight: 500; }
            .util-pill-green { background: rgba(34, 197, 94, 0.18); color: #15803d; }
            .util-pill-amber { background: rgba(245, 158, 11, 0.18); color: #b45309; }
            .util-pill-red   { background: rgba(239, 68, 68, 0.18); color: #b91c1c; }
            .util-pill-blue  { background: rgba(59, 130, 246, 0.18); color: #1d4ed8; }
            .util-pill-grey  { background: rgba(128, 128, 128, 0.18); color: inherit; opacity: 0.75; }
            @media (prefers-color-scheme: dark) {
                .util-pill-green { color: #7ed4a4; }
                .util-pill-amber { color: #f5b958; }
                .util-pill-red   { color: #f08585; }
                .util-pill-blue  { color: #6fa8dc; }
            }
            .stApp[data-theme="dark"] .util-pill-green { color: #7ed4a4; }
            .stApp[data-theme="dark"] .util-pill-amber { color: #f5b958; }
            .stApp[data-theme="dark"] .util-pill-red   { color: #f08585; }
            .stApp[data-theme="dark"] .util-pill-blue  { color: #6fa8dc; }

            .util-callout-red    { border-left: 3px solid #ef4444; }
            .util-callout-amber  { border-left: 3px solid #f59e0b; }
            .util-callout-blue   { border-left: 3px solid #3b82f6; }
            .util-callout-green  { border-left: 3px solid #22c55e; }
            .util-callout-title  { font-size: 12px; font-weight: 600; }
            .util-callout-num    { font-size: 22px; font-weight: 600;
                                   font-variant-numeric: tabular-nums; }
            .util-callout-body   { font-size: 12px; opacity: 0.8; line-height: 1.4; }

            .util-section-label {
                font-size: 13px; font-weight: 700; text-transform: uppercase;
                letter-spacing: 0.8px; color: #4472C4; margin-bottom: 8px;
            }

            .util-legend-label {
                opacity: 0.6; text-transform: uppercase;
                letter-spacing: 0.6px; font-size: 11px; margin-right: 4px;
            }

            /* Employee HTML table */
            .util-emp-table {
                width: 100%; border-collapse: collapse;
                font-family: 'Manrope', sans-serif; font-size: 13px;
                font-variant-numeric: tabular-nums;
                color: inherit;
            }

            .util-emp-table thead tr {
                background: rgba(128,128,128,0.08);
            }
            .util-emp-table th {
                padding: 10px 12px; font-weight: 600; text-align: left;
                opacity: 0.75;
                border-bottom: 1px solid rgba(128,128,128,0.25);
            }
            .util-emp-table th.num    { text-align: right; }
            .util-emp-table th.center { text-align: center; }
            .util-emp-table tbody tr  { border-bottom: 1px solid rgba(128,128,128,0.15); }
            .util-emp-table td        { padding: 9px 12px; vertical-align: middle; }
            .util-emp-table td.num    { text-align: right; }
            .util-emp-table td.center { text-align: center; }
            .util-emp-table td.muted  { opacity: 0.6; }
            .util-emp-avatar {
                display: inline-flex; align-items: center; justify-content: center;
                width: 24px; height: 24px; border-radius: 50%;
                font-size: 10px; font-weight: 600; margin-right: 8px;
                vertical-align: middle; flex-shrink: 0;
            }
            .util-emp-name { display: inline-flex; align-items: center; }
            .util-table-foot {
                padding: 8px 14px;
                border-top: 1px solid rgba(128,128,128,0.15);
                display: flex; justify-content: space-between;
                font-size: 11px; opacity: 0.6;
            }

            /* Sleek reference card (replaces rainbow tables) */
            .util-ref-grid {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 16px;
            }
            .util-ref-section h4 {
                font-size: 11px; font-weight: 600; opacity: 0.65;
                text-transform: uppercase; letter-spacing: 0.8px;
                margin: 0 0 10px 0;
            }
            .util-ref-row {
                display: grid;
                grid-template-columns: auto 1fr;
                gap: 12px; padding: 8px 0;
                border-bottom: 1px solid rgba(128,128,128,0.12);
                align-items: start;
                font-size: 12px;
            }
            .util-ref-row:last-child { border-bottom: none; }
            .util-ref-tag {
                padding: 2px 8px; border-radius: 999px;
                font-size: 11px; font-weight: 500;
                white-space: nowrap;
            }
            .util-ref-desc { opacity: 0.85; line-height: 1.5; }

            /* Movers row */
            .util-mover-row {
                display: flex; justify-content: space-between; align-items: center;
                padding: 7px 0; font-size: 12px;
                border-bottom: 1px solid rgba(128,128,128,0.1);
            }
            .util-mover-row:last-child { border-bottom: none; }
            .util-mover-vals {
                display: inline-flex; align-items: center; gap: 6px;
                font-variant-numeric: tabular-nums;
            }
            .util-mover-from { opacity: 0.6; }

            /* Share bar */
            .util-share-row {
                margin-bottom: 8px;
            }
            .util-share-head {
                display: flex; justify-content: space-between;
                font-size: 12px; margin-bottom: 3px;
            }
            .util-share-bar-bg {
                background: rgba(128,128,128,0.15);
                height: 8px; border-radius: 4px; overflow: hidden;
            }
            .util-share-bar-fg {
                height: 100%; background: #3b82f6; border-radius: 4px;
            }
        </style>
    """, unsafe_allow_html=True)

    # Hero (intentionally dark — brand element, like other pages)
    st.markdown(
        "<div style='background:#050D1F;padding:28px 36px 24px;border-radius:10px;"
        "margin-bottom:14px;font-family:Manrope,sans-serif;'>"
        "<div style='font-size:10px;font-weight:700;letter-spacing:2.5px;"
        "text-transform:uppercase;color:#3B9EFF;margin-bottom:8px'>Professional Services · Tools</div>"
        "<h1 style='color:white;margin:0;font-size:26px;'>Utilization Report</h1>"
        "<p style='color:#aac4d0;margin:6px 0 0 0;font-size:13px;'>"
        "Live utilization credits and capacity from NetSuite. Adjust the period — everything below recomputes automatically.</p>"
        "</div>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # Data sourcing (unchanged)
    # ─────────────────────────────────────────────────────
    _ns_from_session = st.session_state.get("df_ns")
    if _ns_from_session is None:
        st.info("Upload NS Time Detail on the Home page to generate the Utilization Report.")
        return

    df_raw = _ns_from_session.copy()

    _logged_in = st.session_state.get("consultant_name", "")
    from shared.constants import get_role as _gr, resolve_view_as, get_region_consultants
    from shared.config import EMPLOYEE_LOCATION as _EL_u, PS_REGION_MAP as _RM_u, PS_REGION_OVERRIDE as _RO_u
    from shared.constants import ACTIVE_EMPLOYEES as _AE_u
    _role_u = _gr(_logged_in)
    _is_mgr_u = _role_u in ("manager", "manager_only", "reporting_only")

    _home_browse_u = (st.session_state.get("_browse_passthrough") or
                      st.session_state.get("home_browse", ""))

    if _is_mgr_u and not _home_browse_u:
        from shared.constants import CONSULTANT_DROPDOWN as _CD_u
        _active_c_u = sorted([n for n in _CD_u if _gr(n) in ("consultant", "manager")])
        _by_rgn_u = {}
        for _cn_u in _active_c_u:
            _loc_u = _EL_u.get(_cn_u, "")
            _rg_u  = _RO_u.get(_cn_u, _RM_u.get(_loc_u, "Other"))
            _by_rgn_u.setdefault(_rg_u, []).append(_cn_u)
        _bopts_u = ["👥 All team"]
        for _rg_u in sorted(_by_rgn_u.keys()):
            _bopts_u.append(f"── {_rg_u} ──")
            _bopts_u.extend(_by_rgn_u[_rg_u])
        with st.sidebar:
            st.markdown("**View as:**")
            _home_browse_u = st.selectbox("View as", _bopts_u, key="util_view_as", label_visibility="collapsed")

    _browse_clean = _home_browse_u.replace("👥", "").strip() if _home_browse_u else ""
    _is_all_team  = _browse_clean.lower() in ("all team", "") or _home_browse_u in ("👥 All team", "All team")

    if _is_mgr_u and _is_all_team:
        _va_name_u, _va_region_u = None, None
    else:
        _va_name_u, _va_region_u, _ = resolve_view_as(
            _logged_in, _home_browse_u, EMPLOYEE_ROLES, _EL_u, _RM_u, _RO_u, _AE_u)

    if "employee" in df_raw.columns:
        from shared.constants import name_matches
        if _va_region_u:
            _rc_u = get_region_consultants(_va_region_u, _EL_u, _RM_u, _RO_u, _AE_u)
            _filtered_u = df_raw[df_raw["employee"].astype(str).str.strip().str.lower().isin(_rc_u)]
            if not _filtered_u.empty: df_raw = _filtered_u
        elif _va_name_u:
            _filtered_u = df_raw[df_raw["employee"].apply(lambda v: name_matches(v, _va_name_u))]
            if not _filtered_u.empty: df_raw = _filtered_u
        elif not _is_mgr_u and _logged_in:
            _filtered_u = df_raw[df_raw["employee"].apply(lambda v: name_matches(v, _logged_in))]
            if not _filtered_u.empty: df_raw = _filtered_u

    _raw_dates = pd.to_datetime(df_raw["date"], errors="coerce").dropna() if "date" in df_raw.columns else pd.Series(dtype="datetime64[ns]")
    _min_date  = _raw_dates.min().date() if not _raw_dates.empty else date(date.today().year, 1, 1)
    _max_date  = _raw_dates.max().date() if not _raw_dates.empty else date.today()

    # ─────────────────────────────────────────────────────
    # Sticky control bar
    # ─────────────────────────────────────────────────────
    today = date.today()
    _first_this_month_ts = pd.Timestamp(today.year, today.month, 1)
    _last_month_end = (_first_this_month_ts - pd.Timedelta(days=1)).date()
    _last_month_start = _last_month_end.replace(day=1)
    period_options = {
        "This month":    (date(today.year, today.month, 1), today),
        "Last month":    (_last_month_start, _last_month_end),
        "QTD":           (date(today.year, ((today.month - 1) // 3) * 3 + 1, 1), today),
        "YTD":           (date(today.year, 1, 1), today),
        "All available": (_min_date, _max_date),
        "Custom":        (None, None),
    }

    c1, c2, c3, c4, c5 = st.columns([1.2, 1.2, 1.2, 1.0, 1.0])
    with c1:
        preset = st.selectbox("Period", list(period_options.keys()),
                              index=0, key="util_period_preset")
    with c2:
        if preset == "Custom":
            _default_start = st.session_state.get("util_period_start", _min_date)
            period_start = st.date_input("Start", value=_default_start,
                                         min_value=date(2020, 1, 1), max_value=_max_date,
                                         key="util_period_start")
        else:
            _ps, _ = period_options[preset]
            period_start = max(_ps, _min_date)
            st.markdown(f"<div style='font-size:11px;opacity:0.6;text-transform:uppercase;letter-spacing:0.6px;margin-bottom:4px'>Start</div>"
                        f"<div style='border:1px solid rgba(128,128,128,0.25);border-radius:6px;padding:6px 10px;font-size:13px;opacity:0.85'>{period_start.strftime('%Y-%m-%d')}</div>",
                        unsafe_allow_html=True)
    with c3:
        if preset == "Custom":
            _default_end = st.session_state.get("util_period_end", _max_date)
            period_end = st.date_input("End", value=_default_end,
                                       min_value=period_start, max_value=date(2030, 12, 31),
                                       key="util_period_end")
        else:
            _, _pe = period_options[preset]
            period_end = min(_pe, _max_date)
            st.markdown(f"<div style='font-size:11px;opacity:0.6;text-transform:uppercase;letter-spacing:0.6px;margin-bottom:4px'>End</div>"
                        f"<div style='border:1px solid rgba(128,128,128,0.25);border-radius:6px;padding:6px 10px;font-size:13px;opacity:0.85'>{period_end.strftime('%Y-%m-%d')}</div>",
                        unsafe_allow_html=True)
    with c4:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        refresh_clicked = st.button("↻ Refresh", key="util_refresh",
                                    help="Re-pull from source. Only needed if NS data has changed but inputs haven't.")
    with c5:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        _download_slot = st.empty()

    # ─────────────────────────────────────────────────────
    # Auto-run cached engine (cache lives in @st.cache_data on the function;
    # cross-user sharing + 5min TTL).
    # ─────────────────────────────────────────────────────
    _df_drs = st.session_state.get("df_drs")
    if refresh_clicked:
        _run_utilization_engine.clear()
        # Also clear the Excel/Tableau buffer caches so they rebuild against fresh data
        st.session_state.pop("_util_excel_cache", None)
        st.session_state.pop("_util_tableau_cache", None)

    with st.spinner("Computing..."):
        try:
            result = _run_utilization_engine(df_raw, period_start, period_end, _df_drs)
        except Exception as e:
            st.error(f"Processing error: {e}")
            st.exception(e)
            return

    if result.get("empty"):
        st.markdown(
            f"<div class='util-meta-row'>"
            f"<span><span class='util-live-dot'></span><b>Live</b></span>"
            f"<span>NS data Report Period: {_min_date.strftime('%-d %b %Y')} → {_max_date.strftime('%-d %b %Y')}</span>"
            f"<span>·</span><span>No entries for {period_start.strftime('%-d %b %Y')} → {period_end.strftime('%-d %b %Y')}</span>"
            f"</div>", unsafe_allow_html=True)
        return

    df = result["df"]
    consumed = result["consumed"]

    # ─────────────────────────────────────────────────────
    # KPI / summary numbers
    # ─────────────────────────────────────────────────────
    bmask = _billable_mask(df)
    df_bill = df[bmask]

    hours_this_period   = df[df["credit_tag"] != "SKIPPED"]["hours"].sum() if "hours" in df.columns else 0
    total_credit        = df["credit_hrs"].sum()
    total_admin         = df[df["billing_type"].str.lower() == "internal"]["hours"].sum() if "billing_type" in df.columns else 0
    total_proj_overrun  = df[df["credit_tag"].isin(["OVERRUN", "PARTIAL"])]["variance_hrs"].sum() if "variance_hrs" in df.columns else 0

    credit_pct  = total_credit       / hours_this_period if hours_this_period else 0
    overrun_pct = total_proj_overrun / hours_this_period if hours_this_period else 0
    admin_pct   = total_admin        / hours_this_period if hours_this_period else 0
    credit_label = "On target" if credit_pct >= 0.70 else "Below target" if credit_pct >= 0.60 else "At risk"

    # Billable project count by project ID (the fix)
    _pid_col = _proj_id_col(df_bill)
    billable_proj_count = df_bill[_pid_col].astype(str).str.strip().nunique() if _pid_col else 0
    consultant_count = df["employee"].nunique() if "employee" in df.columns else 0

    # Capacity
    _avail_total = 0
    if "employee" in df.columns and "region" in df.columns and "period" in df.columns:
        _emp_region_ui = df.dropna(subset=["region"]).groupby("employee")["region"].first().to_dict()
        _emp_period_ui = df.groupby("employee")["period"].first().to_dict()
        for _e, _r in _emp_region_ui.items():
            _av = get_avail_hours(_r, _emp_period_ui.get(_e, ""))
            if _av: _avail_total += _av
    capacity_pct = hours_this_period / _avail_total if _avail_total else 0

    # Partial period detection
    import calendar as _cal
    _bdays_in = len(pd.bdate_range(period_start, period_end))
    _yr2, _mo2 = period_start.year, period_start.month
    _ms = pd.Timestamp(_yr2, _mo2, 1)
    _me = pd.Timestamp(_yr2, _mo2, _cal.monthrange(_yr2, _mo2)[1])
    _bdays_total_month = len(pd.bdate_range(_ms, _me))
    is_partial = (period_start.month == period_end.month and period_start.year == period_end.year and _bdays_in < _bdays_total_month)

    _last_refresh = st.session_state.get("_util_last_refresh", datetime.now())
    if refresh_clicked:
        st.session_state._util_last_refresh = datetime.now()
        _last_refresh = st.session_state._util_last_refresh
    _ago_min = max(0, int((datetime.now() - _last_refresh).total_seconds() // 60))
    _ago_str = "just now" if _ago_min == 0 else f"{_ago_min} min ago"

    _partial_str = f" · partial period ({_bdays_in} of {_bdays_total_month} days)" if is_partial else ""
    st.markdown(
        f"<div class='util-meta-row'>"
        f"<span><span class='util-live-dot'></span><b>Live</b></span>"
        f"<span>NS data Report Period: {_min_date.strftime('%-d %b %Y')} → {_max_date.strftime('%-d %b %Y')}{_partial_str}</span>"
        f"<span style='margin-left:auto;opacity:0.6'>Updated {_ago_str}</span>"
        f"</div>", unsafe_allow_html=True)

    # Excel download — on-demand. We don't build the buffer until user clicks "Prepare".
    # Excel download — managers + reporting_only only.
    # Two-stage flow: button to trigger build, then download_button replaces it.
    # Cache key built from stable inputs (NS sig + period + view) so it survives reruns.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"utilization_report_{timestamp}.xlsx"
    if _is_mgr_u:
        _excel_cache_key = ("xl", _ns_signature(_ns_from_session),
                            str(period_start), str(period_end),
                            _va_name_u, _va_region_u)
        st.session_state.setdefault("_util_excel_cache", {})

        # If user clicked "Prepare" on a previous render, this flag is set
        if st.session_state.get("_util_excel_prep_requested") == _excel_cache_key:
            st.session_state["_util_excel_prep_requested"] = None
            try:
                with st.spinner("Building Excel..."):
                    st.session_state._util_excel_cache[_excel_cache_key] = build_excel(df, DEFAULT_SCOPE, consumed)
            except Exception as _xl_err:
                st.error(f"Excel build failed: {_xl_err}")
                st.exception(_xl_err)

        _excel_buf = st.session_state._util_excel_cache.get(_excel_cache_key)

        with _download_slot:
            if _excel_buf is not None:
                st.download_button(
                    "⬇ Excel", data=_excel_buf, file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary", key="util_dl_excel_top",
                )
            else:
                if st.button("⬇ Excel", key="util_prep_excel", type="primary",
                             help="Click to generate the Excel report"):
                    st.session_state["_util_excel_prep_requested"] = _excel_cache_key
                    st.rerun()

    # ─────────────────────────────────────────────────────
    # Unmapped / alumni warnings
    # ─────────────────────────────────────────────────────
    _unmapped = []; _alumni = []
    _period_str = str(df["period"].mode().iloc[0]) if "period" in df.columns and len(df) > 0 else None
    for _emp in df["employee"].dropna().unique():
        _emp_s = str(_emp).strip()
        _loc = df[df["employee"]==_emp]["region"].iloc[0] if len(df[df["employee"]==_emp]) > 0 else ""
        _matched_key = None
        for k in EMPLOYEE_LOCATION:
            if _emp_s.lower() == k.lower() or _emp_s.lower().startswith(k.lower()) or k.lower().startswith(_emp_s.lower()):
                _matched_key = k; break
        if not _matched_key and not str(_loc).strip(): _unmapped.append(_emp_s)
        elif _matched_key and _period_str:
            if not _emp_active(_matched_key, _period_str): _alumni.append(_emp_s)
    if _unmapped:
        st.warning(f"**{len(_unmapped)} employee(s) have no location defined** — avail hours and PS region will show as Unknown.\n\n" + ", ".join(sorted(_unmapped)))
    if _alumni:
        st.info(f"**{len(_alumni)} employee(s) have time entries but are outside their active tenure** — excluded from utilization targets.\n\n" + ", ".join(sorted(_alumni)))

    # ─────────────────────────────────────────────────────
    # Helpers used in tabs
    # ─────────────────────────────────────────────────────
    def fmt_hrs(n):
        s = f"{n:,.2f}"
        if "." in s: s = s.rstrip("0").rstrip(".")
        return s

    def rag_pill_html(v, exempt=False):
        if exempt: return "<span class='util-pill util-pill-grey' style='opacity:0.7'>exempt</span>"
        if v is None: return "<span style='opacity:0.5'>—</span>"
        if v >= 0.70:   cls = "util-pill-green"
        elif v >= 0.60: cls = "util-pill-amber"
        else:           cls = "util-pill-red"
        return f"<span class='util-pill {cls}'>{v*100:.1f}%</span>"

    def avatar_html(name):
        parts = [p.strip() for p in str(name).replace(",", " ").split() if p.strip()]
        initials = (parts[0][0] + parts[1][0]).upper() if len(parts) >= 2 else (parts[0][:2].upper() if parts else "??")
        palettes = [
            ("rgba(34,197,94,0.18)", "#15803d"),
            ("rgba(59,130,246,0.18)", "#1d4ed8"),
            ("rgba(245,158,11,0.18)", "#b45309"),
            ("rgba(168,85,247,0.18)", "#7c3aed"),
            ("rgba(236,72,153,0.18)", "#be185d"),
            ("rgba(20,184,166,0.18)", "#0f766e"),
        ]
        bg, fg = palettes[hash(str(name)) % len(palettes)]
        return f"<span class='util-emp-avatar' style='background:{bg};color:{fg}'>{initials}</span>"

    def short_name(n):
        parts = str(n).split(",", 1)
        if len(parts) == 2:
            last = parts[0].strip()
            first = parts[1].strip().split()[0] if parts[1].strip() else ""
            return f"{last}, {first[:1]}." if first else last
        return n

    # ─────────────────────────────────────────────────────
    # Action callouts (for At a glance tab)
    # ─────────────────────────────────────────────────────
    # Build employee summary first — needed by callouts and Consultants tab
    _ep = df[df["credit_tag"] != "SKIPPED"]
    emp_sum = _ep.groupby(["employee", "period"], as_index=False).agg(
        hours_this_period=("hours", "sum"),
        credit_hrs=("credit_hrs", "sum"),
        ff_overrun_hrs=("variance_hrs", "sum"),
    ).sort_values(["employee", "period"])
    _emp_region_ui = df.dropna(subset=["region"]).groupby("employee")["region"].first().to_dict() if "region" in df.columns else {}
    emp_sum["location"]  = emp_sum["employee"].map(_emp_region_ui)
    emp_sum["avail_hrs"] = emp_sum.apply(
        lambda r: get_avail_hours(r["location"], r["period"]) if r["location"] else None, axis=1)
    emp_sum["util_vs_logged"]   = emp_sum.apply(lambda r: r["credit_hrs"] / r["hours_this_period"] if r["hours_this_period"] > 0 else None, axis=1)
    emp_sum["util_vs_capacity"] = emp_sum.apply(lambda r: r["credit_hrs"] / r["avail_hrs"] if r["avail_hrs"] and r["avail_hrs"] > 0 else None, axis=1)
    emp_sum["exempt"] = emp_sum["employee"].apply(_emp_util_exempt)

    # Below-60% consultants
    _below_60 = emp_sum[(~emp_sum["exempt"]) & emp_sum["util_vs_capacity"].notna() & (emp_sum["util_vs_capacity"] < 0.60)]
    below60_count = len(_below_60)
    below60_names = ", ".join(sorted([short_name(n) for n in _below_60["employee"].unique()]))

    # Projects in overrun
    # credit_tag = OVERRUN is only assigned to FF projects by assign_credits, so this is the
    # canonical signal for "FF in overrun" — no project_type string parsing needed.
    _proj_overrun = pd.DataFrame()
    if "variance_hrs" in df.columns and _pid_col and "credit_tag" in df.columns:
        _overrun_pids = df_bill[df_bill["credit_tag"] == "OVERRUN"][_pid_col].astype(str).str.strip().unique()
        _po_grp = df_bill[df_bill[_pid_col].astype(str).str.strip().isin(_overrun_pids)].groupby([_pid_col], as_index=False).agg(
            project=("project", "first") if "project" in df.columns else (_pid_col, "first"),
            project_type=("project_type", "first") if "project_type" in df.columns else (_pid_col, "first"),
            hours_logged=("hours", "sum"),
            credit_hrs=("credit_hrs", "sum"),
            overrun_hrs=("variance_hrs", "sum"),
        )
        _proj_overrun = _po_grp[_po_grp["overrun_hrs"] > 0].copy()
    overrun_count = len(_proj_overrun)
    overrun_total = _proj_overrun["overrun_hrs"].sum() if not _proj_overrun.empty else 0
    overrun_top = ""
    if not _proj_overrun.empty:
        _top = _proj_overrun.sort_values("overrun_hrs", ascending=False).iloc[0]
        overrun_top = f"Largest: <b>{_top.get('project', '')}</b> (+{_top['overrun_hrs']:.1f}h)"

    # No-scope FF
    _noscope = df[df["credit_tag"] == "UNCONFIGURED"] if "credit_tag" in df.columns else df.iloc[0:0]
    noscope_hrs = _noscope["hours"].sum() if not _noscope.empty else 0
    noscope_proj_count = _noscope[_pid_col].astype(str).str.strip().nunique() if _pid_col and not _noscope.empty else 0

    # ─────────────────────────────────────────────────────
    # Wider-window data prep (used by Trend + Task tabs)
    # ─────────────────────────────────────────────────────
    _trend_df_raw = _ns_from_session.copy()
    if "employee" in _trend_df_raw.columns:
        from shared.constants import name_matches as _nm
        if _va_region_u:
            _rc = get_region_consultants(_va_region_u, _EL_u, _RM_u, _RO_u, _AE_u)
            _filt = _trend_df_raw[_trend_df_raw["employee"].astype(str).str.strip().str.lower().isin(_rc)]
            if not _filt.empty: _trend_df_raw = _filt
        elif _va_name_u:
            _filt = _trend_df_raw[_trend_df_raw["employee"].apply(lambda v: _nm(v, _va_name_u))]
            if not _filt.empty: _trend_df_raw = _filt
        elif not _is_mgr_u and _logged_in:
            _filt = _trend_df_raw[_trend_df_raw["employee"].apply(lambda v: _nm(v, _logged_in))]
            if not _filt.empty: _trend_df_raw = _filt

    # The wider-window engine call is deferred — only runs when Trend or Task tab
    # is opened, since those are the only consumers and it's expensive on long periods.
    # Once visited, the result is cached by @st.cache_data on _run_utilization_engine
    # so subsequent renders are free.
    _trend_visited = st.session_state.get("_util_trend_visited", False)
    _task_visited  = st.session_state.get("_util_task_visited", False)
    if _trend_visited or _task_visited:
        try:
            _trend_result = _run_utilization_engine(_trend_df_raw, period_start, period_end, _df_drs)
        except Exception:
            _trend_result = {"empty": True}
    else:
        _trend_result = None  # Will trigger lazy-load placeholder in those tabs

    # ─────────────────────────────────────────────────────
    # Tabs
    # ─────────────────────────────────────────────────────
    overrun_badge = f" · <span style='color:#b91c1c'>{overrun_count}</span>" if overrun_count > 0 else ""
    tab_at_glance, tab_consult, tab_risk, tab_trend, tab_task, tab_detail = st.tabs([
        "At a glance",
        f"Consultants · {consultant_count}",
        f"Projects at risk{' · ' + str(overrun_count + noscope_proj_count) if (overrun_count + noscope_proj_count) > 0 else ''}",
        "Trend",
        "Task analysis",
        f"Detail · {len(df):,}",
    ])

    # ═══════════════════════════════════════════════════════════════════
    # TAB 1 — At a glance
    # ═══════════════════════════════════════════════════════════════════
    with tab_at_glance:
        # KPI strip
        m1, m2, m3, m4, m5 = st.columns(5)
        def _kpi_card(label, value, sub=None, sub_pill_class=None):
            sub_html = ""
            if sub:
                if sub_pill_class:
                    sub_html = f"<div style='display:inline-block;margin-top:4px;padding:2px 9px;border-radius:999px;font-size:11px' class='util-pill {sub_pill_class}'>{sub}</div>"
                else:
                    sub_html = f"<div style='font-size:11px;opacity:0.65;margin-top:4px'>{sub}</div>"
            return (f"<div class='util-kpi'>"
                    f"<div style='font-size:11px;opacity:0.75;margin-bottom:4px;"
                    f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>{label}</div>"
                    f"<div style='font-size:24px;font-weight:600;line-height:1.1;"
                    f"font-variant-numeric:tabular-nums'>{value}</div>"
                    f"{sub_html}</div>")

        with m1: st.markdown(_kpi_card("Billable projects", f"{billable_proj_count:,}", f"across {consultant_count} consultants"), unsafe_allow_html=True)
        with m2: st.markdown(_kpi_card("Hours logged", fmt_hrs(hours_this_period), f"of {_avail_total:,.0f} capacity ({capacity_pct:.1%})" if _avail_total else None), unsafe_allow_html=True)

        _credit_cls = "util-pill-green" if credit_pct >= 0.70 else "util-pill-amber" if credit_pct >= 0.60 else "util-pill-red"
        with m3: st.markdown(_kpi_card("Util credits", fmt_hrs(total_credit), f"{credit_pct:.1%} · {credit_label}", _credit_cls), unsafe_allow_html=True)
        with m4: st.markdown(_kpi_card("FF overrun", fmt_hrs(total_proj_overrun), f"{overrun_pct:.1%} of hrs", "util-pill-amber" if total_proj_overrun > 0 else None), unsafe_allow_html=True)
        with m5: st.markdown(_kpi_card("Admin hrs", fmt_hrs(total_admin), f"{admin_pct:.1%} of hrs"), unsafe_allow_html=True)

        # Where to look — three callouts, always shown, green when clean
        st.markdown("<div style='margin-top:18px;font-size:11px;opacity:0.6;text-transform:uppercase;letter-spacing:0.6px;margin-bottom:8px'>Where to look</div>", unsafe_allow_html=True)

        callout_cols = st.columns(3)

        def _callout(border_color, title_color, icon, title, num, body):
            return (f"<div class='util-card' style='border-left:3px solid {border_color}'>"
                    f"<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:6px'>"
                    f"<span style='font-size:12px;font-weight:600;color:{title_color}'>{icon} {title}</span>"
                    f"<span style='font-size:22px;font-weight:600;color:{title_color};font-variant-numeric:tabular-nums'>{num}</span></div>"
                    f"<div style='font-size:12px;opacity:0.75;line-height:1.4'>{body}</div>"
                    f"</div>")

        # Callout 1 — Below 60%
        if below60_count > 0:
            _names_short = ", ".join(sorted([short_name(n) for n in _below_60["employee"].unique()])[:5])
            _and_more = f" +{below60_count - 5} more" if below60_count > 5 else ""
            _c1 = _callout("#ef4444", "#b91c1c", "⚠", "Consultants below 60%", below60_count, f"{_names_short}{_and_more}")
        else:
            _c1 = _callout("#22c55e", "#15803d", "✓", "Consultants below 60%", 0, "All consultants at or above 60% utilization on capacity.")
        with callout_cols[0]: st.markdown(_c1, unsafe_allow_html=True)

        # Callout 2 — FF in overrun
        if overrun_count > 0:
            _c2 = _callout("#f59e0b", "#b45309", "↑", "FF projects in overrun", overrun_count, f"{overrun_total:.1f} hrs beyond contracted scope. {overrun_top}")
        else:
            _c2 = _callout("#22c55e", "#15803d", "✓", "FF projects in overrun", 0, "No Fixed Fee projects exceeded scope this period.")
        with callout_cols[1]: st.markdown(_c2, unsafe_allow_html=True)

        # Callout 3 — No scope
        if noscope_proj_count > 0:
            _c3 = _callout("#3b82f6", "#1d4ed8", "⊘", "FF: no scope defined", noscope_proj_count, f"{noscope_hrs:.1f} hrs on FF projects with no scope record. Finance follow-up needed.")
        else:
            _c3 = _callout("#22c55e", "#15803d", "✓", "FF: no scope defined", 0, "All FF time logged against scoped projects.")
        with callout_cols[2]: st.markdown(_c3, unsafe_allow_html=True)

        # Inline legend + popover
        st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)
        legend_col, ref_col = st.columns([8, 1])
        with legend_col:
            st.markdown(
                "<div class='util-legend'>"
                "<span class='util-legend-label'>Credit tags</span>"
                "<span class='util-pill util-pill-green'>● Credited</span>"
                "<span class='util-pill util-pill-amber'>● Partial</span>"
                "<span class='util-pill util-pill-red'>● Overrun</span>"
                "<span class='util-pill util-pill-grey'>● Non-billable</span>"
                "<span class='util-pill util-pill-blue'>● FF: no scope</span>"
                "<span class='util-legend-label' style='margin-left:12px'>RAG</span>"
                "<span class='util-pill util-pill-green'>● ≥70%</span>"
                "<span class='util-pill util-pill-amber'>● 60–69%</span>"
                "<span class='util-pill util-pill-red'>● &lt;60%</span>"
                "</div>", unsafe_allow_html=True)
        with ref_col:
            _popover_fn = getattr(st, "popover", None)
            _ref_ctx = _popover_fn("ⓘ How?", help="How are these calculated?") if _popover_fn else st.expander("ⓘ How?", expanded=False)
            with _ref_ctx:
                st.markdown("""
                <div class='util-ref-grid'>
                  <div class='util-ref-section'>
                    <h4>Credit tags</h4>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-green'>Credited</span>
                      <span class='util-ref-desc'>T&amp;M: full hours credited, no cap. Fixed Fee: hours credited up to scoped amount.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-amber'>Partial</span>
                      <span class='util-ref-desc'>Fixed Fee: hours credited up to remaining scope only.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-red'>Overrun</span>
                      <span class='util-ref-desc'>Fixed Fee: hours beyond contracted scope. Not credited.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-grey'>Non-billable</span>
                      <span class='util-ref-desc'>Internal time. Excluded from utilization, tracked as Admin Hours.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-blue'>FF: no scope</span>
                      <span class='util-ref-desc'>No matching scope entry. Flagged for finance follow-up.</span>
                    </div>
                  </div>
                  <div class='util-ref-section'>
                    <h4>RAG status</h4>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-green'>Green ≥ 70%</span>
                      <span class='util-ref-desc'>At or above target utilization.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-amber'>Amber 60–69%</span>
                      <span class='util-ref-desc'>Below target. Monitor and assess project mix.</span>
                    </div>
                    <div class='util-ref-row'>
                      <span class='util-ref-tag util-pill-red'>Red &lt; 60%</span>
                      <span class='util-ref-desc'>Significantly below target. Action required.</span>
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════════════
    # TAB 2 — Consultants
    # ═══════════════════════════════════════════════════════════════════
    with tab_consult:
        # Projection
        _period_bdays = {}
        if "date" in df.columns:
            for _per, _grp in df.groupby("period"):
                _dates = _grp["date"].dropna()
                if len(_dates) == 0: continue
                _days_in = len(pd.bdate_range(_dates.min(), _dates.max()))
                _yr3, _mo3 = _dates.min().year, _dates.min().month
                _ms3 = pd.Timestamp(_yr3, _mo3, 1)
                _me3 = pd.Timestamp(_yr3, _mo3, _cal.monthrange(_yr3, _mo3)[1])
                _total = len(pd.bdate_range(_ms3, _me3))
                _period_bdays[str(_per)] = (_days_in, _total)

        def _proj_util(row):
            per = str(row["period"])
            if per not in _period_bdays: return None
            days_in, total = _period_bdays[per]
            if days_in >= total: return None
            avail = row["avail_hrs"]
            if not avail or avail <= 0 or days_in <= 0: return None
            return (row['credit_hrs'] / days_in * total) / avail

        emp_sum["proj_full_month"] = emp_sum.apply(_proj_util, axis=1)
        _is_partial_emp = emp_sum["proj_full_month"].notna().any()

        _consult_sort_opts = {
            "Util % capacity (low → high)":  ("util_vs_capacity", True),
            "Util % capacity (high → low)":  ("util_vs_capacity", False),
            "Util % logged (low → high)":    ("util_vs_logged", True),
            "Util % logged (high → low)":    ("util_vs_logged", False),
            "Hours logged (high → low)":     ("hours_this_period", False),
            "FF overrun (high → low)":       ("ff_overrun_hrs", False),
            "Consultant (A → Z)":            ("employee", True),
        }
        cs_col, _ = st.columns([1.6, 4])
        with cs_col:
            consult_sort = st.selectbox("Sort by", list(_consult_sort_opts.keys()),
                                        index=1, key="util_consult_sort")
        _csk, _csa = _consult_sort_opts[consult_sort]
        emp_sum_sorted = emp_sum.sort_values(_csk, ascending=_csa, na_position="last").reset_index(drop=True)

        rows_html = []
        for _, r in emp_sum_sorted.iterrows():
            emp = r["employee"]; ex = bool(r["exempt"])
            avail_str = f"{r['avail_hrs']:,.1f}" if r["avail_hrs"] else "—"
            pj = r["proj_full_month"]
            pj_str = f"{pj*100:.1f}%" if pj is not None else "—"
            rows_html.append(
                f"<tr>"
                f"<td><span class='util-emp-name'>{avatar_html(emp)}{short_name(emp)}</span></td>"
                f"<td class='muted'>{r['location'] or '—'}</td>"
                f"<td class='muted'>{r['period']}</td>"
                f"<td class='num'>{avail_str}</td>"
                f"<td class='num'>{r['hours_this_period']:,.2f}</td>"
                f"<td class='num'>{r['credit_hrs']:,.2f}</td>"
                f"<td class='num'>{r['ff_overrun_hrs']:,.2f}</td>"
                f"<td class='center'>{rag_pill_html(r['util_vs_logged'], ex)}</td>"
                f"<td class='center'>{rag_pill_html(r['util_vs_capacity'], ex)}</td>"
                + (f"<td class='num muted'>{pj_str}</td>" if _is_partial_emp else "")
                + "</tr>"
            )

        proj_th = "<th class='num'>Proj full mo</th>" if _is_partial_emp else ""
        st.markdown(
            f"<div class='util-table-header'>"
            f"<span style='font-weight:600'>By consultant{' · projected to full month' if _is_partial_emp else ''}</span>"
            f"<span style='opacity:0.7'>{len(emp_sum_sorted)} of {len(emp_sum_sorted)} · sorted by {consult_sort.lower()}</span>"
            f"</div>"
            f"<div class='util-table-wrap'>"
            f"<table class='util-emp-table'>"
            f"<thead><tr>"
            f"<th>Consultant</th><th>Location</th><th>Period</th>"
            f"<th class='num'>Avail</th><th class='num'>Logged</th>"
            f"<th class='num'>Credits</th><th class='num'>Overrun</th>"
            f"<th class='center'>Util % logged</th><th class='center'>Util % cap</th>"
            f"{proj_th}"
            f"</tr></thead>"
            f"<tbody>{''.join(rows_html)}</tbody>"
            f"</table>"
            f"<div class='util-table-foot'>"
            f"<span>{'Partial period: ' + str(_bdays_in) + ' of ' + str(_bdays_total_month) + ' days. Projected at current daily run rate.' if is_partial else '&nbsp;'}</span>"
            f"<span>Last refresh: {_ago_str} · NetSuite source · {len(df):,} entries</span>"
            f"</div></div>",
            unsafe_allow_html=True
        )

    # ═══════════════════════════════════════════════════════════════════
    # TAB 3 — Projects at risk
    # ═══════════════════════════════════════════════════════════════════
    with tab_risk:
        # Build full project summary
        if _pid_col and len(df_bill) > 0:
            proj_sum = df_bill.groupby([_pid_col], as_index=False).agg(
                project=("project", "first") if "project" in df.columns else (_pid_col, "first"),
                project_type=("project_type", "first") if "project_type" in df.columns else (_pid_col, "first"),
                hours_logged=("hours", "sum"),
                credit_hrs=("credit_hrs", "sum"),
                overrun_hrs=("variance_hrs", "sum"),
            )

            # Consultants per project, ranked by hours logged (this period)
            if "employee" in df_bill.columns:
                _emp_by_proj = (df_bill.groupby([_pid_col, "employee"], as_index=False)["hours"].sum()
                                .sort_values([_pid_col, "hours"], ascending=[True, False]))
                _proj_consultants = _emp_by_proj.groupby(_pid_col)["employee"].apply(list).to_dict()
                proj_sum["consultants"] = proj_sum[_pid_col].map(_proj_consultants).apply(lambda v: v if isinstance(v, list) else [])
            else:
                proj_sum["consultants"] = [[] for _ in range(len(proj_sum))]

            # Resolve scope per project type via DEFAULT_SCOPE substring match
            def _scope_lookup(ptype):
                _matches = [(k, float(v)) for k, v in DEFAULT_SCOPE.items()
                            if k.strip().lower() in str(ptype).strip().lower()]
                return max(_matches, key=lambda x: len(x[0]))[1] if _matches else None
            proj_sum["scoped_hrs"] = proj_sum["project_type"].apply(_scope_lookup)

            # HTD per project from the consumed dict (keyed by project_id, falling back to project name)
            def _htd(r):
                pid = str(r[_pid_col]).strip().replace(".0", "")
                if pid and pid in consumed:
                    return consumed[pid]
                pname = str(r.get("project", "")).strip()
                return consumed.get(pname, None)
            proj_sum["htd_hrs"] = proj_sum.apply(_htd, axis=1)

            # Tag noscope
            _noscope_pids = set(_noscope[_pid_col].astype(str).str.strip().unique()) if _pid_col and not _noscope.empty else set()
            proj_sum["is_noscope"] = proj_sum[_pid_col].astype(str).str.strip().isin(_noscope_pids)

            # Burn % uses HTD vs scope (shows true project health, not just this period)
            def _burn(r):
                sc = r.get("scoped_hrs")
                htd = r.get("htd_hrs")
                if sc is None or sc <= 0 or htd is None: return None
                return htd / sc
            proj_sum["burn_pct"] = proj_sum.apply(_burn, axis=1)

            # Status pill
            def _status(r):
                if r["is_noscope"]:                                                         return ("blue",  "No scope")
                if r["overrun_hrs"] > 0:                                                    return ("red",   "Overrun")
                if r.get("burn_pct") is not None and r["burn_pct"] >= 0.80:                 return ("amber", f"Burn {r['burn_pct']*100:.0f}%")
                return None

            proj_sum["_status"] = proj_sum.apply(_status, axis=1)
            at_risk = proj_sum[proj_sum["_status"].notna()].copy()

            if len(at_risk) == 0:
                st.markdown(
                    "<div class='util-card' style='text-align:center;padding:32px'>"
                    "<div style='font-size:32px;margin-bottom:8px'>✓</div>"
                    "<div style='font-weight:600;margin-bottom:4px'>No projects at risk</div>"
                    "<div style='opacity:0.7;font-size:13px'>No FF overruns, no scope burn over 80%, no missing scope records.</div>"
                    "</div>", unsafe_allow_html=True)
            else:
                # Sort dropdown
                _sort_options = {
                    "Status, then overrun":  ("_status_rank", "overrun_hrs", "burn_pct"),
                    "Overrun (high → low)":  ("overrun_hrs", "burn_pct", None),
                    "Burn % (high → low)":   ("burn_pct", "overrun_hrs", None),
                    "HTD (high → low)":      ("htd_hrs", "overrun_hrs", None),
                    "Logged this period":    ("hours_logged", "overrun_hrs", None),
                    "Project (A → Z)":       ("project_asc", None, None),
                }
                sc1, _ = st.columns([1.6, 4])
                with sc1:
                    sort_choice = st.selectbox("Sort by", list(_sort_options.keys()),
                                               key="util_risk_sort", label_visibility="visible")

                # Apply sort
                at_risk["_status_rank"] = at_risk["_status"].apply(
                    lambda s: 1 if s[0] == "red" else 2 if s[0] == "amber" else 3)
                if sort_choice == "Project (A → Z)":
                    at_risk = at_risk.sort_values("project", ascending=True, na_position="last")
                else:
                    keys = [k for k in _sort_options[sort_choice] if k]
                    ascending = [True if k == "_status_rank" else False for k in keys]
                    at_risk = at_risk.sort_values(keys, ascending=ascending, na_position="last")
                at_risk = at_risk.reset_index(drop=True)

                rows = []
                for _, r in at_risk.iterrows():
                    _scls, _sl = r["_status"]
                    _sc_str = f"{r['scoped_hrs']:,.2f}" if pd.notna(r.get("scoped_hrs")) else "—"
                    _htd_str = f"{r['htd_hrs']:,.2f}" if pd.notna(r.get("htd_hrs")) else "—"
                    _burn_str = f"{r['burn_pct']*100:.0f}%" if pd.notna(r.get("burn_pct")) else "—"
                    _burn_bar = ""
                    if pd.notna(r.get("burn_pct")):
                        _bp = min(r["burn_pct"], 1.5)
                        _bcol = "#ef4444" if _bp > 1.0 else "#f59e0b" if _bp >= 0.8 else "#22c55e"
                        _burn_bar = (f"<div style='display:inline-block;width:60px;height:6px;background:rgba(128,128,128,0.2);"
                                     f"border-radius:3px;overflow:hidden;vertical-align:middle;margin-right:6px'>"
                                     f"<div style='width:{min(_bp*100/1.5, 100)}%;height:100%;background:{_bcol}'></div></div>")

                    # Consultants cell: solo = avatar+name, multi = stacked avatars + count
                    _consultants = r.get("consultants") or []
                    if len(_consultants) == 0:
                        _consult_html = "<span style='opacity:0.4'>—</span>"
                    elif len(_consultants) == 1:
                        _consult_html = f"<span class='util-emp-name'>{avatar_html(_consultants[0])}{short_name(_consultants[0])}</span>"
                    else:
                        # Stack first 3 avatars overlapping, then show count
                        _shown = _consultants[:3]
                        _avatars_html = "".join([
                            f"<span style='display:inline-block;margin-right:-8px' title='{c}'>{avatar_html(c)}</span>"
                            for c in _shown
                        ])
                        _all_names = ", ".join([short_name(c) for c in _consultants])
                        _consult_html = (f"<span class='util-emp-name' title='{_all_names}' style='gap:0'>"
                                         f"{_avatars_html}"
                                         f"<span style='margin-left:14px;font-size:12px;opacity:0.85;white-space:nowrap'>{len(_consultants)} consultants</span>"
                                         f"</span>")

                    rows.append(
                        f"<tr>"
                        f"<td>{r.get('project', '')}</td>"
                        f"<td class='muted'>{r.get('project_type', '')}</td>"
                        f"<td>{_consult_html}</td>"
                        f"<td class='num'>{_sc_str}</td>"
                        f"<td class='num'>{_htd_str}</td>"
                        f"<td class='num'>{r['hours_logged']:,.2f}</td>"
                        f"<td class='num'>{_burn_bar}<span style='vertical-align:middle'>{_burn_str}</span></td>"
                        f"<td class='num'>{r['overrun_hrs']:,.2f}</td>"
                        f"<td class='center'><span class='util-pill util-pill-{_scls}'>{_sl}</span></td>"
                        f"</tr>")

                st.markdown(
                    f"<div class='util-table-header'>"
                    f"<span style='font-weight:600'>Projects requiring attention</span>"
                    f"<span style='opacity:0.7'>{len(at_risk)} of {len(proj_sum)} billable projects · sorted by {sort_choice.lower()}</span>"
                    f"</div>"
                    f"<div class='util-table-wrap'>"
                    f"<table class='util-emp-table'>"
                    f"<thead><tr>"
                    f"<th>Project</th><th>Type</th><th>Consultant</th><th class='num'>Scoped</th>"
                    f"<th class='num'>HTD</th><th class='num'>Logged</th>"
                    f"<th class='num'>Burn</th><th class='num'>Overrun</th>"
                    f"<th class='center'>Status</th>"
                    f"</tr></thead><tbody>{''.join(rows)}</tbody></table>"
                    f"<div class='util-table-foot'>"
                    f"<span>Risk = overrun &gt; 0, burn ≥ 80% (HTD ÷ scope), or no scope record. HTD = hours-to-date all time.</span>"
                    f"<span>Last refresh: {_ago_str}</span>"
                    f"</div></div>",
                    unsafe_allow_html=True)
        else:
            st.info("No billable project data in this period.")

    # ═══════════════════════════════════════════════════════════════════
    # TAB 4 — Trend (matches the Period filter)
    # ═══════════════════════════════════════════════════════════════════
    with tab_trend:
        st.markdown('<div class="util-section-label">Trend — This Period vs Prior Equivalent</div>', unsafe_allow_html=True)
        if not _trend_visited:
            st.info("Trend computes a wider window of weekly data — click below to load.")
            if st.button("Load trend", key="util_load_trend", type="primary"):
                st.session_state._util_trend_visited = True
                st.rerun()
        elif _trend_result is None or _trend_result.get("empty") or "date" not in _trend_result.get("df", pd.DataFrame()).columns:
            st.info("Not enough data to render a trend for this period.")
        else:
            _td = _trend_result["df"].copy()
            _td["_date"] = pd.to_datetime(_td["date"], errors="coerce")
            _td = _td.dropna(subset=["_date"])
            _td["_week"] = _td["_date"].dt.to_period("W").apply(lambda p: p.start_time.date())

            # Weekly aggregates
            weekly = _td.groupby("_week", as_index=False).agg(
                hours=("hours", "sum"),
                credit_hrs=("credit_hrs", "sum"),
                overrun_hrs=("variance_hrs", "sum"),
            )
            # Billable project count per week (distinct project IDs, billable rows only)
            _td_pid = _proj_id_col(_td)
            if _td_pid:
                _bmask_td = _billable_mask(_td)
                _proj_weekly = _td[_bmask_td].groupby("_week", as_index=False).agg(
                    n_projects=(_td_pid, lambda s: s.astype(str).str.strip().nunique())
                )
                weekly = weekly.merge(_proj_weekly, on="_week", how="left")
                weekly["n_projects"] = weekly["n_projects"].fillna(0).astype(int)
            else:
                weekly["n_projects"] = 0
            # noscope hrs computed separately and merged in
            if "credit_tag" in _td.columns:
                _ns_weekly = _td[_td["credit_tag"] == "UNCONFIGURED"].groupby("_week", as_index=False).agg(noscope_hrs=("hours", "sum"))
                weekly = weekly.merge(_ns_weekly, on="_week", how="left")
                weekly["noscope_hrs"] = weekly["noscope_hrs"].fillna(0)
            else:
                weekly["noscope_hrs"] = 0
            weekly["credit_pct"] = weekly.apply(lambda r: r["credit_hrs"] / r["hours"] if r["hours"] > 0 else 0, axis=1)
            weekly = weekly.sort_values("_week").reset_index(drop=True)

            if len(weekly) < 2:
                st.info("Need at least 2 weeks of data to show a trend. Widen the period filter.")
            else:
                # Three side-by-side trend cards
                t1, t2, t3 = st.columns(3)

                def _mini_chart(values, target=None, fmt="pct", color="#3b82f6"):
                    """Render an inline SVG bar chart for trend cards."""
                    if not values: return ""
                    n = len(values)
                    max_v = max([v for v in values if v is not None] + [0.01])
                    bar_w = max(40 / n, 4)
                    gap = bar_w * 0.2
                    chart_w = 100
                    chart_h = 40
                    bar_w_actual = (chart_w - gap * (n - 1)) / n
                    bars = []
                    for i, v in enumerate(values):
                        if v is None or max_v == 0:
                            h = 0
                        else:
                            h = (v / max_v) * chart_h
                        x = i * (bar_w_actual + gap)
                        y = chart_h - h
                        bars.append(f"<rect x='{x:.1f}' y='{y:.1f}' width='{bar_w_actual:.1f}' height='{h:.1f}' fill='{color}' opacity='0.8' rx='1' />")
                    target_line = ""
                    if target is not None and max_v > 0:
                        ty = chart_h - (target / max_v) * chart_h
                        if 0 <= ty <= chart_h:
                            target_line = f"<line x1='0' y1='{ty:.1f}' x2='{chart_w}' y2='{ty:.1f}' stroke='rgba(128,128,128,0.5)' stroke-dasharray='2 2' />"
                    return f"<svg viewBox='0 0 {chart_w} {chart_h}' style='width:100%;height:50px;display:block'>{target_line}{''.join(bars)}</svg>"

                _credit_pcts  = weekly["credit_pct"].tolist()
                _overrun_hrs  = weekly["overrun_hrs"].tolist()
                _noscope_hrs_l= weekly["noscope_hrs"].tolist()

                # ─── Period totals (this period) ───
                _curr_total_hrs        = weekly["hours"].sum()
                _curr_total_credits    = weekly["credit_hrs"].sum()
                _curr_credit_pct       = _curr_total_credits / _curr_total_hrs if _curr_total_hrs > 0 else 0
                _curr_total_overrun    = sum(_overrun_hrs)
                _curr_total_noscope    = sum(_noscope_hrs_l)

                # ─── Prior equivalent period totals ───
                _period_days = (period_end - period_start).days + 1
                _prior_p_end = (pd.Timestamp(period_start) - pd.Timedelta(days=1)).date()
                _prior_p_start = (pd.Timestamp(_prior_p_end) - pd.Timedelta(days=_period_days - 1)).date()

                try:
                    _prior_p_result = _run_utilization_engine(_trend_df_raw, _prior_p_start, _prior_p_end, _df_drs)
                except Exception:
                    _prior_p_result = {"empty": True}

                _prior_credit_pct = None
                _prior_total_overrun = None
                _prior_total_noscope = None
                if not _prior_p_result.get("empty") and not _prior_p_result.get("df", pd.DataFrame()).empty:
                    _pdf2 = _prior_p_result["df"]
                    _p_hrs    = _pdf2["hours"].sum() if "hours" in _pdf2.columns else 0
                    _p_credit = _pdf2["credit_hrs"].sum() if "credit_hrs" in _pdf2.columns else 0
                    _prior_credit_pct = _p_credit / _p_hrs if _p_hrs > 0 else 0
                    _prior_total_overrun = _pdf2[_pdf2["credit_tag"].isin(["OVERRUN", "PARTIAL"])]["variance_hrs"].sum() if "variance_hrs" in _pdf2.columns and "credit_tag" in _pdf2.columns else 0
                    _prior_total_noscope = _pdf2[_pdf2["credit_tag"] == "UNCONFIGURED"]["hours"].sum() if "credit_tag" in _pdf2.columns else 0

                # ─── Deltas (current − prior) ───
                _delta_credit  = (_curr_credit_pct - _prior_credit_pct) if _prior_credit_pct is not None else None
                _delta_overrun = (_curr_total_overrun - _prior_total_overrun) if _prior_total_overrun is not None else None
                _delta_noscope = (_curr_total_noscope - _prior_total_noscope) if _prior_total_noscope is not None else None

                _prior_label = f"vs {_prior_p_start.strftime('%-d %b')}–{_prior_p_end.strftime('%-d %b')}"

                # Headline color (use period total for credit, since that's what RAG measures)
                _headline_color = "#22c55e" if _curr_credit_pct >= 0.70 else "#f59e0b" if _curr_credit_pct >= 0.60 else "#ef4444"

                def _delta_html(delta, unit, lower_is_better=False, threshold=0.005):
                    """Render arrow + magnitude. Higher = better unless lower_is_better."""
                    if delta is None:
                        return f"<div style='font-size:12px;opacity:0.6'>no prior data</div>"
                    if abs(delta) < threshold:
                        return f"<div style='font-size:12px;opacity:0.6'>→ flat</div>"
                    is_up = delta > 0
                    is_good = (is_up and not lower_is_better) or (not is_up and lower_is_better)
                    color = "#15803d" if is_good else "#b91c1c"
                    arrow = "↑" if is_up else "↓"
                    return f"<div style='font-size:12px;color:{color}'>{arrow} {abs(delta):.1f}{unit}</div>"

                with t1:
                    _curr_credit_pp = _curr_credit_pct * 100
                    _delta_credit_pp = _delta_credit * 100 if _delta_credit is not None else None
                    st.markdown(
                        f"<div class='util-card'>"
                        f"<div style='font-size:11px;opacity:0.65;margin-bottom:4px'>Credit % · period total</div>"
                        f"<div style='display:flex;align-items:baseline;gap:8px;margin-bottom:8px'>"
                        f"<div style='font-size:22px;font-weight:600;color:{_headline_color}'>{_curr_credit_pp:.1f}%</div>"
                        f"{_delta_html(_delta_credit_pp, 'pp', lower_is_better=False, threshold=0.5)}"
                        f"</div>"
                        f"{_mini_chart(_credit_pcts, target=0.70, color='#3b82f6')}"
                        f"<div style='font-size:11px;opacity:0.6;margin-top:4px'>{len(weekly)} weeks · target ≥ 70% · {_prior_label}</div>"
                        f"</div>", unsafe_allow_html=True)

                with t2:
                    st.markdown(
                        f"<div class='util-card'>"
                        f"<div style='font-size:11px;opacity:0.65;margin-bottom:4px'>FF overrun · period total</div>"
                        f"<div style='display:flex;align-items:baseline;gap:8px;margin-bottom:8px'>"
                        f"<div style='font-size:22px;font-weight:600'>{fmt_hrs(_curr_total_overrun)} h</div>"
                        f"{_delta_html(_delta_overrun, 'h', lower_is_better=True, threshold=0.5)}"
                        f"</div>"
                        f"{_mini_chart(_overrun_hrs, color='#f59e0b')}"
                        f"<div style='font-size:11px;opacity:0.6;margin-top:4px'>{len(weekly)} weeks · {_prior_label}</div>"
                        f"</div>", unsafe_allow_html=True)

                with t3:
                    st.markdown(
                        f"<div class='util-card'>"
                        f"<div style='font-size:11px;opacity:0.65;margin-bottom:4px'>FF: no scope · period total</div>"
                        f"<div style='display:flex;align-items:baseline;gap:8px;margin-bottom:8px'>"
                        f"<div style='font-size:22px;font-weight:600'>{fmt_hrs(_curr_total_noscope)} h</div>"
                        f"{_delta_html(_delta_noscope, 'h', lower_is_better=True, threshold=0.5)}"
                        f"</div>"
                        f"{_mini_chart(_noscope_hrs_l, color='#3b82f6')}"
                        f"<div style='font-size:11px;opacity:0.6;margin-top:4px'>{len(weekly)} weeks · {_prior_label}</div>"
                        f"</div>", unsafe_allow_html=True)

                # Weekly detail table
                st.markdown("<div style='margin-top:14px;font-size:11px;opacity:0.6;text-transform:uppercase;letter-spacing:0.6px;margin-bottom:6px'>Weekly breakdown</div>", unsafe_allow_html=True)
                _weekly_disp = weekly.copy()
                _weekly_disp["Week"] = _weekly_disp["_week"].apply(lambda d: d.strftime("%Y-W%V") if hasattr(d, "strftime") else str(d))
                _weekly_disp["credit_pct"] = _weekly_disp["credit_pct"] * 100  # Show as percentage points
                _weekly_disp = _weekly_disp[["Week", "n_projects", "hours", "credit_hrs", "credit_pct", "overrun_hrs", "noscope_hrs"]]
                st.dataframe(
                    _weekly_disp, use_container_width=True, hide_index=True,
                    column_config={
                        "Week":         st.column_config.TextColumn("Week", pinned=True),
                        "n_projects":   st.column_config.NumberColumn("Projects", format="%d", help="Distinct billable projects with hours logged this week"),
                        "hours":        st.column_config.NumberColumn("Logged", format="%.2f"),
                        "credit_hrs":   st.column_config.NumberColumn("Credits", format="%.2f"),
                        "credit_pct":   st.column_config.ProgressColumn("Credit %", format="%.1f%%", min_value=0, max_value=100, help="Credits ÷ hours logged"),
                        "overrun_hrs":  st.column_config.NumberColumn("FF overrun", format="%.2f"),
                        "noscope_hrs":  st.column_config.NumberColumn("No scope", format="%.2f"),
                    },
                )

    # ═══════════════════════════════════════════════════════════════════
    # TAB 5 — Task analysis (Billable / Non-billable toggle)
    # ═══════════════════════════════════════════════════════════════════
    with tab_task:
        st.markdown('<div class="util-section-label">Task Analysis — This Period vs Prior</div>', unsafe_allow_html=True)
        if not _task_visited:
            st.info("Task analysis computes a wider window plus prior-period comparison — click below to load.")
            if st.button("Load task analysis", key="util_load_task", type="primary"):
                st.session_state._util_task_visited = True
                st.rerun()
        else:
            # Mode selector
            _bill_hrs_total = df_bill["hours"].sum() if not df_bill.empty else 0
            _nonbill_hrs_total = df[df["billing_type"].fillna("").str.lower() == "internal"]["hours"].sum() if "billing_type" in df.columns else 0

            ts_col1, ts_col2, ts_col3 = st.columns([1.2, 1.2, 2.5])
            with ts_col1:
                task_mode = st.radio(
                    "Mode",
                    options=["Billable", "Non-billable"],
                    horizontal=True, label_visibility="collapsed",
                    key="util_task_mode",
                    format_func=lambda x: f"Billable · {fmt_hrs(_bill_hrs_total)}h" if x == "Billable" else f"Non-billable · {fmt_hrs(_nonbill_hrs_total)}h",
                )
            with ts_col2:
                compare_options = {
                    "Prior equivalent period": "prior_eq",
                    "Prior 4 weeks (per-week avg)": "prior_4w",
                    "Prior 8 weeks (per-week avg)": "prior_8w",
                }
                compare_label = st.selectbox("Compare vs", list(compare_options.keys()), key="util_task_compare", label_visibility="collapsed")
                compare_mode = compare_options[compare_label]
            with ts_col3:
                _norm_note = " · normalized to /week" if compare_mode != "prior_eq" else ""
                st.markdown(f"<div style='padding:8px 0;font-size:11px;opacity:0.6;text-align:right'>vs <b>{compare_label}</b>{_norm_note}</div>", unsafe_allow_html=True)

            # Determine task column
            if task_mode == "Billable":
                _task_df = df_bill.copy()
                _task_col = "ff_task" if "ff_task" in _task_df.columns and _task_df["ff_task"].notna().any() else ("task" if "task" in _task_df.columns else None)
            else:
                _task_df = df[df["billing_type"].fillna("").str.lower() == "internal"].copy() if "billing_type" in df.columns else df.iloc[0:0].copy()
                _task_col = "task" if "task" in _task_df.columns else None

            if not _task_col or len(_task_df) == 0:
                st.info(f"No {task_mode.lower()} time entries in this period.")
            else:
                # ─── Top-left: share by task ───
                share = _task_df.groupby(_task_col, as_index=False).agg(hours=("hours", "sum"))
                share = share[share["hours"] > 0].sort_values("hours", ascending=False).reset_index(drop=True)
                total_share = share["hours"].sum()
                # Collapse < 5% into "Other"
                _big = share[share["hours"] / total_share >= 0.05].copy() if total_share > 0 else share.copy()
                _small = share[share["hours"] / total_share < 0.05].copy() if total_share > 0 else share.iloc[0:0]
                if len(_small) > 0:
                    _other_row = pd.DataFrame([{_task_col: f"Other ({len(_small)} tasks)", "hours": _small["hours"].sum()}])
                    _big = pd.concat([_big, _other_row], ignore_index=True)
                _big["pct"] = _big["hours"] / total_share if total_share > 0 else 0

                share_rows = []
                for i, r in _big.iterrows():
                    _opacity = max(0.4, 1.0 - (i * 0.1))
                    _muted = " style='opacity:0.65'" if str(r[_task_col]).startswith("Other") else ""
                    share_rows.append(
                        f"<div class='util-share-row'>"
                        f"<div class='util-share-head'{_muted}>"
                        f"<span>{r[_task_col]}</span>"
                        f"<span style='opacity:0.7;font-variant-numeric:tabular-nums'>{r['hours']:,.2f} h · {r['pct']*100:.1f}%</span>"
                        f"</div>"
                        f"<div class='util-share-bar-bg'>"
                        f"<div class='util-share-bar-fg' style='width:{min(r['pct']*100, 100):.1f}%;opacity:{_opacity}'></div>"
                        f"</div></div>")

                # ─── Top-right: movers vs prior period ───
                if compare_mode == "prior_eq":
                    _delta = (period_end - period_start).days + 1
                    _prior_end = (pd.Timestamp(period_start) - pd.Timedelta(days=1)).date()
                    _prior_start = (pd.Timestamp(_prior_end) - pd.Timedelta(days=_delta - 1)).date()
                elif compare_mode == "prior_4w":
                    _prior_end = (pd.Timestamp(period_start) - pd.Timedelta(days=1)).date()
                    _prior_start = (pd.Timestamp(_prior_end) - pd.Timedelta(weeks=4)).date()
                else:  # prior_8w
                    _prior_end = (pd.Timestamp(period_start) - pd.Timedelta(days=1)).date()
                    _prior_start = (pd.Timestamp(_prior_end) - pd.Timedelta(weeks=8)).date()

                _curr_weeks  = max(((period_end - period_start).days + 1) / 7, 1/7)
                _prior_weeks = max(((_prior_end - _prior_start).days + 1) / 7, 1/7)

                # Use the wider df (df_raw before period filter) for comparison
                try:
                    _prior_result = _run_utilization_engine(_trend_df_raw, _prior_start, _prior_end, _df_drs)
                except Exception:
                    _prior_result = {"empty": True}

                movers_html = ""
                if _prior_result.get("empty") or _prior_result.get("df", pd.DataFrame()).empty:
                    movers_html = "<div style='opacity:0.6;font-size:12px;padding:20px 0;text-align:center'>No data for the comparison window.</div>"
                else:
                    _pdf = _prior_result["df"]
                    if task_mode == "Billable":
                        _pdf_filt = _pdf[_billable_mask(_pdf)]
                        _pcol = "ff_task" if "ff_task" in _pdf_filt.columns and _pdf_filt["ff_task"].notna().any() else ("task" if "task" in _pdf_filt.columns else None)
                    else:
                        _pdf_filt = _pdf[_pdf["billing_type"].fillna("").str.lower() == "internal"] if "billing_type" in _pdf.columns else _pdf.iloc[0:0]
                        _pcol = "task" if "task" in _pdf_filt.columns else None

                    if _pcol and len(_pdf_filt) > 0:
                        prior_share = _pdf_filt.groupby(_pcol, as_index=False).agg(prior_hours=("hours", "sum"))
                        merged = share.merge(prior_share, left_on=_task_col, right_on=_pcol, how="outer").fillna(0)
                        if _task_col != _pcol and _pcol in merged.columns:
                            merged[_task_col] = merged[_task_col].where(merged[_task_col] != 0, merged[_pcol])

                        # Per-week normalization for rolling options
                        _normalize = (compare_mode != "prior_eq")
                        if _normalize:
                            merged["curr_per_wk"]  = merged["hours"] / _curr_weeks
                            merged["prior_per_wk"] = merged["prior_hours"] / _prior_weeks
                            merged["delta"] = merged["curr_per_wk"] - merged["prior_per_wk"]
                            _curr_disp_col, _prior_disp_col = "curr_per_wk", "prior_per_wk"
                            _unit = "h/wk"
                        else:
                            merged["delta"] = merged["hours"] - merged["prior_hours"]
                            _curr_disp_col, _prior_disp_col = "hours", "prior_hours"
                            _unit = "h"

                        merged["abs_delta"] = merged["delta"].abs()
                        movers = merged.sort_values("abs_delta", ascending=False).head(5)

                        _rows = []
                        for _, r in movers.iterrows():
                            _d = r["delta"]
                            _cls = "util-pill-green" if _d > 0 else "util-pill-red" if _d < 0 else "util-pill-grey"
                            _sign = "+" if _d > 0 else ""
                            _rows.append(
                                f"<div class='util-mover-row'>"
                                f"<span>{r[_task_col]}</span>"
                                f"<span class='util-mover-vals'>"
                                f"<span class='util-mover-from'>{r[_prior_disp_col]:.2f} → {r[_curr_disp_col]:.2f} {_unit}</span>"
                                f"<span class='util-pill {_cls}'>{_sign}{_d:.2f}</span>"
                                f"</span></div>"
                            )

                        # Auto-interpretation
                        _interp = ""
                        if not movers.empty:
                            _top = movers.iloc[0]
                            _delta_unit = "h/wk" if _normalize else "h"
                            if _top["delta"] > 0:
                                _interp = f"{_top[_task_col]} growing fastest (+{_top['delta']:.1f} {_delta_unit})."
                            else:
                                _interp = f"{_top[_task_col]} declining most ({_top['delta']:.1f} {_delta_unit})."

                        movers_html = "".join(_rows) + (f"<div style='font-size:11px;opacity:0.6;margin-top:10px;padding-top:8px;border-top:1px solid rgba(128,128,128,0.15)'>{_interp}</div>" if _interp else "")
                    else:
                        movers_html = "<div style='opacity:0.6;font-size:12px;padding:20px 0;text-align:center'>No comparable task data.</div>"

                # Render top row
                tr1, tr2 = st.columns([1.4, 1])
                with tr1:
                    st.markdown(
                        f"<div class='util-card'>"
                        f"<div style='font-size:12px;font-weight:600;margin-bottom:4px'>{task_mode} hours by task · this period</div>"
                        f"<div style='font-size:11px;opacity:0.65;margin-bottom:12px'>{total_share:,.2f} hrs across {len(share)} task type{'s' if len(share) != 1 else ''}</div>"
                        f"{''.join(share_rows)}"
                        f"</div>", unsafe_allow_html=True)
                with tr2:
                    if compare_mode != "prior_eq":
                        _subtitle = (f"Current ÷ {_curr_weeks:.1f}w  vs  prior ÷ {_prior_weeks:.1f}w "
                                     f"({_prior_start.strftime('%-d %b')} → {_prior_end.strftime('%-d %b')})")
                    else:
                        _subtitle = f"{_prior_start.strftime('%-d %b')} → {_prior_end.strftime('%-d %b')}"
                    st.markdown(
                        f"<div class='util-card'>"
                        f"<div style='font-size:12px;font-weight:600;margin-bottom:4px'>Movers · vs {compare_label.lower()}</div>"
                        f"<div style='font-size:11px;opacity:0.65;margin-bottom:8px'>{_subtitle}</div>"
                        f"{movers_html}"
                        f"</div>", unsafe_allow_html=True)

                # ─── Bottom: stacked weekly trend ───
                if not _trend_result.get("empty"):
                    _td2 = _trend_result["df"].copy()
                    if task_mode == "Billable":
                        _td2 = _td2[_billable_mask(_td2)]
                    else:
                        _td2 = _td2[_td2["billing_type"].fillna("").str.lower() == "internal"] if "billing_type" in _td2.columns else _td2.iloc[0:0]
                    _td2["_date"] = pd.to_datetime(_td2["date"], errors="coerce")
                    _td2 = _td2.dropna(subset=["_date"])
                    _td2["_week"] = _td2["_date"].dt.to_period("W").apply(lambda p: p.start_time.date())

                    _tcol_trend = _task_col if _task_col in _td2.columns else None
                    if _tcol_trend and len(_td2) > 0:
                        # Top 4 tasks by total hours; rest = Other
                        _totals = _td2.groupby(_tcol_trend)["hours"].sum().sort_values(ascending=False)
                        _top4 = _totals.head(4).index.tolist()
                        _td2["_task_grp"] = _td2[_tcol_trend].apply(lambda v: v if v in _top4 else "Other")
                        _stack = _td2.groupby(["_week", "_task_grp"], as_index=False).agg(hours=("hours", "sum"))
                        _pivot = _stack.pivot(index="_week", columns="_task_grp", values="hours").fillna(0).sort_index()
                        _categories = _top4 + (["Other"] if "Other" in _pivot.columns else [])
                        _categories = [c for c in _categories if c in _pivot.columns]

                        if not _pivot.empty and len(_categories) > 0:
                            _palette = ["#1d4ed8", "#3b82f6", "#60a5fa", "#93c5fd", "#bfdbfe"]
                            _max_total = _pivot[_categories].sum(axis=1).max()
                            _w = 700; _h = 200
                            _bar_area_w = _w - 50
                            _n = len(_pivot.index)
                            _bw = (_bar_area_w / _n) * 0.7
                            _gap = (_bar_area_w / _n) * 0.3

                            bars_svg = []
                            labels_svg = []
                            for i, (wk, row) in enumerate(_pivot.iterrows()):
                                x = 50 + i * (_bw + _gap)
                                _y_cursor = _h - 20
                                for j, cat in enumerate(_categories):
                                    v = row.get(cat, 0)
                                    if v <= 0: continue
                                    bh = (v / _max_total) * (_h - 40) if _max_total > 0 else 0
                                    _y_cursor -= bh
                                    bars_svg.append(f"<rect x='{x:.1f}' y='{_y_cursor:.1f}' width='{_bw:.1f}' height='{bh:.1f}' fill='{_palette[j % len(_palette)]}' />")
                                wk_label = wk.strftime("%-d %b") if hasattr(wk, "strftime") else str(wk)
                                labels_svg.append(f"<text x='{x + _bw/2:.1f}' y='{_h - 5}' fill='currentColor' opacity='0.6' font-size='9' text-anchor='middle'>{wk_label}</text>")

                            # y-axis grid
                            grid = ""
                            if _max_total > 0:
                                for frac in [0.25, 0.5, 0.75, 1.0]:
                                    gy = (_h - 20) - frac * (_h - 40)
                                    gv = frac * _max_total
                                    grid += f"<line x1='50' y1='{gy:.1f}' x2='{_w}' y2='{gy:.1f}' stroke='currentColor' opacity='0.1' stroke-dasharray='2 4' />"
                                    grid += f"<text x='45' y='{gy + 3:.1f}' fill='currentColor' opacity='0.5' font-size='9' text-anchor='end'>{gv:.0f}h</text>"

                            legend_html = "".join([
                                f"<span style='display:inline-flex;align-items:center;gap:5px;margin-right:14px;font-size:11px'>"
                                f"<span style='display:inline-block;width:10px;height:10px;background:{_palette[j % len(_palette)]};border-radius:2px'></span>"
                                f"{cat}</span>"
                                for j, cat in enumerate(_categories)
                            ])

                            st.markdown(
                                f"<div class='util-card' style='margin-top:12px'>"
                                f"<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:4px'>"
                                f"<span style='font-size:12px;font-weight:600'>{task_mode} hours by task · weekly</span>"
                                f"<span style='font-size:11px;opacity:0.6'>stacked · top 4 + other</span>"
                                f"</div>"
                                f"<svg viewBox='0 0 {_w} {_h}' style='width:100%;height:200px'>{grid}{''.join(bars_svg)}{''.join(labels_svg)}</svg>"
                                f"<div style='margin-top:8px;color:currentColor'>{legend_html}</div>"
                                f"</div>", unsafe_allow_html=True)

        # ═══════════════════════════════════════════════════════════════════
        # TAB 6 — Detail
    # ═══════════════════════════════════════════════════════════════════
    with tab_detail:
        display_cols_t5 = ["employee","region","project","project_type","billing_type",
                            "hours_to_date","date","hours","credit_hrs","variance_hrs",
                            "previous_htd","credit_tag","notes"]
        existing_t5 = [c for c in display_cols_t5 if c in df.columns]

        _glyph_map = {"CREDITED":"🟢","OVERRUN":"🔴","PARTIAL":"🟡","NON-BILLABLE":"⚫","UNCONFIGURED":"⚪","SKIPPED":"·"}
        _detail_view = df[existing_t5].head(500).copy()
        if "credit_tag" in _detail_view.columns:
            _detail_view["credit_tag"] = _detail_view["credit_tag"].map(lambda t: f"{_glyph_map.get(t,'')} {t}" if pd.notna(t) else "")

        st.dataframe(
            _detail_view, use_container_width=True, hide_index=True,
            column_config={
                "employee":     st.column_config.TextColumn("Employee", pinned=True),
                "region":       st.column_config.TextColumn("Region"),
                "project":      st.column_config.TextColumn("Project"),
                "project_type": st.column_config.TextColumn("Type"),
                "billing_type": st.column_config.TextColumn("Billing"),
                "hours_to_date":st.column_config.NumberColumn("HTD", format="%.2f"),
                "date":         st.column_config.DateColumn("Date"),
                "hours":        st.column_config.NumberColumn("Hours", format="%.2f"),
                "credit_hrs":   st.column_config.NumberColumn("Credit hrs", format="%.2f"),
                "variance_hrs": st.column_config.NumberColumn("Variance", format="%.2f"),
                "previous_htd": st.column_config.NumberColumn("Prev HTD", format="%.2f"),
                "credit_tag":   st.column_config.TextColumn("Credit tag", help="🟢 Credited · 🟡 Partial · 🔴 Overrun · ⚫ Non-billable · ⚪ FF no scope"),
                "notes":        st.column_config.TextColumn("Notes"),
            },
        )
        if len(df) > 500:
            st.caption(f"Showing first 500 of {len(df):,} rows. Full data in Excel download.")

    # ─────────────────────────────────────────────────────
    # Tableau export — secondary, on-demand build
    # ─────────────────────────────────────────────────────
    st.divider()
    if _is_mgr_u:
        _tab_cache_key = ("tab", _ns_signature(_ns_from_session),
                          str(period_start), str(period_end),
                          _va_name_u, _va_region_u)
        st.session_state.setdefault("_util_tableau_cache", {})
        tableau_filename = f"utilization_tableau_{timestamp}.xlsx"

        # Build runs OUTSIDE the expander so any error/exception renders at page level
        # where the user can see it (rather than buried in a collapsed expander body).
        if st.session_state.get("_util_tab_prep_requested") == _tab_cache_key:
            st.session_state["_util_tab_prep_requested"] = None
            try:
                with st.spinner("Building Tableau export..."):
                    st.session_state._util_tableau_cache[_tab_cache_key] = build_tableau_excel(df, DEFAULT_SCOPE, consumed)
            except Exception as _tab_err:
                st.error(f"Tableau build failed: {_tab_err}")
                st.exception(_tab_err)

        _tab_buf = st.session_state._util_tableau_cache.get(_tab_cache_key)

        # Auto-open the expander after a click so the user sees what happened
        _expand_now = st.session_state.pop("_util_tab_expand", False)
        with st.expander("Tableau export (advanced)", expanded=_expand_now or _tab_buf is not None):
            if _tab_buf is not None:
                st.download_button(
                    label="⬇ Download Tableau Export",
                    data=_tab_buf, file_name=tableau_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="util_dl_tableau",
                )
            else:
                if st.button("Build Tableau export", key="util_prep_tableau",
                             help="Generate the 3-sheet Tableau workbook"):
                    st.session_state["_util_tab_prep_requested"] = _tab_cache_key
                    st.session_state["_util_tab_expand"] = True
                    st.rerun()
            st.caption("3 flat sheets: fact_utilization · fact_processed_time_entries · fact_ff_overrun_by_type")
def build_tableau_excel(df, scope_map, consumed):
    import io as _io
    from openpyxl import Workbook as _WB

    wb = _WB()
    wb.remove(wb.active)

    def _flat_sheet(wb, name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for r in rows: ws.append(r)
        return ws

    def _pct(num, denom):
        try: return round(num / denom, 6) if denom and denom > 0 else None
        except Exception: return None

    _skipped = df["credit_tag"] == "SKIPPED" if "credit_tag" in df.columns else df.index == -1
    _df = df[~_skipped].copy()
    _admin = (_df[_df["billing_type"].str.lower() == "internal"]
              .groupby(["employee","period"])["hours"].sum()
              .reset_index().rename(columns={"hours":"admin_hrs"}))
    _sum = (_df.groupby(["employee","period"], as_index=False)
            .agg(hours_logged=("hours","sum"), credit_hrs=("credit_hrs","sum"), ff_overrun_hrs=("variance_hrs","sum"))
            .sort_values(["employee","period"]))
    _sum = _sum.merge(_admin, on=["employee","period"], how="left")
    _sum["admin_hrs"] = _sum["admin_hrs"].fillna(0)

    fu_headers = ["employee","location","ps_region","role","period","hours_capacity","hours_logged",
                  "credit_hrs","admin_hrs","ff_overrun_hrs","util_pct_vs_logged","util_pct_vs_capacity",
                  "project_util_pct","gap_pts","util_rag"]
    fu_rows = []
    for _, row in _sum.iterrows():
        emp = row["employee"]; period = row["period"]
        loc = df[df["employee"] == emp]["region"].iloc[0] if len(df[df["employee"] == emp]) > 0 else ""
        ps_reg = df[df["employee"] == emp]["ps_region"].iloc[0] if "ps_region" in df.columns and len(df[df["employee"] == emp]) > 0 else ""
        info = EMPLOYEE_ROLES.get(emp, {}); role = info.get("role", "Consultant")
        avail = get_avail_hours(loc, period) if loc else None
        u_log = _pct(row["credit_hrs"], row["hours_logged"])
        u_cap = _pct(row["credit_hrs"], avail)
        u_proj = _pct(row["credit_hrs"] + row["ff_overrun_hrs"], avail)
        gap = round(u_proj - u_cap, 6) if u_proj is not None and u_cap is not None else None
        rag = "Unknown" if u_cap is None else "Green" if u_cap >= 0.70 else "Amber" if u_cap >= 0.60 else "Red"
        fu_rows.append([emp, loc, ps_reg, role, period, avail,
                        round(row["hours_logged"], 2), round(row["credit_hrs"], 2),
                        round(row["admin_hrs"], 2), round(row["ff_overrun_hrs"], 2),
                        u_log, u_cap, u_proj, gap, rag])
    _flat_sheet(wb, "fact_utilization", fu_headers, fu_rows)

    ft_headers = ["employee","location","ps_region","customer_region","project_manager","project",
                  "project_type","billing_type","hrs_to_date","date","hours_logged","approval",
                  "task_case","non_billable","credit_hrs","variance_hrs","previous_hrs_to_date",
                  "credit_tag","period","notes","project_phase","start_date","days_active",
                  "scoped_hrs","variance_flag"]

    def _tbl_scoped(ptype):
        _m = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in str(ptype).strip().lower()]
        return round(max(_m, key=lambda x: len(x[0]))[1], 2) if _m else None

    def _tbl_vflag(tag, scoped):
        t = str(tag).strip().upper()
        if t in ("SKIPPED", "NON-BILLABLE", "CREDITED"): return "N/A"
        if t == "UNCONFIGURED": return "No Scope Set"
        if t in ("OVERRUN", "PARTIAL"): return "True Overrun" if scoped and scoped > 0 else "No Scope Set"
        return "Within Budget"

    ft_rows = []
    for _, row in df.iterrows():
        def _g(col, default=""):
            v = row.get(col, default)
            if v is None or (hasattr(v, "__class__") and v.__class__.__name__ == "float" and str(v) == "nan"): return default
            return v
        _tag = str(_g("credit_tag")).strip(); _scoped = _tbl_scoped(_g("project_type"))
        ft_rows.append([
            _g("employee"), _g("region"), _g("ps_region",""), _g("customer_region"), _g("project_manager"),
            _g("project"), _g("project_type"), _g("billing_type"),
            round(float(_g("hours_to_date", 0) or 0), 2),
            str(_g("date"))[:10], round(float(_g("hours", 0) or 0), 2),
            _g("approval"), _g("task", _g("case_task_event", "")),
            1 if str(_g("non_billable","")).lower() in ("true","yes","1","x") else 0,
            round(float(_g("credit_hrs", 0) or 0), 2), round(float(_g("variance_hrs", 0) or 0), 2),
            round(float(_g("previous_htd", 0) or 0), 2),
            _g("credit_tag"), _g("period"), _g("notes",""),
            _g("project_phase"), str(_g("start_date_display", _g("start_date", "")))[:10],
            _g("days_active",""), _scoped if _scoped is not None else "", _tbl_vflag(_tag, _scoped),
        ])
    _flat_sheet(wb, "fact_processed_time_entries", ft_headers, ft_rows)

    fot_headers = ["project_type","scoped_hrs","total_projects","projects_over_budget",
                   "pct_over_budget","total_overrun_hrs","avg_scoped_hrs","avg_actual_hrs","avg_over_under_hrs"]
    _fot_ff = _df[_df.get("billing_type", pd.Series(dtype="object")).str.lower() == "fixed fee"].copy() \
        if "billing_type" in _df.columns else _df.copy()

    def _fot_scope(ptype):
        _m = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in str(ptype).strip().lower()]
        return max(_m, key=lambda x: len(x[0]))[1] if _m else None

    _fot_proj = _fot_ff.groupby(["project","project_type"], as_index=False).agg(
        hours_total=("hours","sum"), overrun_hrs=("variance_hrs","sum"))
    _fot_proj["scoped_hrs"] = _fot_proj["project_type"].apply(_fot_scope)
    _fot_proj = _fot_proj[_fot_proj["scoped_hrs"].notna()].copy()
    _fot_proj["over_under"] = _fot_proj["hours_total"].astype(float) - _fot_proj["scoped_hrs"].astype(float)
    _fot_proj["is_over"] = (_fot_proj["over_under"] > 0).astype(int)
    _fot_type = _fot_proj.groupby("project_type", as_index=False).agg(
        scoped_hrs=("scoped_hrs","mean"), total_projects=("project","count"),
        projects_over_budget=("is_over","sum"), total_overrun_hrs=("overrun_hrs","sum"),
        avg_scoped_hrs=("scoped_hrs","mean"), avg_actual_hrs=("hours_total","mean"),
        avg_over_under_hrs=("over_under","mean"),
    ).sort_values("total_overrun_hrs", ascending=False)
    _fot_type["pct_over_budget"] = _fot_type.apply(
        lambda r: round(r["projects_over_budget"] / r["total_projects"], 6) if r["total_projects"] > 0 else None, axis=1)
    fot_rows = []
    for _, r in _fot_type.iterrows():
        fot_rows.append([r["project_type"], round(float(r["scoped_hrs"]),2), int(r["total_projects"]),
                         int(r["projects_over_budget"]), r["pct_over_budget"],
                         round(float(r["total_overrun_hrs"]),2), round(float(r["avg_scoped_hrs"]),2),
                         round(float(r["avg_actual_hrs"]),2), round(float(r["avg_over_under_hrs"]),2)])
    _flat_sheet(wb, "fact_ff_overrun_by_type", fot_headers, fot_rows)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == "__main__":
    main()
